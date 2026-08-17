import streamlit as st
import re
try:
    import matplotlib.pyplot as plt
    MATPLOT = True
except ImportError:
    MATPLOT = False
import pandas as pd
import os
import shutil
import time
import io
import json
import hashlib
import requests
from datetime import datetime, timedelta
from PIL import Image
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CACHE DE DADOS — evita recarregar arquivo JSON a cada rerun
# ============================================================
@st.cache_data(ttl=300, show_spinner=False)
def _cached_carregar_compras_local(path: str, mtime: float = 0.0):
    """Carrega dados do JSON com cache de 5 min (invalidado por mtime)."""
    if not os.path.exists(path):
        return [], []
    try:
        with open(path, "r", encoding="utf-8") as f:
            conteudo = f.read()
        if not conteudo.strip():
            return [], []
        dados = json.loads(conteudo)
        return dados.get("solicitacoes", []), dados.get("entregas", [])
    except Exception:
        return [], []

# ====================== GOOGLE SHEETS INTEGRAÇÃO ======================
# Verifica se as credenciais do Google Sheets estão configuradas
GS_ENABLED = False
gc = None
GS_ID_FUNCIONARIOS = None
GS_ID_DIARIAS = None
GS_ID_COMPRAS = None

try:
    import gspread
    from google.oauth2.service_account import Credentials
    
    if "gspread" in st.secrets:
        creds_dict = dict(st.secrets["gspread"])
        creds = Credentials.from_service_account_info(creds_dict, scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ])
        gc = gspread.authorize(creds)
        GS_ID_FUNCIONARIOS = st.secrets.get("gsheets", {}).get("id_funcionarios", "")
        GS_ID_DIARIAS = st.secrets.get("gsheets", {}).get("id_diarias", "")
        GS_ID_COMPRAS = st.secrets.get("gsheets", {}).get("id_compras", "")
        if GS_ID_FUNCIONARIOS or GS_ID_DIARIAS or GS_ID_COMPRAS:
            GS_ENABLED = True
except Exception:
    pass


# ====================== MÓDULO DE COMPRAS EMBUTIDO ======================
# -*- coding: utf-8 -*-
"""Módulo Sistema de Compras e Entregas - Streamlit"""
import streamlit as st
import pandas as pd
import uuid
from datetime import datetime, date
import io
import base64

# ========== FRAGMENT SUPPORT (Streamlit >= 1.38) ==========
# Isola reruns para não recarregar todo o app a cada clique
try:
    fragment = st.fragment
except AttributeError:
    try:
        fragment = st.experimental_fragment
    except AttributeError:
        def fragment(func):
            """Fallback: executa normalmente se fragment não disponível."""
            return func


# ========== DADOS ==========
# Fallback estático (usado apenas quando Auxiliares não tem dados de Cliente/Loja)
CLIENTES_FALLBACK = ["Smart Fit", "Self Fit", "Assaí Atacadista"]

LOJAS_POR_CLIENTE_FALLBACK = {
    "Smart Fit": [
        "Smart Fit Shopping Manoa", "Smart Fit Shopping Cidade Leste", "Smart Fit Macapá Shopping",
        "Smart Fit Shopping Grande Circular", "Smart Fit Shopping Via Norte", "Smart Fit Cidade Nova",
        "Smart Fit Parque Mosaico", "Smart Fit Cachoeirinha", "Smart Fit Flores", "Smart Fit Ponta Negra",
        "Smart Fit Nova Porto Velho", "Smart Fit Porto Velho Flodoaldo", "Smart Fit Alvorada",
        "Smart Fit Novo Aleixo", "Smart Fit São José do Operário", "Smart Fit Santana Macapá",
        "Smart Fit Toequato Tapajós"
    ],
    "Self Fit": [
        "Self Fit Hiper DB Ponta Negra", "Self Fit Manaus Plaza Shopping", "Self Fit Vieira Alves"
    ],
    "Assaí Atacadista": [
        "Assaí Atacadista Batista Campos", "Assaí Atacadista Almirante Barroso", "Assaí Atacadista Castanhal",
        "Assaí Atacadista Ananindeua", "Assaí Atacadista Augusto Monte Negro", "Assaí Atacadista Boa Vista",
        "Assaí Atacadista Manaus", "Assaí Atacadista Macapá", "Assaí Atacadista Belém"
    ]
}

MATERIAIS_POR_CLIENTE_FALLBACK = {
    "Smart Fit": [
        "ÁGUA SANITÁRIA", "ASPIRADOR SEMI-INDUSTRIAL 23L", "BALDE 15L", "BALDE 6L",
        "BALDE ESPREMEDOR COMPLETO", "CABO DE ALUMINIO SEM ROSCA", "DISCO VERMELHO 510",
        "ENCERADEIRA INDUSTRIAL", "ESCOVA DE MÃO", "ESCOVA SANITÁRIA", "ESPONJA DUPLA FACE",
        "EXTENSÃO DE 30 M", "FIBRAS DE LIMPEZA PESADA", "FLANELAS", "KIT LIMPA VIDRO 2 EM 1 BRALIMPIA",
        "LIMPA TUDO (MINI LOK)", "PÁ DE LIXO COMPLETA", "PANO DE CHÃO ALVEJADO", "PLACAS SINALIZADORA",
        "REFIL MOP ÁGUA", "RODO DE 60cm COMPLETO", "SABÃO EM PÓ", "SACO DE LIXO 100 L",
        "SACO DE LIXO 40 L", "SACO DE LIXO 60 L", "VASSOURA DE NYLON COMPLETA", "VASSOURA DE TETO COMPLETA"
    ],
    "Self Fit": [
        "ASPIRADOR DE PÓ E BATERIA SEM FIO", "BAUDE EXPREMEDOR", "REFIL MOP ÁGUA", "REFIL MOP PÓ",
        "PLACAS SINALIZADORA", "ENCERADEIRA INDUSTRIAL", "KIT LIMPA VIDRO 2 EM 1 BRALIMPIA",
        "PÁ COLETORA DE LIXO", "PULVERIZADOR", "CESTA MULTIUSO"
    ],
    "Assaí Atacadista": [
        "Disco 510 mm - Preto", "Disco 510 mm - Marron para Remoção", "Disco 510 mm - vermelho",
        "Disco 510 mm - Verde", "Disco pelo de porco para Polidora", "Disco champanhe para Polidora",
        "Starlock frange 510mm aço", "Starlock frange 510mm com Velcon", "Starlock frange 510mm escova",
        "Enceradeira industrial 510", "Armação Mop Pó 60 cm", "Armação Mop Pó 1,20 cm", "Refil mop cera",
        "Suporte mop cera", "Lt", "Esponja p/LT", "Rodo madeira 1,20mts", "Rodo 60 cm",
        "Cabeleira Mop Agua", "Mop Pó 60 cm", "Mop Pó 1,20 cm", "Saco Amarelo P/Carrinho",
        "Pa coletora azul pop", "Raspador pesado sem cabo", "Raspador pesado com cabo", "Garra mop Agua",
        "Carro coletor de lixo 240lts", "Cabo de aluminio com rosca", "Cabo de aluminio sem rosca",
        "Lâminas p/ raspdor", "Raspador de mão", "Extenção cabo PP 3x2,5", "Borracha Organizadora Carrinho Funcional",
        "Carrinho funcional kit completo", "Balde expremedor", "Kit manunteção para carrinho", "Vassoura nylon",
        "Vassoura Piassava", "Vassourão gari 60 cm", "Alongador 9mts", "Regador"
    ]
}

EPIS_POR_CLIENTE_FALLBACK = {
    "Smart Fit": [
        "Luva látex", "Óculos de proteção", "Luva de Vinil", "Máscara de Proteção",
        "Protetor auricular plug", "Protetor tipo concha", "Luva para jardineiro", "Avental de raspa",
        "Viseira", "Perneira", "Meia Térmica", "Japonha Térmica", "Calça Térmica", "Luvas térmicas",
        "Capuz Térmico", "Avental Térmico", "Bota C.Médio Nº34", "Bota C.Médio Nº35", "Bota C.Médio Nº36",
        "Bota C.Médio Nº37", "Bota C.Médio Nº38", "Bota C.Médio Nº39", "Bota C.Médio Nº40",
        "Bota C.Médio Nº41", "Bota C.Médio Nº42", "Bota C.Médio Nº43", "Bota C.Médio Nº44",
        "Bota C.Médio Nº45", "Bota C.Médio Nº46", "Bota de Couro Nº34", "Bota de Couro Nº35",
        "Bota de Couro Nº36", "Bota de Couro Nº37", "Bota de Couro Nº38", "Bota de Couro Nº39",
        "Bota de Couro Nº40", "Bota de Couro Nº41", "Bota de Couro Nº42", "Bota de Couro Nº43",
        "Bota de Couro Nº44", "Bota de Couro Nº45", "Bota de Couro Nº46", "Sapato Ant-derrapante Nº34",
        "Sapato Ant-derrapante Nº35", "Sapato Ant-derrapante Nº36", "Sapato Ant-derrapante Nº37",
        "Sapato Ant-derrapante Nº38", "Sapato Ant-derrapante Nº39", "Sapato Ant-derrapante Nº40",
        "Sapato Ant-derrapante Nº41", "Sapato Ant-derrapante Nº42", "Sapato Ant-derrapante Nº43",
        "Sapato Ant-derrapante Nº44", "Sapato Ant-derrapante Nº45", "Sapato Ant-derrapante Nº46",
        "Farda C.Feminino (P)", "Farda C.Feminino (M)", "Farda C.Feminino (G)", "Farda C.Feminino (GG)",
        "Farda C.Feminino (XG)", "Farda C.Masculino (P)", "Farda C.Masculino (M)", "Farda C.Masculino (G)",
        "Farda C.Masculino (GG)", "Farda C.Masculino (XG)", "Farda p/ Jardineiro", "Farda p/ Encarregado & Líder",
        "Farda p/ Supervisor", "Camisa Branca", "Cauça", "Chapéu"
    ],
    "Self Fit": [
        "Luva látex", "Óculos de proteção", "Luva de Vinil", "Máscara de Proteção",
        "Protetor auricular plug", "Protetor tipo concha", "Luva para jardineiro", "Avental de raspa",
        "Viseira", "Perneira", "Meia Térmica", "Japonha Térmica", "Calça Térmica", "Luvas térmicas",
        "Capuz Térmico", "Avental Térmico", "Bota C.Médio Nº34", "Bota C.Médio Nº35", "Bota C.Médio Nº36",
        "Bota C.Médio Nº37", "Bota C.Médio Nº38", "Bota C.Médio Nº39", "Bota C.Médio Nº40",
        "Bota C.Médio Nº41", "Bota C.Médio Nº42", "Bota C.Médio Nº43", "Bota C.Médio Nº44",
        "Bota C.Médio Nº45", "Bota C.Médio Nº46", "Bota de Couro Nº34", "Bota de Couro Nº35",
        "Bota de Couro Nº36", "Bota de Couro Nº37", "Bota de Couro Nº38", "Bota de Couro Nº39",
        "Bota de Couro Nº40", "Bota de Couro Nº41", "Bota de Couro Nº42", "Bota de Couro Nº43",
        "Bota de Couro Nº44", "Bota de Couro Nº45", "Bota de Couro Nº46", "Sapato Ant-derrapante Nº34",
        "Sapato Ant-derrapante Nº35", "Sapato Ant-derrapante Nº36", "Sapato Ant-derrapante Nº37",
        "Sapato Ant-derrapante Nº38", "Sapato Ant-derrapante Nº39", "Sapato Ant-derrapante Nº40",
        "Sapato Ant-derrapante Nº41", "Sapato Ant-derrapante Nº42", "Sapato Ant-derrapante Nº43",
        "Sapato Ant-derrapante Nº44", "Sapato Ant-derrapante Nº45", "Sapato Ant-derrapante Nº46",
        "Farda C.Feminino (P)", "Farda C.Feminino (M)", "Farda C.Feminino (G)", "Farda C.Feminino (GG)",
        "Farda C.Feminino (XG)", "Farda C.Masculino (P)", "Farda C.Masculino (M)", "Farda C.Masculino (G)",
        "Farda C.Masculino (GG)", "Farda C.Masculino (XG)", "Farda p/ Jardineiro", "Farda p/ Encarregado & Líder",
        "Farda p/ Supervisor", "Camisa Branca", "Cauça", "Chapéu"
    ],
    "Assaí Atacadista": [
        "Luva látex", "Óculos de proteção", "Luva de Vinil", "Máscara de Proteção",
        "Protetor auricular plug", "Protetor tipo concha", "Luva para jardineiro", "Avental de raspa",
        "Viseira", "Perneira", "Meia Térmica", "Japonha Térmica", "Calça Térmica", "Luvas térmicas",
        "Capuz Térmico", "Avental Térmico", "Bota C.Médio Nº34", "Bota C.Médio Nº35", "Bota C.Médio Nº36",
        "Bota C.Médio Nº37", "Bota C.Médio Nº38", "Bota C.Médio Nº39", "Bota C.Médio Nº40",
        "Bota C.Médio Nº41", "Bota C.Médio Nº42", "Bota C.Médio Nº43", "Bota C.Médio Nº44",
        "Bota C.Médio Nº45", "Bota C.Médio Nº46", "Bota de Couro Nº34", "Bota de Couro Nº35",
        "Bota de Couro Nº36", "Bota de Couro Nº37", "Bota de Couro Nº38", "Bota de Couro Nº39",
        "Bota de Couro Nº40", "Bota de Couro Nº41", "Bota de Couro Nº42", "Bota de Couro Nº43",
        "Bota de Couro Nº44", "Bota de Couro Nº45", "Bota de Couro Nº46", "Sapato Ant-derrapante Nº34",
        "Sapato Ant-derrapante Nº35", "Sapato Ant-derrapante Nº36", "Sapato Ant-derrapante Nº37",
        "Sapato Ant-derrapante Nº38", "Sapato Ant-derrapante Nº39", "Sapato Ant-derrapante Nº40",
        "Sapato Ant-derrapante Nº41", "Sapato Ant-derrapante Nº42", "Sapato Ant-derrapante Nº43",
        "Sapato Ant-derrapante Nº44", "Sapato Ant-derrapante Nº45", "Sapato Ant-derrapante Nº46",
        "Farda C.Feminino (P)", "Farda C.Feminino (M)", "Farda C.Feminino (G)", "Farda C.Feminino (GG)",
        "Farda C.Feminino (XG)", "Farda C.Masculino (P)", "Farda C.Masculino (M)", "Farda C.Masculino (G)",
        "Farda C.Masculino (GG)", "Farda C.Masculino (XG)", "Farda p/ Jardineiro", "Farda p/ Encarregado & Líder",
        "Farda p/ Supervisor", "Camisa Branca", "Cauça", "Chapéu"
    ]
}

STATUS_OPCOES = ["Pendente", "Aprovado", "Em Trânsito", "Entregue", "Cancelado"]
TIPOS_SOLICITACAO = ["Material", "EPI"]
PRIORIDADES = ["Normal", "Urgente", "Baixa"]

TAMANHOS_EPI = ["", "P", "M", "G", "GG", "37", "38", "39", "40", "41", "42", "43", "44"]


def gerar_id():
    return "SOL-" + uuid.uuid4().hex[:8].upper()


def formatar_data_br(d):
    if not d:
        return "-"
    if isinstance(d, str):
        if "T" in d:
            d = d.split("T")[0]
        try:
            y, m, day = d.split("-")
            return f"{day}/{m}/{y}"
        except Exception:
            return d
    if isinstance(d, (datetime, date)):
        return d.strftime("%d/%m/%Y")
    return str(d)


def formatar_moeda(v):
    if not v:
        return "R$ 0,00"
    try:
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v)


def init_session_state():
    if "compras_solicitacoes" not in st.session_state:
        sols, ents = [], []
        # Tenta carregar do Google Sheets
        if GS_ENABLED and GS_ID_COMPRAS:
            try:
                sols, ents = _carregar_compras_gs()
            except Exception:
                sols, ents = [], []
        # Se não conseguiu do GS, carrega do arquivo local (com cache)
        if not sols and not ents:
            mtime = os.path.getmtime(ARQUIVO_COMPRAS) if os.path.exists(ARQUIVO_COMPRAS) else 0.0
            sols, ents = _cached_carregar_compras_local(ARQUIVO_COMPRAS, mtime)
            # Fallback de seguranca: se o arquivo existe e tem conteudo mas o cache retornou vazio, recarrega direto
            if not sols and not ents and os.path.exists(ARQUIVO_COMPRAS) and os.path.getsize(ARQUIVO_COMPRAS) > 10:
                try:
                    with open(ARQUIVO_COMPRAS, "r", encoding="utf-8") as f:
                        dados = json.load(f)
                    sols = dados.get("solicitacoes", [])
                    ents = dados.get("entregas", [])
                except Exception:
                    pass
        st.session_state["compras_solicitacoes"] = sols
        st.session_state["compras_entregas"] = ents
    if "compras_entregas" not in st.session_state:
        st.session_state["compras_entregas"] = []
    if "compras_page" not in st.session_state:
        st.session_state["compras_page"] = "Dashboard"
    if "compras_edit_id" not in st.session_state:
        st.session_state["compras_edit_id"] = None
    if "compras_nova_itens_material" not in st.session_state:
        st.session_state["compras_nova_itens_material"] = []
    if "compras_nova_itens_epi" not in st.session_state:
        st.session_state["compras_nova_itens_epi"] = []
    if "compras_form_reset" not in st.session_state:
        st.session_state["compras_form_reset"] = False
    if "compras_gs_loaded" not in st.session_state:
        st.session_state["compras_gs_loaded"] = True


def _rerun_fragment():
    try:
        st.rerun(scope="fragment")
    except TypeError:
        st.rerun()

def switch_page(page):
    st.session_state["compras_page"] = page
    st.session_state["compras_edit_id"] = None
    _rerun_fragment()


def _cb_switch_page(page):
    """Callback para navegação por botão. Executa antes do script, pode modificar key do radio."""
    st.session_state["compras_page"] = page
    st.session_state["compras_edit_id"] = None
    if "nav_compras" in st.session_state:
        st.session_state["nav_compras"] = page


def _cb_voltar_solicitacoes():
    """Callback para voltar à lista de solicitações."""
    st.session_state["compras_page"] = "Solicitações"
    st.session_state["compras_ver_id"] = None
    if "nav_compras" in st.session_state:
        st.session_state["nav_compras"] = "Solicitações"


def _cb_ver_detalhes(sol_id):
    """Callback para ver detalhes de uma solicitação."""
    st.session_state["compras_ver_id"] = sol_id
    st.session_state["compras_page"] = "Detalhes"


def _cb_editar_solicitacao(sol_id):
    """Callback para editar uma solicitação."""
    st.session_state["compras_edit_id"] = sol_id
    st.session_state["compras_page"] = "Nova Solicitação"
    if "nav_compras" in st.session_state:
        st.session_state["nav_compras"] = "Nova Solicitação"


def get_badge_color(status):
    return {
        "Pendente": "🔴",
        "Aprovado": "🟢",
        "Em Trânsito": "🔵",
        "Entregue": "🟢",
        "Cancelado": "⚫"
    }.get(status, "⚪")


# ========== GERAÇÃO DE DOCUMENTOS ==========
def gerar_doc_epi(sol):
    data_geracao = datetime.now().strftime("%d/%m/%Y")
    cliente = sol.get("cliente", "_________________")
    loja = sol.get("loja", "_________________")
    map_itens = {i.get("epi", ""): i for i in sol.get("itens", [])}

    epi_list = [
        "Luva látex", "Luva para jardineiro", "Japona Térmica",
        "Óculos de proteção", "Avental de raspa", "Calça Térmica",
        "Máscara de Proteção", "Viseira", "Luvas térmicas",
        "Protetor auricular plug ( ) ou concha ( )", "Perneira", "Cap", "", "", ""
    ]
    epi_rows = []
    for i in range(0, len(epi_list), 3):
        e1 = epi_list[i] if i < len(epi_list) else ""
        e2 = epi_list[i+1] if i+1 < len(epi_list) else ""
        e3 = epi_list[i+2] if i+2 < len(epi_list) else ""
        q1 = map_itens.get(e1, {}).get("qtd", "&nbsp;") if e1 else "&nbsp;"
        q2 = map_itens.get(e2, {}).get("qtd", "&nbsp;") if e2 else "&nbsp;"
        q3 = map_itens.get(e3, {}).get("qtd", "&nbsp;") if e3 else "&nbsp;"
        epi_rows.append(
            f'<tr><td style="border:1px solid #000;padding:3px 5px;font-size:10px">{e1 or "&nbsp;"}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px;text-align:center">{q1}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px">{e2 or "&nbsp;"}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px;text-align:center">{q2}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px">{e3 or "&nbsp;"}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px;text-align:center">{q3}</td></tr>'
        )

    itens_arr = [(i.get("epi", i.get("nome", "Item")) + (f' x{i.get("qtd", "")}' if i.get("qtd") else "")) for i in sol.get("itens", [])]
    botas_rows = []
    total_rows = max(17, len(itens_arr))
    for i in range(total_rows):
        item_obj = sol.get("itens", [])[i] if i < len(sol.get("itens", [])) else None
        nome = item_obj.get("colaborador", sol.get("nomeFuncionario", "&nbsp;")) if item_obj else "&nbsp;"
        loja_val = sol.get("loja", "&nbsp;") if item_obj else "&nbsp;"
        enc = sol.get("encarregado", "&nbsp;") if item_obj else "&nbsp;"
        sup = sol.get("supervisor", "&nbsp;") if item_obj else "&nbsp;"
        item = itens_arr[i] if i < len(itens_arr) else "&nbsp;"
        botas_rows.append(
            f'<tr><td style="border:1px solid #000;padding:3px 5px;font-size:10px">{nome}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px">{loja_val}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px">{enc}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px">{sup}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px">{item}</td>'
            f'<td style="border:1px solid #000;padding:3px 5px;font-size:10px">&nbsp;</td></tr>'
        )

    states = 'ALAGOAS ( &nbsp;) &nbsp;&nbsp;BAHIA ( &nbsp;) &nbsp;&nbsp;CEARÁ ( &nbsp;) &nbsp;&nbsp;MARANHÃO ( &nbsp;) &nbsp;&nbsp;PARAIBA ( &nbsp;) &nbsp;&nbsp;PARÁ ( &nbsp;) &nbsp;&nbsp;PERNAMBUCO ( &nbsp;) &nbsp;&nbsp;PIAUÍ ( &nbsp;) &nbsp;&nbsp;RIO GRANDE DO NORTE ( &nbsp;) &nbsp;&nbsp;SERGIPE ( &nbsp;) &nbsp;&nbsp;AMAPÁ ( &nbsp;) &nbsp;&nbsp;RORAIMA ( &nbsp;) &nbsp;&nbsp;AMAZONAS ( &nbsp;) &nbsp;&nbsp;RONDÔNIA ( &nbsp;)'

    html = f"""<html xmlns:o='urn:schemas-microsoft-com:office:office' xmlns:w='urn:schemas-microsoft-com:office:word' xmlns='http://www.w3.org/TR/REC-html40'>
<head><meta charset='utf-8'><title>Solicitacao de EPI - {cliente} - {loja}</title>
<style>@page Section1 {{ size: 841.95pt 595.35pt; margin: 36pt 36pt 36pt 36pt; mso-page-orientation: landscape; }}
div.Section1 {{ page: Section1; }}</style></head>
<body style="font-family:Arial,sans-serif;font-size:10pt;margin:0;padding:0">
<div class="Section1" style="mso-page-orientation: landscape;">
<table style="width:100%;border-collapse:collapse">
<tr><td colspan="6" style="border:1px solid #000;padding:3px 5px;font-size:9px;text-align:center">{states}</td></tr>
<tr><td colspan="6" style="border:1px solid #000;padding:3px 5px;font-size:11px;font-weight:bold;background:#DEEBF6;text-align:center">BOTAS</td></tr>
<tr>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Nome</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Loja</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Encarregado</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Supervisor</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Itens da Solicitação de EPI</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Situação ( * )</th>
</tr>
{''.join(botas_rows)}
<tr><td colspan="6" style="border:1px solid #000;padding:3px 5px;font-size:11px;font-weight:bold;background:#DEEBF6;text-align:center">EPIS</td></tr>
<tr>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Equipamento de Proteção</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Quantidade</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Equipamento de Proteção</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Quantidade</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Equipamento de Proteção</th>
<th style="border:1px solid #000;padding:3px 5px;font-size:10px;font-weight:bold;background:#DEEBF6">Quantidade</th>
</tr>
{''.join(epi_rows)}
<tr><td colspan="6" style="border:1px solid #000;padding:3px 5px;font-size:9px">Observação: A coluna Situação ( * ), é para o setor SST preencher.</td></tr>
</table>
<p style="font-size:8pt;color:#666;text-align:center;margin-top:6px">Data de geração: {data_geracao}</p>
</div></body></html>"""
    return html


def gerar_xls_material(sol):
    data_geracao = datetime.now().strftime("%d/%m/%Y")
    cliente = sol.get("cliente", "")
    itens = sol.get("itens", [])

    def build_rows_smart_self(items):
        if not items:
            return '<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px">&nbsp;</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center">&nbsp;</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center">UN.</td></tr>'
        return "".join([
            f'<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;white-space:nowrap">{i.get("material","")}</td>'
            f'<td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center;white-space:nowrap">{i.get("qtd","&nbsp;")}</td>'
            f'<td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center;white-space:nowrap">UN.</td></tr>'
            for i in items
        ])

    def build_rows_assai(items):
        if not items:
            return '<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center">1</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px">&nbsp;</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px">&nbsp;</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center">&nbsp;</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center">peças</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:right">&nbsp;</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:right">&nbsp;</td></tr>'
        rows = []
        for idx, i in enumerate(items, 1):
            qtd = i.get("qtd", "&nbsp;")
            unit = f'R$ {i["valorUnit"]:.2f}'.replace(".", ",") if i.get("valorUnit") else "&nbsp;"
            total = f'R$ {(i["valorUnit"]*i["qtd"]):.2f}'.replace(".", ",") if i.get("valorUnit") and i.get("qtd") else "&nbsp;"
            rows.append(
                f'<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center;white-space:nowrap">{idx}</td>'
                f'<td style="border:1px solid #000;padding:4px 6px;font-size:11px;white-space:nowrap">{i.get("material","")}</td>'
                f'<td style="border:1px solid #000;padding:4px 6px;font-size:11px;white-space:nowrap">&nbsp;</td>'
                f'<td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center;white-space:nowrap">{qtd}</td>'
                f'<td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:center;white-space:nowrap">peças</td>'
                f'<td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:right;white-space:nowrap">{unit}</td>'
                f'<td style="border:1px solid #000;padding:4px 6px;font-size:11px;text-align:right;white-space:nowrap">{total}</td></tr>'
            )
        return "".join(rows)

    if cliente == "Smart Fit":
        rows = build_rows_smart_self(itens)
        html = f"""<html xmlns:x="urn:schemas-microsoft-com:office:excel"><head><meta charset="utf-8"><style>table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #000}}</style></head><body><table>
<tr><th colspan="3" style="border:1px solid #000;padding:6px;font-size:13px;font-weight:bold;background:#DEEBF6;text-align:center">SMART FIT</th></tr>
<tr><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">SMART FIT</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">QTD</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">REF.</th></tr>
{rows}
<tr><td colspan="3" style="border:1px solid #000;padding:4px 6px;font-size:10px">Solicitação: {sol.get("id","")} | Loja: {sol.get("loja","")} | Data: {data_geracao}</td></tr>
</table></body></html>"""
        filename = f'Pedido_Material_SmartFit_{sol.get("id","")}.xls'
    elif cliente == "Self Fit":
        rows = build_rows_smart_self(itens)
        html = f"""<html xmlns:x="urn:schemas-microsoft-com:office:excel"><head><meta charset="utf-8"><style>table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #000}}</style></head><body><table>
<tr><th colspan="3" style="border:1px solid #000;padding:6px;font-size:13px;font-weight:bold;background:#DEEBF6;text-align:center">SELF FIT</th></tr>
<tr><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">SELF FIT</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">QTD</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">REF.</th></tr>
{rows}
<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">DATA DO PEDIDO</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">&nbsp;</td></tr>
<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">MÊS DE REFERÊNCIA</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">&nbsp;</td></tr>
<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">SEPARADO</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">&nbsp;</td></tr>
<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">ENVIADO</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">&nbsp;</td></tr>
<tr><td colspan="3" style="border:1px solid #000;padding:4px 6px;font-size:10px">Solicitação: {sol.get("id","")} | Loja: {sol.get("loja","")} | Data: {data_geracao}</td></tr>
</table></body></html>"""
        filename = f'Pedido_Material_SelfFit_{sol.get("id","")}.xls'
    elif cliente == "Assaí Atacadista":
        item_rows = build_rows_assai(itens)
        html = f"""<html xmlns:x="urn:schemas-microsoft-com:office:excel"><head><meta charset="utf-8"><style>table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #000}}</style></head><body>
<table style="width:100%;border-collapse:collapse">
<tr><td colspan="7" style="border:1px solid #000;padding:6px;font-size:11px;text-align:center;font-weight:bold">R: Sgto Jeter Augusto Pereira N° 02 e 04 - São Paulo - CEP: 02188-070 - E-mail: vendas@thamesjlara.com.br - Site www.thamesjlara.com.br</td></tr>
<tr><td colspan="7" style="height:10px"></td></tr>
<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">Orçamento</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">At.: Sr.(a): Mendonça</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">Data: {data_geracao}</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">Vendedor: Hélio</td></tr>
<tr><td colspan="7" style="height:10px"></td></tr>
<tr><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">Razão Social</td><td colspan="3" style="border:1px solid #000;padding:4px 6px;font-size:11px">FG Services Eireli - ME</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">CNPJ/CPF</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px">23.585.374/0001-11</td></tr>
<tr><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">Endereço</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">Av. Barão de Vera Cruz, 586 BR 101 Norte</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">Bairro</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px">Cruz de Rebouças</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">Cidade</td></tr>
<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px">Igarassu</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">UF</td><td style="border:1px solid #000;padding:4px 6px;font-size:11px">PE</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">CEP</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">53635-015</td></tr>
<tr><td colspan="7" style="height:10px"></td></tr>
<tr><td style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">Telefone</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">(0xx81) 3545-3990</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold">E-mail</td><td colspan="2" style="border:1px solid #000;padding:4px 6px;font-size:11px">&nbsp;</td></tr>
<tr><td colspan="7" style="height:10px"></td></tr>
<tr><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">Item</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">Descrição</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">Marca</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">Qtde</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">Unid</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">Valor Unit.</th><th style="border:1px solid #000;padding:4px 6px;font-size:11px;font-weight:bold;background:#DEEBF6;white-space:nowrap">Total</th></tr>
{item_rows}
<tr><td colspan="7" style="border:1px solid #000;padding:4px 6px;font-size:10px">Solicitação: {sol.get("id","")} | Loja: {sol.get("loja","")} | Data: {data_geracao}</td></tr>
</table></body></html>"""
        filename = f'Orcamento_Assai_{sol.get("id","")}.xls'
    else:
        return None, None
    return html, filename


# ========== PÁGINAS ==========
# ========== CACHE DE DATAFRAMES ==========
@st.cache_data(show_spinner=False)
def _cached_df_solicitacoes(solicitacoes_json):
    sols = json.loads(solicitacoes_json)
    if not sols:
        return pd.DataFrame()
    return pd.DataFrame([
        {
            "ID": s["id"],
            "Data": formatar_data_br(s.get("data")),
            "Loja": s.get("loja", ""),
            "Cliente": s.get("cliente", ""),
            "Tipo": s.get("tipo", ""),
            "Solicitante": s.get("solicitante", ""),
            "Itens": len(s.get("itens", [])),
            "Valor": formatar_moeda(s.get("valorTotal", 0)),
            "Status": f"{get_badge_color(s.get('status'))} {s.get('status', '')}",
        }
        for s in sols
    ])

@st.cache_data(show_spinner=False)
def _cached_df_entregas(entregas_json):
    ents = json.loads(entregas_json)
    if not ents:
        return pd.DataFrame()
    return pd.DataFrame([
        {
            "ID": e.get("id", ""),
            "Solicitação": e.get("idSolicitacao", ""),
            "Data": formatar_data_br(e.get("data")),
            "Transportadora": e.get("transportadora", ""),
            "Rastreio": e.get("rastreio", ""),
            "Status": e.get("status", ""),
        }
        for e in ents
    ])

def page_dashboard():
    st.markdown("### 📊 Dashboard")
    sols = st.session_state["compras_solicitacoes"]
    ents = st.session_state["compras_entregas"]

    total = len(sols)
    pendentes = sum(1 for s in sols if s.get("status") == "Pendente")
    transito = sum(1 for s in sols if s.get("status") == "Em Trânsito")
    entregues = sum(1 for s in sols if s.get("status") == "Entregue")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Solicitações", total)
    c2.metric("Pendentes", pendentes)
    c3.metric("Em Trânsito", transito)
    c4.metric("Entregues", entregues)

    st.markdown("---")
    st.markdown("#### 📋 Últimas Solicitações")
    if sols:
        recentes = sorted(sols, key=lambda x: x.get("dataCriacao", ""), reverse=True)[:10]
        df = _cached_df_solicitacoes(json.dumps(recentes, ensure_ascii=False, default=str))
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhuma solicitação cadastrada.")

    st.button("➕ Nova Solicitação", type="primary", key="btn_nova_sol_dashboard",
              on_click=_cb_switch_page, args=("Nova Solicitação",))



def _salvar_compras_local(solicitacoes, entregas):
    """Salva dados de compras em arquivo JSON local (apenas se houver mudanças)."""
    try:
        dados = {"solicitacoes": solicitacoes, "entregas": entregas}
        dados_json = json.dumps(dados, ensure_ascii=False, sort_keys=True)
        hash_novo = hashlib.md5(dados_json.encode("utf-8")).hexdigest()
        chave_cache = "_compras_hash"
        hash_antigo = st.session_state.get(chave_cache, "")
        if hash_novo == hash_antigo:
            return  # sem mudanças, ignora
        
        # BACKUP AUTOMÁTICO antes de salvar
        try:
            pasta_bkp, arqs = criar_backup_automatico()
        except Exception:
            pass
        # Protecao: nao salvar dados vazios por cima de arquivo existente com dados
        if not solicitacoes and not entregas and os.path.exists(ARQUIVO_COMPRAS) and os.path.getsize(ARQUIVO_COMPRAS) > 10:
            try:
                with open(ARQUIVO_COMPRAS, "r", encoding="utf-8") as f:
                    dados_existentes = json.load(f)
                if dados_existentes.get("solicitacoes") or dados_existentes.get("entregas"):
                    st.warning("⚠️ Tentativa de salvar dados vazios detectada e bloqueada. Arquivo original preservado.")
                    return
            except Exception:
                pass
        st.session_state[chave_cache] = hash_novo
        with open(ARQUIVO_COMPRAS, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
    except Exception as e:
        st.error(f"❌ Erro ao salvar compras localmente: {e}")


def _carregar_compras_local():
    """Carrega dados de compras do arquivo JSON local."""
    caminho = os.path.abspath(ARQUIVO_COMPRAS)
    if not os.path.exists(ARQUIVO_COMPRAS):
        st.toast(f"📂 Nenhum arquivo de compras encontrado em: {caminho}")
        return [], []
    try:
        with open(ARQUIVO_COMPRAS, "r", encoding="utf-8") as f:
            conteudo = f.read()
        if not conteudo.strip():
            st.warning(f"⚠️ Arquivo de compras vazio: {caminho}")
            # Tentar restaurar do backup
            sols, ents = _tentar_restaurar_compras()
            if sols or ents:
                st.toast("🛡️ Compras recuperadas automaticamente do backup!")
                return sols, ents
            return [], []
        dados = json.loads(conteudo)
        sols = dados.get("solicitacoes", [])
        ents = dados.get("entregas", [])
        # PROTEÇÃO CONTRA DADOS VAZIOS: tentar restaurar do último backup se os dados foram perdidos
        if _compras_estao_vazias(sols, ents):
            sols_bkp, ents_bkp = _tentar_restaurar_compras()
            if sols_bkp or ents_bkp:
                st.toast("🛡️ Compras recuperadas automaticamente do backup!")
                return sols_bkp, ents_bkp
        st.toast(f"📂 Compras carregadas do arquivo ({len(sols)} solicitações, {len(ents)} entregas)")
        return sols, ents
    except Exception as e:
        st.error(f"❌ Erro ao carregar compras do arquivo {caminho}: {e}")
        # Tentar restaurar do backup em caso de erro
        sols, ents = _tentar_restaurar_compras()
        if sols or ents:
            st.toast("🛡️ Compras recuperadas automaticamente do backup!")
            return sols, ents
        return [], []


def _tentar_restaurar_compras():
    """Tenta restaurar compras do último backup válido."""
    try:
        backups = listar_backups_automaticos()
        if not backups:
            return [], []
        for _, _, bkp_path in backups:
            try:
                with zipfile.ZipFile(bkp_path, "r") as zf:
                    if os.path.basename(ARQUIVO_COMPRAS) in zf.namelist():
                        conteudo = zf.read(os.path.basename(ARQUIVO_COMPRAS)).decode("utf-8")
                        dados = json.loads(conteudo)
                        sols = dados.get("solicitacoes", [])
                        ents = dados.get("entregas", [])
                        if sols or ents:
                            return sols, ents
            except Exception:
                continue
        return [], []
    except Exception:
        return [], []


def _salvar_compras_automatico():
    """Salva automaticamente o estado atual do módulo de compras no Google Sheets e localmente."""
    import sys, threading
    sols = st.session_state.get("compras_solicitacoes", [])
    ents = st.session_state.get("compras_entregas", [])
    # Sempre salva localmente para garantir persistência (com cache de hash)
    _salvar_compras_local(sols, ents)
    if not GS_ENABLED or not GS_ID_COMPRAS:
        return

    def _bg_save():
        try:
            _salvar_compras_gs(sols, ents)
        except Exception as e:
            print(f"[DEBUG] Falha ao salvar no GS: {e}", file=sys.stderr)

    threading.Thread(target=_bg_save, daemon=True).start()


# Inicialização Google Sheets: garante que abas existam

def page_solicitacoes():
    st.markdown("### 📋 Todas as Solicitações")

    sols = st.session_state["compras_solicitacoes"]

    # ── Filtros (compactos, 1 linha) ──
    with st.expander("🔍 Filtros", expanded=True):
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        with c1:
            busca = st.text_input("Buscar", placeholder="ID, loja, solicitante...", key="sol_busca")
        with c2:
            fstatus = st.selectbox("Status", ["Todos"] + STATUS_OPCOES, key="sol_fstatus")
        with c3:
            floja = st.selectbox("Loja", ["Todas"] + sorted(list(set(l for c in lista_lojas_por_cliente().values() for l in c))), key="sol_floja")
        with c4:
            fcliente = st.selectbox("Cliente", ["Todos"] + lista_clientes(), key="sol_fcliente")
        with c5:
            fdi_txt = st.text_input("Data Início (DD/MM/AAAA)", value="", key="sol_fdi")
            fdi = _parse_data_br(fdi_txt) if fdi_txt.strip() else None
        with c6:
            fdf_txt = st.text_input("Data Fim (DD/MM/AAAA)", value="", key="sol_fdf")
            fdf = _parse_data_br(fdf_txt) if fdf_txt.strip() else None

    # ── Filtragem (lista em memória, rápida) ──
    filtradas = []
    for s in sols:
        txt = f"{s.get('id','')} {s.get('loja','')} {s.get('solicitante','')}".lower()
        if busca and busca.lower() not in txt:
            continue
        if fstatus != "Todos" and s.get("status") != fstatus:
            continue
        if floja != "Todas" and s.get("loja") != floja:
            continue
        if fcliente != "Todos" and s.get("cliente") != fcliente:
            continue
        if fdi:
            try:
                if s.get("data") and str(s["data"]) < str(fdi):
                    continue
            except Exception:
                pass
        if fdf:
            try:
                if s.get("data") and str(s["data"]) > str(fdf):
                    continue
            except Exception:
                pass
        filtradas.append(s)

    if not filtradas:
        st.info("Nenhuma solicitação encontrada.")
        return

    # ── Tabela única (st.dataframe) em vez de 1 container + 3 colunas + 3 botões POR ITEM ──
    df = _cached_df_solicitacoes(json.dumps(filtradas, ensure_ascii=False, default=str))
    st.dataframe(df, use_container_width=True, hide_index=True)

    # ── Ações em lote (3 widgets fixos, independente da quantidade de solicitações) ──
    st.markdown("---")
    st.markdown("#### ⚡ Ação Rápida")
    ids_disponiveis = [s["id"] for s in filtradas]
    c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
    with c1:
        sel_id = st.selectbox("Selecionar solicitação", ids_disponiveis, key="sol_sel_id")
    with c2:
        st.button("👁️ Ver", use_container_width=True, key="btn_ver_sel",
                  on_click=_cb_ver_detalhes, args=(sel_id,))
    with c3:
        st.button("✏️ Editar", use_container_width=True, key="btn_edit_sel",
                  on_click=_cb_editar_solicitacao, args=(sel_id,))
    with c4:
        if st.button("🗑️ Excluir", use_container_width=True, key="btn_del_sel"):
            st.session_state["compras_solicitacoes"] = [x for x in sols if x["id"] != sel_id]
            st.session_state["compras_entregas"] = [e for e in st.session_state["compras_entregas"] if e.get("idSolicitacao") != sel_id]
            _salvar_compras_automatico()
            st.success("Excluído!")
            # Fragmento reexecuta automaticamente

    # ── Exportar CSV ──
    csv_data = []
    for s in filtradas:
        csv_data.append([
            s["id"], s.get("data",""), s.get("loja",""), s.get("cliente",""),
            s.get("tipo",""), s.get("solicitante",""), len(s.get("itens",[])),
            s.get("valorTotal",0), s.get("status","")
        ])
    df_csv = pd.DataFrame(csv_data, columns=["ID","Data","Loja","Cliente","Tipo","Solicitante","Qtd Itens","Valor Total","Status"])
    csv_buffer = io.StringIO()
    df_csv.to_csv(csv_buffer, index=False, sep=";", encoding="utf-8")
    st.download_button("📥 Exportar CSV", data=csv_buffer.getvalue().encode("utf-8"), file_name="solicitacoes.csv", mime="text/csv")


def _build_df_mat(itens):
    return pd.DataFrame(itens)

def _build_df_epi(itens):
    return pd.DataFrame(itens)

def page_nova_solicitacao():
    st.markdown("### ➕ Nova Solicitação de Compra")

    edit_id = st.session_state.get("compras_edit_id")
    edit_sol = None
    if edit_id:
        for s in st.session_state["compras_solicitacoes"]:
            if s["id"] == edit_id:
                edit_sol = s
                break

    # Detectar mudança de contexto (nova solicitação, edição diferente, tipo mudou, cliente mudou)
    # e limpar estado
    last_edit_id = st.session_state.get("compras_last_edit_id")
    if last_edit_id != edit_id:
        keys_to_clear = ["nova_cliente", "nova_loja", "nova_tipo", "nova_solicitante",
                         "nova_data", "nova_prioridade", "nova_previsao", "nova_obs",
                         "nova_nome_func", "nova_encarregado", "nova_supervisor", "nova_data_bota",
                         "compras_edit_loaded"]
        for k in keys_to_clear:
            if k in st.session_state:
                del st.session_state[k]
        st.session_state["compras_last_edit_id"] = edit_id

    # ── INICIALIZA VALORES PADRÃO NO SESSION_STATE ──
    if "nova_cliente" not in st.session_state:
        st.session_state["nova_cliente"] = edit_sol["cliente"] if edit_sol and edit_sol.get("cliente") in lista_clientes() else lista_clientes()[0]
    if "nova_loja" not in st.session_state:
        st.session_state["nova_loja"] = edit_sol["loja"] if edit_sol and edit_sol.get("loja") else ""
    if "nova_tipo" not in st.session_state:
        st.session_state["nova_tipo"] = edit_sol["tipo"] if edit_sol and edit_sol.get("tipo") in TIPOS_SOLICITACAO else TIPOS_SOLICITACAO[0]
    if "nova_solicitante" not in st.session_state:
        st.session_state["nova_solicitante"] = edit_sol.get("solicitante", "") if edit_sol else ""
    if "nova_data" not in st.session_state:
        if edit_sol and edit_sol.get("data"):
            st.session_state["nova_data"] = _iso_para_br(edit_sol["data"])
        else:
            st.session_state["nova_data"] = datetime.now().strftime("%d/%m/%Y")
    if "nova_prioridade" not in st.session_state:
        st.session_state["nova_prioridade"] = edit_sol["prioridade"] if edit_sol and edit_sol.get("prioridade") in PRIORIDADES else PRIORIDADES[0]
    if "nova_previsao" not in st.session_state:
        if edit_sol and edit_sol.get("previsao"):
            st.session_state["nova_previsao"] = _iso_para_br(edit_sol["previsao"])
        else:
            st.session_state["nova_previsao"] = (datetime.now() + timedelta(days=7)).strftime("%d/%m/%Y")
    if "nova_obs" not in st.session_state:
        st.session_state["nova_obs"] = edit_sol.get("observacoes", "") if edit_sol else ""
    if "nova_nome_func" not in st.session_state:
        st.session_state["nova_nome_func"] = edit_sol.get("nomeFuncionario", "") if edit_sol else ""
    if "nova_encarregado" not in st.session_state:
        st.session_state["nova_encarregado"] = edit_sol.get("encarregado", "") if edit_sol else ""
    if "nova_supervisor" not in st.session_state:
        st.session_state["nova_supervisor"] = edit_sol.get("supervisor", "") if edit_sol else ""
    if "nova_data_bota" not in st.session_state:
        if edit_sol and edit_sol.get("dataUltimaBota"):
            st.session_state["nova_data_bota"] = _iso_para_br(edit_sol["dataUltimaBota"])
        else:
            st.session_state["nova_data_bota"] = ""

    # ── WIDGETS COM KEY (sem st.form) ──
    col_sel1, col_sel2, col_sel3 = st.columns(3)
    with col_sel1:
        st.selectbox("Cliente *", lista_clientes(), key="nova_cliente")
    with col_sel2:
        lojas = lista_lojas_por_cliente(st.session_state["nova_cliente"])
        if st.session_state["nova_loja"] not in lojas:
            st.session_state["nova_loja"] = lojas[0] if lojas else ""
        st.selectbox("Loja *", lojas, key="nova_loja")
    with col_sel3:
        st.selectbox("Tipo *", TIPOS_SOLICITACAO, key="nova_tipo")

    c4, c5, c6 = st.columns(3)
    with c4:
        st.text_input("Solicitante *", key="nova_solicitante")
    with c5:
        st.text_input("Data (DD/MM/AAAA)", key="nova_data")
    with c6:
        st.selectbox("Prioridade", PRIORIDADES, key="nova_prioridade")

    st.text_input("Previsão de Entrega (DD/MM/AAAA)", key="nova_previsao")
    st.text_area("Observações", key="nova_obs")

    if st.session_state["nova_tipo"] == "EPI":
        st.markdown("---")
        st.markdown("#### 👷 Informações EPI")
        c7, c8, c9 = st.columns(3)
        with c7:
            st.text_input("Nome do Funcionário", key="nova_nome_func")
        with c8:
            st.text_input("Encarregado", key="nova_encarregado")
        with c9:
            st.text_input("Supervisor", key="nova_supervisor")
        st.text_input("Data Última Bota (DD/MM/AAAA)", key="nova_data_bota")

    # ── VALORES FINAIS DO CABEÇALHO ──
    cliente = st.session_state.get("nova_cliente", lista_clientes()[0])
    loja = st.session_state.get("nova_loja", "")
    tipo = st.session_state.get("nova_tipo", TIPOS_SOLICITACAO[0])
    solicitante = st.session_state.get("nova_solicitante", "")
    data_sol = _parse_data_br(st.session_state.get("nova_data", ""))
    prioridade = st.session_state.get("nova_prioridade", PRIORIDADES[0])
    previsao = _parse_data_br(st.session_state.get("nova_previsao", ""))
    observacoes = st.session_state.get("nova_obs", "")
    nome_func = st.session_state.get("nova_nome_func", "")
    encarregado = st.session_state.get("nova_encarregado", "")
    supervisor = st.session_state.get("nova_supervisor", "")
    data_bota = _parse_data_br(st.session_state.get("nova_data_bota", "")) if st.session_state.get("nova_data_bota") else None

    st.markdown("---")
    st.markdown("#### 📦 Itens")

    # Inicializa listas de itens no session_state (única vez por contexto)
    ctx_key = f"items_{edit_id or 'new'}_{tipo}"
    if st.session_state.get("compras_items_ctx_key") != ctx_key:
        if tipo == "Material" and edit_sol:
            st.session_state["compras_nova_itens_material"] = [
                {"material": i.get("material", ""), "qtd": i.get("qtd", 1), "valorUnit": i.get("valorUnit", 0)}
                for i in edit_sol.get("itens", [])
            ]
        elif tipo == "Material":
            st.session_state["compras_nova_itens_material"] = []
        elif tipo == "EPI" and edit_sol:
            st.session_state["compras_nova_itens_epi"] = [
                {"epi": i.get("epi", ""), "colaborador": i.get("colaborador", ""), "qtd": i.get("qtd", 1), "tamanho": i.get("tamanho", "")}
                for i in edit_sol.get("itens", [])
            ]
        elif tipo == "EPI":
            st.session_state["compras_nova_itens_epi"] = []
        st.session_state["compras_items_ctx_key"] = ctx_key

    if tipo == "Material":
        st.markdown("**Materiais**")
        materiais = lista_materiais_por_cliente(cliente)
        itens_mat = st.session_state.get("compras_nova_itens_material", [])

        if itens_mat:
            st.dataframe(pd.DataFrame(itens_mat), use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum material adicionado.")

        with st.form("add_mat_form", clear_on_submit=True):
            c1, c2, c3 = st.columns([3, 1, 1])
            with c1:
                mat_sel = st.selectbox("Material", materiais, key="add_mat_sel")
            with c2:
                qtd = st.number_input("Qtd", min_value=1, value=1, step=1, key="add_mat_qtd")
            with c3:
                val = st.number_input("Valor Unit.", min_value=0.0, value=0.0, step=0.01, format="%.2f", key="add_mat_val")
            if st.form_submit_button("➕ Adicionar"):
                itens_mat.append({"material": mat_sel, "qtd": qtd, "valorUnit": val})
                st.session_state["compras_nova_itens_material"] = itens_mat
            # Fragmento reexecuta automaticamente

        if itens_mat:
            idx = st.selectbox("Item para remover", range(len(itens_mat)),
                               format_func=lambda i: f"{itens_mat[i]['material']} (x{itens_mat[i]['qtd']})",
                               key="rem_mat_sel")
            if st.button("🗑️ Remover", key="rem_mat_btn"):
                itens_mat.pop(idx)
                st.session_state["compras_nova_itens_material"] = itens_mat
            # Fragmento reexecuta automaticamente

        itens_para_salvar = itens_mat

    elif tipo == "EPI":
        st.markdown("**EPIs**")
        epis = lista_epis_por_cliente(cliente) or lista_epis_por_cliente("Smart Fit")
        itens_epi = st.session_state.get("compras_nova_itens_epi", [])

        if itens_epi:
            st.dataframe(pd.DataFrame(itens_epi), use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum EPI adicionado.")

        with st.form("add_epi_form", clear_on_submit=True):
            c1, c2, c3, c4 = st.columns([2, 2, 1, 1])
            with c1:
                epi_sel = st.selectbox("EPI", epis, key="add_epi_sel")
            with c2:
                colab = st.text_input("Colaborador", key="add_epi_colab")
            with c3:
                qtd = st.number_input("Qtd", min_value=1, value=1, step=1, key="add_epi_qtd")
            with c4:
                tam = st.selectbox("Tamanho", TAMANHOS_EPI, key="add_epi_tam")
            if st.form_submit_button("➕ Adicionar"):
                itens_epi.append({"epi": epi_sel, "colaborador": colab, "qtd": qtd, "tamanho": tam})
                st.session_state["compras_nova_itens_epi"] = itens_epi
            # Fragmento reexecuta automaticamente

        if itens_epi:
            idx = st.selectbox("Item para remover", range(len(itens_epi)),
                               format_func=lambda i: f"{itens_epi[i]['epi']} (x{itens_epi[i]['qtd']})",
                               key="rem_epi_sel")
            if st.button("🗑️ Remover", key="rem_epi_btn"):
                itens_epi.pop(idx)
                st.session_state["compras_nova_itens_epi"] = itens_epi
            # Fragmento reexecuta automaticamente

        itens_para_salvar = itens_epi

    submitted = st.button("💾 Salvar Solicitação", type="primary")

    if submitted:
        itens = itens_para_salvar
        valor_total = 0
        if tipo == "Material":
            valor_total = sum(i.get("valorUnit", 0) * i.get("qtd", 0) for i in itens)
            if not itens:
                st.error("Adicione pelo menos um material.")
                return
        elif tipo == "EPI":
            if not itens:
                st.error("Adicione pelo menos um EPI.")
                return

        nova_sol = {
            "id": edit_sol["id"] if edit_sol else gerar_id(),
            "data": str(data_sol),
            "cliente": cliente,
            "loja": loja,
            "tipo": tipo,
            "solicitante": solicitante,
            "nomeFuncionario": nome_func if tipo == "EPI" else "",
            "encarregado": encarregado if tipo == "EPI" else "",
            "supervisor": supervisor if tipo == "EPI" else "",
            "dataUltimaBota": str(data_bota) if tipo == "EPI" and data_bota else "",
            "prioridade": prioridade,
            "previsao": str(previsao),
            "observacoes": observacoes,
            "itens": itens,
            "valorTotal": valor_total,
            "anexos": [],
            "status": edit_sol.get("status", "Pendente") if edit_sol else "Pendente",
            "dataCriacao": edit_sol.get("dataCriacao", datetime.now().isoformat()) if edit_sol else datetime.now().isoformat()
        }

        if tipo == "EPI":
            html_doc = gerar_doc_epi(nova_sol)
            nova_sol["anexos"].append({
                "nome": f'Solicitacao_de_EPI_{nova_sol["id"]}_{cliente}_{loja}.doc',
                "conteudo": html_doc,
                "tipo": "doc"
            })
        elif tipo == "Material":
            html_xls, filename_xls = gerar_xls_material(nova_sol)
            if html_xls:
                nova_sol["anexos"].append({
                    "nome": filename_xls,
                    "conteudo": html_xls,
                    "tipo": "xls"
                })

        sols = st.session_state["compras_solicitacoes"]
        if edit_sol:
            sols = [s for s in sols if s["id"] != edit_id]
        sols.append(nova_sol)
        st.session_state["compras_solicitacoes"] = sols

        if not edit_sol:
            st.session_state["compras_entregas"].append({
                "idSolicitacao": nova_sol["id"],
                "loja": loja,
                "tipo": tipo,
                "dataEnvio": "",
                "dataPrevista": str(previsao),
                "dataEntrega": "",
                "transportadora": "",
                "rastreio": "",
                "status": "Pendente",
                "observacoes": ""
            })

        # Limpa estado
        for k in ["nova_cliente", "nova_loja", "nova_tipo", "nova_solicitante",
                  "nova_data", "nova_prioridade", "nova_previsao", "nova_obs",
                  "nova_nome_func", "nova_encarregado", "nova_supervisor", "nova_data_bota",
                  "df_mat", "df_epi", "compras_items_ctx", "compras_edit_loaded",
                  "compras_nova_itens_material", "compras_nova_itens_epi",
                  "compras_items_ctx_key"]:
            if k in st.session_state:
                del st.session_state[k]
        _salvar_compras_automatico()
        st.success("Solicitação salva com sucesso!")
        st.session_state["compras_page"] = "Solicitações"
        if "nav_compras" in st.session_state:
            del st.session_state["nav_compras"]
        _rerun_fragment()

def page_detalhes():
    ver_id = st.session_state.get("compras_ver_id")
    s = None
    for sol in st.session_state["compras_solicitacoes"]:
        if sol["id"] == ver_id:
            s = sol
            break
    if not s:
        st.error("Solicitação não encontrada.")
        return

    st.markdown(f"### 👁️ Detalhes da Solicitação {s['id']}")

    col1, col2 = st.columns(2)
    with col1:
        st.write(f"**Data:** {formatar_data_br(s.get('data'))}")
        st.write(f"**Cliente:** {s.get('cliente','')}")
        st.write(f"**Loja:** {s.get('loja','')}")
        st.write(f"**Tipo:** {s.get('tipo','')}")
        st.write(f"**Solicitante:** {s.get('solicitante','')}")
    with col2:
        st.write(f"**Prioridade:** {s.get('prioridade','')}")
        st.write(f"**Status:** {get_badge_color(s.get('status'))} {s.get('status','')}")
        st.write(f"**Previsão:** {formatar_data_br(s.get('previsao'))}")
        st.write(f"**Valor Total:** {formatar_moeda(s.get('valorTotal',0))}")

    if s.get("observacoes"):
        st.write(f"**Observações:** {s['observacoes']}")

    if s.get("tipo") == "EPI":
        st.markdown("---")
        st.markdown("#### 👷 Informações EPI")
        st.write(f"**Funcionário:** {s.get('nomeFuncionario','-')}")
        st.write(f"**Encarregado:** {s.get('encarregado','-')}")
        st.write(f"**Supervisor:** {s.get('supervisor','-')}")
        st.write(f"**Data Última Bota:** {formatar_data_br(s.get('dataUltimaBota'))}")

    st.markdown("---")
    st.markdown("#### 📦 Itens")
    if s.get("itens"):
        if s["tipo"] == "Material":
            df = pd.DataFrame([{"Material": i.get("material",""), "Qtd": i.get("qtd",1), "Valor Unit.": formatar_moeda(i.get("valorUnit",0)), "Total": formatar_moeda(i.get("valorUnit",0)*i.get("qtd",0))} for i in s["itens"]])
        else:
            df = pd.DataFrame([{"EPI": i.get("epi",""), "Colaborador": i.get("colaborador",""), "Qtd": i.get("qtd",1), "Tamanho": i.get("tamanho","-")} for i in s["itens"]])
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhum item.")

    # Anexos
    if s.get("anexos"):
        st.markdown("---")
        st.markdown("#### 📎 Anexos Gerados")
        for a in s["anexos"]:
            st.download_button(
                label=f"📥 {a['nome']}",
                data=a["conteudo"].encode("utf-8"),
                file_name=a["nome"],
                mime="application/octet-stream",
                key=f"anexo_{a['nome']}"
            )

    st.button("🔙 Voltar", on_click=_cb_voltar_solicitacoes)


def page_entregas():
    st.markdown("### 🚚 Controle de Entregas")

    ents = st.session_state["compras_entregas"]

    # ── Filtros compactos ──
    with st.expander("🔍 Filtros", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            busca = st.text_input("Buscar", placeholder="ID, loja...", key="ent_busca")
        with c2:
            fstatus = st.selectbox("Status", ["Todos"] + STATUS_OPCOES, key="ent_fstatus")

    # ── Filtragem em memória ──
    filtradas = []
    for e in ents:
        txt = f"{e.get('idSolicitacao','')} {e.get('loja','')}".lower()
        if busca and busca.lower() not in txt:
            continue
        if fstatus != "Todos" and e.get("status") != fstatus:
            continue
        filtradas.append(e)

    if not filtradas:
        st.info("Nenhuma entrega encontrada.")
        return

    # ── Tabela única (st.dataframe) — ZERO widgets por linha ──
    df = _cached_df_entregas(json.dumps(filtradas, ensure_ascii=False, default=str))
    st.dataframe(df, use_container_width=True, hide_index=True)

    # ── Edição rápida (3 widgets fixos, independente da quantidade) ──
    st.markdown("---")
    st.markdown("#### ✏️ Editar Entrega")
    ids_disp = [e.get("idSolicitacao", "") for e in filtradas]
    c1, c2 = st.columns([3, 1])
    with c1:
        sel_id = st.selectbox("Selecionar entrega", ids_disp, key="ent_sel_id")
    with c2:
        if st.button("📝 Editar", use_container_width=True, key="ent_btn_edit"):
            st.session_state["compras_entrega_edit_id"] = sel_id
            # Fragmento reexecuta automaticamente

    # ── Formulário de edição inline (apenas 1 form, aberto sob demanda) ──
    edit_id = st.session_state.get("compras_entrega_edit_id")
    if edit_id and edit_id in ids_disp:
        e = next((x for x in filtradas if x.get("idSolicitacao") == edit_id), None)
        if e:
            st.markdown("---")
            with st.form("form_edit_entrega"):
                ne_status = st.selectbox("Status", STATUS_OPCOES, index=STATUS_OPCOES.index(e.get("status", "Pendente")) if e.get("status") in STATUS_OPCOES else 0)
                ne_transportadora = st.text_input("Transportadora", value=e.get("transportadora", ""))
                ne_rastreio = st.text_input("Rastreio", value=e.get("rastreio", ""))
                ne_data_envio_str = _iso_para_br(e.get("dataEnvio", ""))
                ne_data_envio_txt = st.text_input("Data Envio (DD/MM/AAAA)", value=ne_data_envio_str)
                ne_data_envio = _parse_data_br(ne_data_envio_txt) if ne_data_envio_txt.strip() else None
                ne_data_entrega_str = _iso_para_br(e.get("dataEntrega", ""))
                ne_data_entrega_txt = st.text_input("Data Entrega (DD/MM/AAAA)", value=ne_data_entrega_str)
                ne_data_entrega = _parse_data_br(ne_data_entrega_txt) if ne_data_entrega_txt.strip() else None
                ne_obs = st.text_area("Observações", value=e.get("observacoes", ""))
                c_btn1, c_btn2 = st.columns([1, 1])
                with c_btn1:
                    submitted = st.form_submit_button("💾 Salvar", use_container_width=True)
                with c_btn2:
                    if st.form_submit_button("❌ Cancelar", use_container_width=True):
                        st.session_state["compras_entrega_edit_id"] = None
            # Fragmento reexecuta automaticamente
                if submitted:
                    e["status"] = ne_status
                    e["transportadora"] = ne_transportadora
                    e["rastreio"] = ne_rastreio
                    e["dataEnvio"] = str(ne_data_envio) if ne_data_envio else ""
                    e["dataEntrega"] = str(ne_data_entrega) if ne_data_entrega else ""
                    e["observacoes"] = ne_obs
                    # Sincroniza status da solicitação
                    for s in st.session_state["compras_solicitacoes"]:
                        if s["id"] == e["idSolicitacao"]:
                            s["status"] = ne_status
                            break
                    _salvar_compras_automatico()
                    st.session_state["compras_entrega_edit_id"] = None
                    st.success("Entrega atualizada!")
            # Fragmento reexecuta automaticamente


def page_materiais():
    st.markdown("### 📋 Catálogo de Materiais e EPIs")

    tab_mat, tab_epi, tab_cad = st.tabs(["📦 Materiais", "👷 EPIs", "➕ Cadastrar Novo"])

    with tab_mat:
        fcliente = st.selectbox("Filtrar por Cliente", ["Todos"] + lista_clientes(), key="mat_fcliente_tab")
        dados = []
        for cliente, lista in lista_materiais_por_cliente().items():
            if fcliente != "Todos" and cliente != fcliente:
                continue
            for idx, nome in enumerate(lista, 1):
                dados.append({"ID": idx, "Nome": nome, "Cliente": cliente, "Categoria": "Material"})
        if dados:
            df = pd.DataFrame(dados)
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum material encontrado.")

    with tab_epi:
        fcliente_epi = st.selectbox("Filtrar por Cliente", ["Todos"] + lista_clientes(), key="epi_fcliente_tab")
        dados_epi = []
        for cliente, lista in lista_epis_por_cliente().items():
            if fcliente_epi != "Todos" and cliente != fcliente_epi:
                continue
            for idx, nome in enumerate(lista, 1):
                dados_epi.append({"ID": idx, "Nome": nome, "Cliente": cliente, "Categoria": "EPI"})
        if dados_epi:
            df_epi = pd.DataFrame(dados_epi)
            st.dataframe(df_epi, use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum EPI encontrado.")

    with tab_cad:
        st.markdown("#### ➕ Cadastrar Novo Item")
        tipo_cad = st.selectbox("Tipo", ["Material", "EPI"], key="cad_tipo")
        cliente_cad = st.selectbox("Cliente", lista_clientes(), key="cad_cliente")
        nome_cad = st.text_input("Nome do Item *", key="cad_nome")
        if st.button("💾 Salvar Item", type="primary", key="cad_salvar"):
            if not nome_cad.strip():
                st.error("Informe o nome do item.")
            else:
                if tipo_cad == "Material":
                    mats_existentes = lista_materiais_por_cliente(cliente_cad)
                    if nome_cad.strip() not in mats_existentes:
                        dados_salvar = carregar_dados()
                        nova_linha = {"Loja": "", "Cargo": "", "Cliente": cliente_cad, "Material": nome_cad.strip(), "EPI": ""}
                        # Garantir colunas existam
                        for col in ["Cliente", "Material", "EPI"]:
                            if col not in dados_salvar["Auxiliares"].columns:
                                dados_salvar["Auxiliares"][col] = ""
                        dados_salvar["Auxiliares"] = pd.concat([dados_salvar["Auxiliares"], pd.DataFrame([nova_linha])], ignore_index=True)
                        if salvar_dados(dados_salvar):
                            st.success(f"Material '{nome_cad.strip()}' cadastrado para {cliente_cad}!")
                        else:
                            st.error("❌ Não foi possível salvar os dados.")
                    else:
                        st.warning("Este material já está cadastrado para este cliente.")
                else:
                    epis_existentes = lista_epis_por_cliente(cliente_cad)
                    if nome_cad.strip() not in epis_existentes:
                        dados_salvar = carregar_dados()
                        nova_linha = {"Loja": "", "Cargo": "", "Cliente": cliente_cad, "Material": "", "EPI": nome_cad.strip()}
                        # Garantir colunas existam
                        for col in ["Cliente", "Material", "EPI"]:
                            if col not in dados_salvar["Auxiliares"].columns:
                                dados_salvar["Auxiliares"][col] = ""
                        dados_salvar["Auxiliares"] = pd.concat([dados_salvar["Auxiliares"], pd.DataFrame([nova_linha])], ignore_index=True)
                        if salvar_dados(dados_salvar):
                            st.success(f"EPI '{nome_cad.strip()}' cadastrado para {cliente_cad}!")
                        else:
                            st.error("❌ Não foi possível salvar os dados.")
                    else:
                        st.warning("Este EPI já está cadastrado para este cliente.")


def page_lojas():
    st.markdown("### 🏬 Lojas e Clientes")

    tab1, tab2, tab3 = st.tabs(["Lojas", "Clientes", "➕ Adicionar Cliente"])
    with tab1:
        dados = []
        for cliente in lista_clientes():
            for loja in lista_lojas_por_cliente(cliente):
                dados.append({"Loja": loja, "Cliente": cliente})
        df = pd.DataFrame(dados) if dados else pd.DataFrame(columns=["Loja", "Cliente"])
        st.dataframe(df, use_container_width=True, hide_index=True)
    with tab2:
        clientes_atuais = lista_clientes()
        df = pd.DataFrame([{"Cliente": c, "Qtd Lojas": len(lista_lojas_por_cliente(c)), "Qtd Materiais": len(lista_materiais_por_cliente(c)), "Qtd EPIs": len(lista_epis_por_cliente(c))} for c in clientes_atuais])
        st.dataframe(df, use_container_width=True, hide_index=True)
    with tab3:
        st.markdown("#### ➕ Adicionar Novo Cliente")
        novo_cliente = st.text_input("Nome do novo cliente:", key="novo_cliente_input")
        if st.button("💾 Cadastrar Cliente", key="btn_cad_cliente"):
            if novo_cliente.strip():
                clientes_existentes = lista_clientes()
                if novo_cliente.strip() not in clientes_existentes:
                    dados_salvar = carregar_dados()
                    nova_linha = {"Loja": "", "Cargo": "", "Cliente": novo_cliente.strip(), "Material": "", "EPI": ""}
                    for col in ["Cliente", "Material", "EPI"]:
                        if col not in dados_salvar["Auxiliares"].columns:
                            dados_salvar["Auxiliares"][col] = ""
                    dados_salvar["Auxiliares"] = pd.concat([dados_salvar["Auxiliares"], pd.DataFrame([nova_linha])], ignore_index=True)
                    if salvar_dados(dados_salvar):
                        st.success(f"Cliente '{novo_cliente.strip()}' cadastrado com sucesso!")
                        st.rerun()
                    else:
                        st.error("Não foi possível salvar os dados.")
                else:
                    st.warning("Este cliente já está cadastrado.")
            else:
                st.warning("Digite um nome válido para o cliente.")


def page_relatorios():
    st.markdown("### 📁 Relatórios e Downloads")

    st.markdown("#### 📝 Templates de Documentos")
    st.info("Os documentos são gerados automaticamente ao salvar uma solicitação.")

    # Listar solicitações com anexos
    sols_com_anexo = [s for s in st.session_state["compras_solicitacoes"] if s.get("anexos")]
    if sols_com_anexo:
        # ── Tabela única em vez de N containers + download buttons ──
        dados_anexos = []
        for s in sols_com_anexo:
            for a in s["anexos"]:
                dados_anexos.append({
                    "Solicitação": s["id"],
                    "Loja": s.get("loja", ""),
                    "Cliente": s.get("cliente", ""),
                    "Tipo": s.get("tipo", ""),
                    "Documento": a["nome"],
                    "Tipo Arq": a.get("tipo", "-"),
                })
        df_anexos = pd.DataFrame(dados_anexos)
        st.dataframe(df_anexos, use_container_width=True, hide_index=True)

        # ── Download único por seleção (3 widgets fixos) ──
        st.markdown("---")
        st.markdown("#### 📥 Baixar Documento")
        ids_disp = list(dict.fromkeys([s["id"] for s in sols_com_anexo]))
        c1, c2, c3 = st.columns([2, 3, 1])
        with c1:
            sel_id = st.selectbox("Solicitação", ids_disp, key="rel_sel_id")
        with c2:
            sol_sel = next((s for s in sols_com_anexo if s["id"] == sel_id), None)
            nomes_anexos = [a["nome"] for a in sol_sel["anexos"]] if sol_sel else []
            sel_doc = st.selectbox("Documento", nomes_anexos, key="rel_sel_doc") if nomes_anexos else None
        with c3:
            if sel_doc and st.button("📥 Baixar", use_container_width=True, key="rel_btn_down"):
                a = next((a for a in sol_sel["anexos"] if a["nome"] == sel_doc), None)
                if a:
                    st.download_button(
                        label=f"Confirmar download: {a['nome']}",
                        data=a["conteudo"].encode("utf-8"),
                        file_name=a["nome"],
                        mime="application/octet-stream",
                        key="rel_confirm_down"
                    )
    else:
        st.info("Nenhum documento gerado ainda. Crie uma solicitação para gerar documentos.")

    # Download dados completos
    st.markdown("---")
    st.markdown("#### 💾 Exportar Dados")
    if st.session_state["compras_solicitacoes"]:
        sols = st.session_state["compras_solicitacoes"]
        df = pd.DataFrame([
            {
                "ID": s["id"], "Data": s.get("data",""), "Loja": s.get("loja",""),
                "Cliente": s.get("cliente",""), "Tipo": s.get("tipo",""),
                "Solicitante": s.get("solicitante",""), "Status": s.get("status",""),
                "Valor Total": s.get("valorTotal",0)
            }
            for s in sols
        ])
        csv = df.to_csv(index=False, sep=";").encode("utf-8")
        st.download_button("📥 Exportar Solicitações (CSV)", data=csv, file_name="solicitacoes_completo.csv", mime="text/csv")


# ========== MAIN ==========
@st.fragment
def render_compras():
    init_session_state()

    st.subheader("🛒 SISTEMA DE COMPRAS E ENTREGAS")
    st.info("Módulo completo de controle de compras, solicitações e entregas.")

    # Menu lateral interno (dentro da aba, não na sidebar global)
    col_menu, col_conteudo = st.columns([1, 4])
    with col_menu:
        st.markdown("#### Navegação Compras")
        # "Detalhes" é uma tela secundária (acessada via 👁️) — NÃO aparece no menu.
        # Quando o usuário está em Detalhes, mostramos um botão Voltar em vez do radio.
        MENU_OPCOES = ["Dashboard", "Solicitações", "Nova Solicitação", "Controle de Entregas", "Catálogo de Materiais", "Lojas e Clientes", "Relatórios e Downloads"]
        if st.session_state["compras_page"] == "Detalhes":
            st.info("👁️ Visualizando detalhes")
            st.button("🔙 Voltar para Solicitações", use_container_width=True,
                      on_click=_cb_voltar_solicitacoes)
        else:
            menu = st.radio(
                "",
                MENU_OPCOES,
                index=MENU_OPCOES.index(st.session_state["compras_page"])
                if st.session_state["compras_page"] in MENU_OPCOES
                else 0,
                key="nav_compras",
            )
            if menu != st.session_state["compras_page"]:
                st.session_state["compras_page"] = menu
                st.session_state["compras_edit_id"] = None
                st.session_state["compras_edit_loaded"] = False
                _rerun_fragment()

    with col_conteudo:
        page = st.session_state["compras_page"]
        if page == "Dashboard":
            page_dashboard()
        elif page == "Solicitações":
            page_solicitacoes()
        elif page == "Nova Solicitação":
            page_nova_solicitacao()
        elif page == "Detalhes":
            page_detalhes()
        elif page == "Controle de Entregas":
            page_entregas()
        elif page == "Catálogo de Materiais":
            page_materiais()
        elif page == "Lojas e Clientes":
            page_lojas()
        elif page == "Relatórios e Downloads":
            page_relatorios()


# ====================== CONFIGURAÇÕES GERAIS ======================
# Diretório base fixo onde o script está (garante persistência entre reinícios)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Log de diagnóstico no stderr para debug de persistência
import sys
print(f"[DEBUG] BASE_DIR={BASE_DIR}", file=sys.stderr)
print(f"[DEBUG] ARQUIVO_COMPRAS={os.path.join(BASE_DIR, 'dados_compras.json')}", file=sys.stderr)

ARQUIVO = os.path.join(BASE_DIR, "dados_funcionarios.xlsx")
ARQUIVO_DIARIAS = os.path.join(BASE_DIR, "controle_diarias.xlsx")
ARQUIVO_VIAGENS = os.path.join(BASE_DIR, "registro_viagens.xlsx")
ARQUIVO_COMPRAS = os.path.join(BASE_DIR, "dados_compras.json")
PASTA_DOCS = os.path.join(BASE_DIR, "Documentos_Lojas")
PASTA_DOCS_FUNC = os.path.join(BASE_DIR, "Documentos_Funcionarios")
PASTA_FOTOS = os.path.join(BASE_DIR, "Fotos_Funcionarios")
PASTA_COMPROVANTES = os.path.join(BASE_DIR, "Comprovantes_Diarias")
os.makedirs(PASTA_DOCS, exist_ok=True)
os.makedirs(PASTA_DOCS_FUNC, exist_ok=True)
os.makedirs(PASTA_FOTOS, exist_ok=True)
os.makedirs(PASTA_COMPROVANTES, exist_ok=True)

# ====================== SISTEMA DE BACKUP AUTOMÁTICO ======================
PASTA_BACKUPS = os.path.join(BASE_DIR, "Backups_Auto")
os.makedirs(PASTA_BACKUPS, exist_ok=True)
MAX_BACKUPS = 50  # Mantém os últimos 50 backups
INTERVALO_BACKUP_MINUTOS = 30  # Faz backup automático a cada 30 minutos

ARQUIVOS_CRITICOS = [ARQUIVO, ARQUIVO_DIARIAS, ARQUIVO_VIAGENS, ARQUIVO_COMPRAS]


def _dados_estao_vazios(dados):
    """Verifica se os dados carregados estão vazios ou corrompidos."""
    if not dados or not isinstance(dados, dict):
        return True
    total_registros = 0
    for aba, df in dados.items():
        if isinstance(df, pd.DataFrame):
            total_registros += len(df)
    return total_registros == 0


def _diarias_estao_vazias(df):
    """Verifica se o DataFrame de diárias está vazio."""
    if df is None or not isinstance(df, pd.DataFrame):
        return True
    return len(df) == 0


def _viagens_estao_vazias(df):
    """Verifica se o DataFrame de viagens está vazio."""
    if df is None or not isinstance(df, pd.DataFrame):
        return True
    return len(df) == 0


def _compras_estao_vazias(solicitacoes, entregas):
    """Verifica se os dados de compras estão vazios."""
    sol_vazio = not solicitacoes or len(solicitacoes) == 0
    ent_vazio = not entregas or len(entregas) == 0
    return sol_vazio and ent_vazio


def criar_backup_automatico():
    """Cria um backup com timestamp de todos os arquivos críticos."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pasta_backup = os.path.join(PASTA_BACKUPS, f"backup_{timestamp}")
    os.makedirs(pasta_backup, exist_ok=True)
    
    arquivos_backupeados = []
    for arq in ARQUIVOS_CRITICOS:
        if os.path.exists(arq) and os.path.getsize(arq) > 0:
            destino = os.path.join(pasta_backup, os.path.basename(arq))
            shutil.copy2(arq, destino)
            arquivos_backupeados.append(os.path.basename(arq))
    
    # Também faz backup das pastas de documentos (fotos, comprovantes, docs)
    for pasta_origem in [PASTA_DOCS, PASTA_DOCS_FUNC, PASTA_FOTOS, PASTA_COMPROVANTES]:
        if os.path.exists(pasta_origem):
            nome_pasta = os.path.basename(pasta_origem)
            pasta_destino = os.path.join(pasta_backup, nome_pasta)
            os.makedirs(pasta_destino, exist_ok=True)
            for root, dirs, files in os.walk(pasta_origem):
                for file in files:
                    src = os.path.join(root, file)
                    rel = os.path.relpath(src, pasta_origem)
                    dst = os.path.join(pasta_destino, rel)
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    shutil.copy2(src, dst)
    
    # Limpa backups antigos (mantém apenas os MAX_BACKUPS mais recentes)
    try:
        backups = sorted([d for d in os.listdir(PASTA_BACKUPS) if d.startswith("backup_")])
        while len(backups) > MAX_BACKUPS:
            pasta_antiga = os.path.join(PASTA_BACKUPS, backups.pop(0))
            shutil.rmtree(pasta_antiga, ignore_errors=True)
    except Exception:
        pass
    
    return pasta_backup, arquivos_backupeados


def listar_backups_automaticos():
    """Retorna lista de backups automáticos ordenados do mais recente ao mais antigo."""
    if not os.path.exists(PASTA_BACKUPS):
        return []
    backups = [d for d in os.listdir(PASTA_BACKUPS) if d.startswith("backup_")]
    backups.sort(reverse=True)
    return backups


def restaurar_backup_automatico(nome_backup):
    """Restaura todos os arquivos críticos a partir de um backup automático."""
    pasta_backup = os.path.join(PASTA_BACKUPS, nome_backup)
    if not os.path.exists(pasta_backup):
        return False, "Pasta de backup não encontrada."
    
    restaurados = []
    erros = []
    
    # Restaura arquivos principais
    for arq in ARQUIVOS_CRITICOS:
        nome = os.path.basename(arq)
        src = os.path.join(pasta_backup, nome)
        if os.path.exists(src):
            try:
                shutil.copy2(src, arq)
                restaurados.append(nome)
            except Exception as e:
                erros.append(f"{nome}: {e}")
    
    # Restaura pastas de documentos
    for pasta_nome in ["Documentos_Lojas", "Documentos_Funcionarios", "Fotos_Funcionarios", "Comprovantes_Diarias"]:
        src_pasta = os.path.join(pasta_backup, pasta_nome)
        if os.path.exists(src_pasta):
            dst_pasta = os.path.join(BASE_DIR, pasta_nome)
            os.makedirs(dst_pasta, exist_ok=True)
            for root, dirs, files in os.walk(src_pasta):
                for file in files:
                    src = os.path.join(root, file)
                    rel = os.path.relpath(src, src_pasta)
                    dst = os.path.join(dst_pasta, rel)
                    os.makedirs(os.path.dirname(dst), exist_ok=True)
                    try:
                        shutil.copy2(src, dst)
                    except Exception as e:
                        erros.append(f"{pasta_nome}/{rel}: {e}")
            if pasta_nome not in restaurados:
                restaurados.append(pasta_nome)
    
    if restaurados:
        st.cache_data.clear()
        return True, f"Restaurado: {', '.join(restaurados)}"
    else:
        return False, "Nenhum arquivo restaurado."


def backup_automatico_periodico():
    """Verifica se já passou o intervalo desde o último backup e cria um novo se necessário."""
    try:
        agora = datetime.now()
        chave_ultimo_backup = "_ultimo_backup_periodico"
        chave_ultimo_notificado = "_ultimo_backup_notificado"
        ultimo_backup = st.session_state.get(chave_ultimo_backup)
        
        if ultimo_backup is not None:
            # Verifica se já passou o intervalo
            minutos_desde_ultimo = (agora - ultimo_backup).total_seconds() / 60.0
            if minutos_desde_ultimo < INTERVALO_BACKUP_MINUTOS:
                return  # Ainda não passou o tempo, não faz nada
        
        # Cria o backup e atualiza o timestamp
        pasta_bkp, arqs = criar_backup_automatico()
        st.session_state[chave_ultimo_backup] = agora
        
        # Notifica o usuário apenas se este backup ainda não foi notificado
        ultimo_notificado = st.session_state.get(chave_ultimo_notificado)
        if ultimo_notificado != pasta_bkp:
            st.toast(f"🛡️ Backup automático criado! ({len(arqs)} arquivos protegidos)", icon="💾")
            st.session_state[chave_ultimo_notificado] = pasta_bkp
        
        # Mostra no log também
        print(f"[BACKUP PERIÓDICO] Backup criado em: {pasta_bkp} | Arquivos: {arqs}")
    except Exception as e:
        print(f"[BACKUP PERIÓDICO] Erro ao criar backup: {e}")


def _tentar_restaurar_dados_vazios():
    """Tenta restaurar dados automaticamente de um backup se detectar dados vazios."""
    backups = listar_backups_automaticos()
    if not backups:
        return False
    for nome_backup in backups:
        sucesso, msg = restaurar_backup_automatico(nome_backup)
        if sucesso:
            return True
    return False


MESES = ["Todos", "jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]

# ====================== FUNÇÕES DE BUSCA INTELIGENTE ======================
import unicodedata

def normalizar_texto(texto):
    """Remove acentos, converte para maiúsculas e remove espaços extras."""
    if pd.isna(texto):
        return ""
    texto = str(texto).strip()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    texto = ' '.join(texto.split())  # remove espaços duplos
    return texto.upper()

def busca_palavras(serie, texto_busca):
    """
    Verifica se TODAS as palavras do texto_busca estão contidas em cada valor da série.
    Ex: buscar 'CARLOS IVAN' encontra 'CARLOS IVAN SILVA VAZ'.
    Ignora acentos, case e espaços extras.
    """
    texto_busca = normalizar_texto(texto_busca)
    palavras = texto_busca.split()
    if not palavras:
        return pd.Series([True] * len(serie), index=serie.index)
    serie_norm = serie.apply(normalizar_texto)
    mascara = pd.Series([True] * len(serie), index=serie.index)
    for palavra in palavras:
        mascara = mascara & serie_norm.str.contains(palavra, case=False, regex=False, na=False)
    return mascara

SEMANAS = ["Todas", "1º Semana", "2º Semana", "3º Semana", "4º Semana"]
SITUACOES_DIARIA = ["Todas", "FALTA ENVIAR AO FINANCEIRO", "ENVIADO/PENDENTE", "PAGO"]
ANOS = [str(a) for a in range(2020, datetime.now().year + 2)]

SITUACOES = [
    "Ativo", "Pré-cadastro", "Abandono", "Desistente", "Término de Contrato",
    "Demitido S/JC", "Demitido C/JC", "Pedido de Conta",
    "Rescisão Indireta", "Férias", "Doença", "Acidente", "Maternidade"
]

# ====================== GOOGLE SHEETS HELPERS ======================
def _gsheet_to_df(worksheet):
    """Converte uma worksheet do gspread para DataFrame."""
    try:
        records = worksheet.get_all_records()
        if not records:
            return pd.DataFrame()
        df = pd.DataFrame(records)
        df = df.astype(str)
        df = df.replace("nan", "")
        df = df.replace("None", "")
        return df
    except Exception:
        return pd.DataFrame()

def _df_to_gsheet(df, worksheet):
    """Sobrescreve uma worksheet do gspread com os dados de um DataFrame."""
    worksheet.clear()
    if df.empty:
        worksheet.update([df.columns.tolist()])
        return
    data = [df.columns.tolist()] + df.fillna("").astype(str).values.tolist()
    # Garante que não ultrapasse limites do Google Sheets
    if len(data) > 1000:
        data = data[:1000]
    worksheet.update(data)

def _garantir_abas_gs(spreadsheet, abas_necessarias, padrao_cols):
    """Garante que todas as abas existam na planilha do Google Sheets."""
    abas_existentes = {ws.title: ws for ws in spreadsheet.worksheets()}
    for aba_nome, cols in abas_necessarias.items():
        if aba_nome not in abas_existentes:
            spreadsheet.add_worksheet(title=aba_nome, rows=1000, cols=len(cols))
            ws = spreadsheet.worksheet(aba_nome)
            ws.update([cols])

def _carregar_dados_gs():
    """Carrega dados do Google Sheets."""
    spreadsheet = gc.open_by_key(GS_ID_FUNCIONARIOS)
    abas = {ws.title: ws for ws in spreadsheet.worksheets()}
    dados = {}
    for aba_nome, ws in abas.items():
        dados[aba_nome] = _gsheet_to_df(ws)
    return dados

def _salvar_dados_gs(dados):
    """Salva dados no Google Sheets."""
    spreadsheet = gc.open_by_key(GS_ID_FUNCIONARIOS)
    for aba_nome, df in dados.items():
        try:
            ws = spreadsheet.worksheet(aba_nome)
        except gspread.exceptions.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title=aba_nome, rows=1000, cols=len(df.columns) if not df.empty else 10)
        # Garante que o DataFrame tenha todas as colunas esperadas antes de salvar
        if not df.empty and ws.row_count > 0:
            try:
                header_row = ws.row_values(1)
                if header_row:
                    for col in header_row:
                        if col not in df.columns:
                            df[col] = ""
                    df = df[header_row]
            except Exception:
                pass
        _df_to_gsheet(df, ws)

def _carregar_diarias_gs():
    """Carrega diárias do Google Sheets."""
    spreadsheet = gc.open_by_key(GS_ID_DIARIAS)
    ws = spreadsheet.sheet1
    return _gsheet_to_df(ws)

def _salvar_diarias_gs(df):
    """Salva diárias no Google Sheets."""
    spreadsheet = gc.open_by_key(GS_ID_DIARIAS)
    ws = spreadsheet.sheet1
    _df_to_gsheet(df, ws)


# ====================== COMPRAS - GOOGLE SHEETS ======================
def _normalizar_data_iso(valor):
    """Converte uma data de qualquer formato conhecido para ISO (YYYY-MM-DD)."""
    if not valor or str(valor).strip() == "":
        return ""
    s = str(valor).strip()
    # Se já está no formato ISO, retorna como está
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    # Tenta DD/MM/YYYY (prioridade para formato brasileiro)
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        try:
            dt = datetime.strptime(s, "%d/%m/%Y")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    # Tenta MM/DD/YYYY (formato americano) — somente se DD/MM falhou
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        try:
            dt = datetime.strptime(s, "%m/%d/%Y")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    # Tenta YYYY/MM/DD
    if re.match(r"^\d{4}/\d{2}/\d{2}$", s):
        try:
            dt = datetime.strptime(s, "%Y/%m/%d")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    # Se não conseguiu converter, retorna o valor original
    return s


def _iso_para_br(valor):
    """Converte data ISO (YYYY-MM-DD) para formato brasileiro (DD/MM/YYYY)."""
    if not valor or str(valor).strip() == "":
        return ""
    s = str(valor).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            dt = datetime.strptime(s, "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        except Exception:
            pass
    return s


def _parse_data_br(texto):
    """Converte texto DD/MM/YYYY para objeto date. Retorna date.today() se inválido."""
    if not texto or str(texto).strip() == "":
        return date.today()
    s = str(texto).strip()
    # Se já vem no formato ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            pass
    # Formato brasileiro DD/MM/YYYY
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        try:
            return datetime.strptime(s, "%d/%m/%Y").date()
        except Exception:
            pass
    return date.today()


def _solicitacao_to_row(s):
    """Converte dict de solicitação em linha para o GS."""
    return {
        "id": s.get("id", ""),
        "data": _iso_para_br(s.get("data", "")),
        "cliente": s.get("cliente", ""),
        "loja": s.get("loja", ""),
        "tipo": s.get("tipo", ""),
        "solicitante": s.get("solicitante", ""),
        "nomeFuncionario": s.get("nomeFuncionario", ""),
        "encarregado": s.get("encarregado", ""),
        "supervisor": s.get("supervisor", ""),
        "dataUltimaBota": _iso_para_br(s.get("dataUltimaBota", "")),
        "prioridade": s.get("prioridade", ""),
        "previsao": _iso_para_br(s.get("previsao", "")),
        "observacoes": s.get("observacoes", ""),
        "itens": json.dumps(s.get("itens", []), ensure_ascii=False),
        "valorTotal": s.get("valorTotal", 0),
        "anexos": json.dumps(s.get("anexos", []), ensure_ascii=False),
        "status": s.get("status", "Pendente"),
        "dataCriacao": _iso_para_br(s.get("dataCriacao", ""))
    }

def _row_to_solicitacao(row):
    """Converte linha do GS em dict de solicitação."""
    s = dict(row)
    try:
        s["itens"] = json.loads(s.get("itens", "[]")) if s.get("itens") else []
    except Exception:
        s["itens"] = []
    try:
        s["anexos"] = json.loads(s.get("anexos", "[]")) if s.get("anexos") else []
    except Exception:
        s["anexos"] = []
    try:
        s["valorTotal"] = float(s.get("valorTotal", 0)) if s.get("valorTotal") else 0
    except Exception:
        s["valorTotal"] = 0
    # Normaliza campos de data para ISO
    for campo_data in ["data", "previsao", "dataUltimaBota", "dataCriacao"]:
        if campo_data in s:
            s[campo_data] = _normalizar_data_iso(s.get(campo_data, ""))
    return s

def _entrega_to_row(e):
    """Converte dict de entrega em linha para o GS."""
    return {
        "idSolicitacao": e.get("idSolicitacao", ""),
        "loja": e.get("loja", ""),
        "tipo": e.get("tipo", ""),
        "dataEnvio": _iso_para_br(e.get("dataEnvio", "")),
        "dataPrevista": _iso_para_br(e.get("dataPrevista", "")),
        "dataEntrega": _iso_para_br(e.get("dataEntrega", "")),
        "transportadora": e.get("transportadora", ""),
        "rastreio": e.get("rastreio", ""),
        "status": e.get("status", "Pendente")
    }

def _row_to_entrega(row):
    """Converte linha do GS em dict de entrega."""
    e = dict(row)
    for campo_data in ["dataEnvio", "dataPrevista", "dataEntrega"]:
        if campo_data in e:
            e[campo_data] = _normalizar_data_iso(e.get(campo_data, ""))
    return e

@st.cache_data(ttl=30, show_spinner=False)
def _carregar_compras_gs():
    """Carrega dados de compras do Google Sheets."""
    spreadsheet = gc.open_by_key(GS_ID_COMPRAS)
    abas = {ws.title: ws for ws in spreadsheet.worksheets()}
    solicitacoes = []
    entregas = []
    if "Solicitacoes" in abas:
        df = _gsheet_to_df(abas["Solicitacoes"])
        for _, row in df.iterrows():
            solicitacoes.append(_row_to_solicitacao(row))
    if "Entregas" in abas:
        df = _gsheet_to_df(abas["Entregas"])
        for _, row in df.iterrows():
            entregas.append(_row_to_entrega(row))
    return solicitacoes, entregas

def _salvar_compras_gs(solicitacoes, entregas):
    """Salva dados de compras no Google Sheets."""
    spreadsheet = gc.open_by_key(GS_ID_COMPRAS)
    abas = {ws.title: ws for ws in spreadsheet.worksheets()}

    # Aba Solicitacoes
    if "Solicitacoes" not in abas:
        spreadsheet.add_worksheet(title="Solicitacoes", rows=1000, cols=20)
    ws_sol = spreadsheet.worksheet("Solicitacoes")
    if solicitacoes:
        df_sol = pd.DataFrame([_solicitacao_to_row(s) for s in solicitacoes])
    else:
        df_sol = pd.DataFrame(columns=[
            "id","data","cliente","loja","tipo","solicitante","nomeFuncionario",
            "encarregado","supervisor","dataUltimaBota","prioridade","previsao",
            "observacoes","itens","valorTotal","anexos","status","dataCriacao"
        ])
    _df_to_gsheet(df_sol, ws_sol)

    # Aba Entregas
    if "Entregas" not in abas:
        spreadsheet.add_worksheet(title="Entregas", rows=1000, cols=10)
    ws_ent = spreadsheet.worksheet("Entregas")
    if entregas:
        df_ent = pd.DataFrame([_entrega_to_row(e) for e in entregas])
    else:
        df_ent = pd.DataFrame(columns=[
            "idSolicitacao","loja","tipo","dataEnvio","dataPrevista",
            "dataEntrega","transportadora","rastreio","status"
        ])
    _df_to_gsheet(df_ent, ws_ent)


if GS_ENABLED:
    try:
        # Garante abas na planilha de funcionários
        padrao_func = {
            "Base_Dados": [
                "Matricula","Nome","CPF","RG","PIS","Nascimento","Admissao",
                "Telefone","Endereco","Loja","Cargo","Salario","Situacao",
                "DataAvisoPrevio","DiasAvisoPrevio","DataTerminoAviso",
                "DataFeriasInicio","DiasFerias","DataRetornoFerias",
                "DataPedidoConta","DataRescisao","DataAbandono","DataDesistencia",
                "DataTerminoContrato",
                "DataLicenca","DiasLicenca","DataTerminoLicenca",
                "DataAfastamento","DiasAfastamento","DataRetornoAfastamento",
                "CaminhoFoto"
            ],
            "Historico": [
                "DataEvento","TipoEvento","Matricula","Nome","CPF","RG","PIS",
                "Nascimento","Admissao","Telefone","Endereco","Loja","Cargo",
                "Salario","Situacao","DataAvisoPrevio","DiasAvisoPrevio","DataTerminoAviso",
                "DataFeriasInicio","DiasFerias","DataRetornoFerias",
                "DataPedidoConta","DataRescisao","DataAbandono","DataDesistencia",
                "DataTerminoContrato",
                "DataLicenca","DiasLicenca","DataTerminoLicenca",
                "DataAfastamento","DiasAfastamento","DataRetornoAfastamento","Detalhes"
            ],
            "Auxiliares": ["Loja", "Cargo", "Cliente", "Material", "EPI"],
            "Docs_Lojas": ["Loja","Mes","Ano","NomeArquivo","Caminho","DataAnexado","Responsavel"],
            "Docs_Funcionarios": ["Matricula","Nome","TipoDoc","NomeArquivo","Caminho","DataAnexado"]
        }
        spreadsheet = gc.open_by_key(GS_ID_FUNCIONARIOS)
        abas_existentes = {ws.title for ws in spreadsheet.worksheets()}
        for aba_nome, cols in padrao_func.items():
            if aba_nome not in abas_existentes:
                ws = spreadsheet.add_worksheet(title=aba_nome, rows=1000, cols=len(cols))
                ws.update([cols])
    except Exception:
        pass

# ====================== BANCO DE DADOS ======================
@st.cache_data(ttl=0, show_spinner=False)
def carregar_dados():
    dados = {}
    # Tenta carregar do Google Sheets primeiro
    if GS_ENABLED:
        try:
            dados = _carregar_dados_gs()
            # Também salva localmente como cache/fallback
            try:
                with pd.ExcelWriter(ARQUIVO, engine="openpyxl", mode="w") as f:
                    for aba, df in dados.items():
                        df.to_excel(f, sheet_name=aba, index=False)
            except Exception:
                pass
        except Exception as e:
            st.warning(f"⚠️ Erro ao carregar do Google Sheets: {e}. Usando arquivo local.")
            try:
                dados = pd.read_excel(ARQUIVO, sheet_name=None, dtype=str, keep_default_na=False)
            except:
                dados = {}
    else:
        try:
            dados = pd.read_excel(ARQUIVO, sheet_name=None, dtype=str, keep_default_na=False)
        except:
            dados = {}
    
    # PROTEÇÃO: Se dados estiverem vazios/corrompidos, tenta restaurar do backup automático
    if _dados_estao_vazios(dados):
        st.warning("⚠️ Dados principais parecem vazios ou corrompidos. Tentando restaurar do backup automático...")
        if _tentar_restaurar_dados_vazios():
            try:
                dados = pd.read_excel(ARQUIVO, sheet_name=None, dtype=str, keep_default_na=False)
                st.success("✅ Dados restaurados automaticamente do backup mais recente!")
            except:
                dados = {}
        else:
            st.error("❌ Não foi possível restaurar automaticamente. Verifique a aba 'BACKUP / RESTAURAÇÃO' para restaurar manualmente.")
    
    padrao = {
        "Base_Dados": [
            "Matricula","Nome","CPF","RG","PIS","Nascimento","Admissao",
            "Telefone","Endereco","Loja","Cargo","Salario","Situacao",
            "DataAvisoPrevio","DiasAvisoPrevio","DataTerminoAviso",
            "DataFeriasInicio","DiasFerias","DataRetornoFerias",
            "DataPedidoConta","DataRescisao","DataAbandono","DataDesistencia",
            "DataTerminoContrato",
            "DataLicenca","DiasLicenca","DataTerminoLicenca",
            "DataAfastamento","DiasAfastamento","DataRetornoAfastamento",
            "TipoAfastamento","CaminhoFoto"
        ],
        "Historico": [
            "DataEvento","TipoEvento","Matricula","Nome","CPF","RG","PIS",
            "Nascimento","Admissao","Telefone","Endereco","Loja","Cargo",
            "Salario","Situacao","DataAvisoPrevio","DiasAvisoPrevio","DataTerminoAviso",
            "DataFeriasInicio","DiasFerias","DataRetornoFerias",
            "DataPedidoConta","DataRescisao","DataAbandono","DataDesistencia",
            "DataTerminoContrato",
            "DataLicenca","DiasLicenca","DataTerminoLicenca",
            "DataAfastamento","DiasAfastamento","DataRetornoAfastamento","Detalhes"
        ],
        "Auxiliares": ["Loja", "Cargo", "Cliente", "Material", "EPI"],
        "Docs_Lojas": ["Loja","Mes","Ano","NomeArquivo","Caminho","DataAnexado","Responsavel"],
        "Docs_Funcionarios": ["Matricula","Nome","TipoDoc","NomeArquivo","Caminho","DataAnexado"]
    }
    
    for aba, cols in padrao.items():
        if aba not in dados:
            dados[aba] = pd.DataFrame(columns=cols)
        else:
            for c in cols:
                if c not in dados[aba].columns:
                    dados[aba][c] = ""
            if "Matricula" in dados[aba].columns:
                dados[aba]["Matricula"] = dados[aba]["Matricula"].astype(str).str.strip()
            if "Situacao" in dados[aba].columns:
                dados[aba]["Situacao"] = dados[aba]["Situacao"].astype(str).str.strip()
    
    # SEED: Popular colunas Cliente/Material/EPI do Auxiliares com dados dos fallbacks
    # se estiverem todas vazias (primeira execução com colunas novas)
    if "Auxiliares" in dados:
        aux = dados["Auxiliares"]
        precisa_seed = False
        for col in ["Cliente", "Material", "EPI"]:
            if col in aux.columns:
                if aux[col].astype(str).str.strip().replace("", None).dropna().empty:
                    precisa_seed = True
                    break
            else:
                precisa_seed = True
                break
        if precisa_seed:
            linhas_seed = []
            for cliente in CLIENTES_FALLBACK:
                # Cria linha com nome do cliente
                linhas_seed.append({"Loja": "", "Cargo": "", "Cliente": cliente, "Material": "", "EPI": ""})
                # Lojas para este cliente
                for loja in LOJAS_POR_CLIENTE_FALLBACK.get(cliente, []):
                    linhas_seed.append({"Loja": loja, "Cargo": "", "Cliente": cliente, "Material": "", "EPI": ""})
                # Materiais para este cliente
                for mat in MATERIAIS_POR_CLIENTE_FALLBACK.get(cliente, []):
                    linhas_seed.append({"Loja": "", "Cargo": "", "Cliente": cliente, "Material": mat, "EPI": ""})
                # EPIs para este cliente
                for epi in EPIS_POR_CLIENTE_FALLBACK.get(cliente, []):
                    linhas_seed.append({"Loja": "", "Cargo": "", "Cliente": cliente, "Material": "", "EPI": epi})
            if linhas_seed:
                df_seed = pd.DataFrame(linhas_seed)
                for col in ["Cliente", "Material", "EPI"]:
                    if col not in aux.columns:
                        aux[col] = ""
                dados["Auxiliares"] = pd.concat([aux, df_seed], ignore_index=True)
                # Salvar seed no arquivo para persistir
                try:
                    with pd.ExcelWriter(ARQUIVO, engine="openpyxl", mode="w") as f:
                        for aba_nome, df_abas in dados.items():
                            df_abas.to_excel(f, sheet_name=aba_nome, index=False)
                except Exception:
                    pass
    
    return dados

@st.cache_data(ttl=0, show_spinner=False)
def carregar_diarias():
    cols_padrao = [
        "LOJA","NOME COLABORADOR","CPF","DATA EXECUCAO","QTDE DE DIARIAS","VALOR UNITARIO","VALOR TOTAL",
        "DADOS BANCÁRIOS","SUBSTITUICAO","MOTIVO","DATA PAGAMENTO","SITUACAO","MES","SEMANA","ANO",
        "CARGO","DATA CADASTRO","COMPROVANTE","OBSERVACAO"
    ]
    
    # Tenta carregar do Google Sheets primeiro
    df = None
    if GS_ENABLED:
        try:
            df = _carregar_diarias_gs()
            # Salva local como fallback
            try:
                df.to_excel(ARQUIVO_DIARIAS, index=False, engine="openpyxl")
            except Exception:
                pass
        except Exception as e:
            st.warning(f"⚠️ Erro ao carregar diárias do Google Sheets: {e}. Usando arquivo local.")
    
    if df is None:
        if not os.path.exists(ARQUIVO_DIARIAS):
            return pd.DataFrame(columns=cols_padrao)
    
    # Mapeamento de colunas da planilha externa para o padrão do app
    rename_map = {
        'NOME COMPLETO DO COLABORADOR': 'NOME COLABORADOR',
        'QUANT.': 'QTDE DE DIARIAS',
        'VALOR UNI.': 'VALOR UNITARIO',
        'TOTAL': 'VALOR TOTAL',
        'MOTIVO DA DIARIA': 'MOTIVO',
        'situação': 'SITUACAO',
        'Mês': 'MES',
        'semana': 'SEMANA',
        'DATA DA EXECUÇÃO': 'DATA EXECUCAO',
        'SUBSTITUIÇÃO': 'SUBSTITUICAO',
        'DATA DE PAGAM.': 'DATA PAGAMENTO'
    }
    
    # Se já carregou do Google Sheets, verifica se está no formato do app
    if df is not None:
        colunas_encontradas = [c for c in cols_padrao if c in df.columns]
        if len(colunas_encontradas) >= 3:
            # Já está no formato do app, retorna
            for c in cols_padrao:
                if c not in df.columns:
                    df[c] = ""
            return df
        # Caso contrário, processa como planilha externa
    
    if df is None:
        # ESTRATÉGIA 1: Tenta header=0 (formato do app - cabeçalho na 1ª linha)
        try:
            df_test = pd.read_excel(ARQUIVO_DIARIAS, header=0, dtype=str, keep_default_na=False)
            colunas_encontradas = [c for c in cols_padrao if c in df_test.columns]
            if len(colunas_encontradas) >= 3:
                df = df_test
        except Exception:
            pass
        
        # ESTRATÉGIA 2: Tenta header=1 (formato da planilha do usuário com instruções na 1ª linha)
        if df is None:
            try:
                df_test = pd.read_excel(ARQUIVO_DIARIAS, header=1, dtype=str, keep_default_na=False)
                df_test = df_test.rename(columns=rename_map)
                colunas_encontradas = [c for c in cols_padrao if c in df_test.columns]
                if len(colunas_encontradas) >= 3:
                    df = df_test
            except Exception:
                pass
    
    # ESTRATÉGIA 3: Último recurso - tenta ler de qualquer jeito
    if df is None:
        try:
            df = pd.read_excel(ARQUIVO_DIARIAS, dtype=str, keep_default_na=False)
        except Exception:
            df = pd.DataFrame(columns=cols_padrao)
    
    if df is None:
        df = pd.DataFrame(columns=cols_padrao)

    # Garante que todas as colunas padrão existam
    for col in cols_padrao:
        if col not in df.columns:
            df[col] = ""

    # Extrai ANO da DATA PAGAMENTO quando estiver vazio
    for i in df.index:
        if str(df.at[i, "ANO"]).strip() == "":
            try:
                dt = pd.to_datetime(str(df.at[i, "DATA PAGAMENTO"]).strip())
                df.at[i, "ANO"] = str(dt.year)
            except Exception:
                df.at[i, "ANO"] = str(datetime.now().year)

    # Limpa strings
    for col in ["NOME COLABORADOR", "CPF", "LOJA", "MOTIVO", "SITUACAO", "MES", "SEMANA"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # Garante ordem correta das colunas
    df = df[[c for c in cols_padrao if c in df.columns]]
    
    # PROTEÇÃO: Se diárias estiverem vazias/corrompidas, tenta restaurar do backup automático
    if _diarias_estao_vazias(df):
        st.warning("⚠️ Dados de diárias parecem vazios ou corrompidos. Tentando restaurar do backup automático...")
        if _tentar_restaurar_dados_vazios():
            try:
                df_restaurado = pd.read_excel(ARQUIVO_DIARIAS, header=0, dtype=str, keep_default_na=False)
                colunas_encontradas = [c for c in cols_padrao if c in df_restaurado.columns]
                if len(colunas_encontradas) >= 3:
                    df = df_restaurado
                    for c in cols_padrao:
                        if c not in df.columns:
                            df[c] = ""
                    st.success("✅ Diárias restauradas automaticamente do backup mais recente!")
            except Exception:
                pass
    
    return df

def salvar_dados(dados):
    # BACKUP AUTOMÁTICO antes de salvar
    try:
        pasta_bkp, arqs = criar_backup_automatico()
    except Exception:
        pass
    
    sucesso = False
    # Sempre salva localmente como fallback
    try:
        with pd.ExcelWriter(ARQUIVO, engine="openpyxl", mode="w") as f:
            for aba, df in dados.items():
                df.to_excel(f, sheet_name=aba, index=False)
        sucesso = True
    except Exception as e:
        st.error(f"❌ Erro ao salvar arquivo local: {e}")
    
    # Se Google Sheets ativo, salva na nuvem também
    if GS_ENABLED:
        try:
            _salvar_dados_gs(dados)
        except Exception as e:
            st.warning(f"⚠️ Erro ao salvar no Google Sheets: {e}")
    
    st.cache_data.clear()
    return sucesso

def salvar_diarias(df_diarias):
    # BACKUP AUTOMÁTICO antes de salvar
    try:
        pasta_bkp, arqs = criar_backup_automatico()
    except Exception:
        pass
    
    # Sempre salva localmente como fallback
    try:
        with pd.ExcelWriter(ARQUIVO_DIARIAS, engine="openpyxl", mode="w") as f:
            df_diarias.to_excel(f, sheet_name="Diarias", index=False)
    except Exception:
        pass
    
    # Se Google Sheets ativo, salva na nuvem também
    if GS_ENABLED:
        try:
            _salvar_diarias_gs(df_diarias)
        except Exception as e:
            st.warning(f"⚠️ Erro ao salvar diárias no Google Sheets: {e}")
    
    st.cache_data.clear()

def exportar_diarias_formatado(df, caminho):
    """Exporta DataFrame de diárias para Excel com a mesma formatação da planilha padrão."""
    df_export = df.copy()
    df_export.to_excel(caminho, index=False, engine="openpyxl")
    try:
        wb = load_workbook(caminho)
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.append(list(df_export.columns))
        for _, row in df_export.iterrows():
            ws.append(list(row))
        wb.save(caminho)
    ws = wb.active
    ws.title = "Diarias"

    # Cores e estilos exatos da planilha padrão
    fill_instrucoes = PatternFill(start_color="1B2D4F", end_color="1B2D4F", fill_type="solid")
    font_instrucoes = Font(name="Calibri", size=11, bold=False, color="FFFFFF")
    align_instrucoes = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fill_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_header = Alignment(horizontal="center", vertical="center")

    borda = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    fill_pendente = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_pendente = Font(name="Calibri", size=11, bold=False, color="9C0006")

    font_data = Font(name="Calibri", size=11, bold=False, color="000000")
    align_data_center = Alignment(horizontal="center", vertical="center")
    align_data_left = Alignment(horizontal="left", vertical="center")

    # Mapear colunas do app para posição
    headers_app = list(df_export.columns)
    num_cols = len(headers_app)

    # Inserir linha de instruções no topo e mesclar
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell_instr = ws.cell(row=1, column=1)
    cell_instr.value = (
        "- Pagamento da diária será efetuado em até 5 dias úteis.\n"
        "- Os pagamentos de diárias só serão efetuados via transferência bancária.\n"
        "- Não é permitido pagamento em conta de terceiros."
    )
    cell_instr.fill = fill_instrucoes
    cell_instr.font = font_instrucoes
    cell_instr.alignment = align_instrucoes
    cell_instr.border = borda
    ws.row_dimensions[1].height = 63

    # Aplicar borda nas células mescladas da linha 1
    for c in range(2, num_cols + 1):
        ws.cell(row=1, column=c).border = borda

    # Formatar cabeçalho (linha 2 após inserção)
    for col_idx, header in enumerate(headers_app, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = borda

    # Formatar dados (a partir da linha 3)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = borda
            cell.font = font_data
            header = ws.cell(row=2, column=cell.column).value
            # Nome alinha à esquerda, resto centralizado
            if header and "NOME" in str(header).upper():
                cell.alignment = align_data_left
            else:
                cell.alignment = align_data_center

            # Destacar ENVIADO/PENDENTE na coluna SITUACAO
            if header == "SITUACAO" and cell.value == "ENVIADO/PENDENTE":
                cell.fill = fill_pendente
                cell.font = font_pendente

            # Formato de moeda nas colunas de valor
            if header in ["VALOR UNITARIO", "VALOR TOTAL"]:
                try:
                    val = float(str(cell.value).replace(",", "."))
                    cell.number_format = r'_-"R$"\ * #,##0.00_-;\-"R$"\ * #,##0.00_-;_-"R$"\ * "-"??_-;_-@_-'
                    cell.value = val
                except:
                    pass

    # Larguras de coluna
    larguras = {
        "LOJA": 12, "MES": 10, "SEMANA": 12, "ANO": 8,
        "NOME COLABORADOR": 38, "CPF": 16, "CARGO": 18,
        "DADOS BANCÁRIOS": 35,
        "MOTIVO": 28, "QTDE DE DIARIAS": 10, "VALOR UNITARIO": 14,
        "VALOR TOTAL": 14, "SITUACAO": 18, "DATA CADASTRO": 18,
        "COMPROVANTE": 35, "OBSERVACAO": 40
    }
    for col_idx, header in enumerate(headers_app, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = larguras.get(header, 18)

    wb.save(caminho)
    wb.close()


@st.cache_data(ttl=0, show_spinner=False)
def carregar_viagens():
    """Carrega o registro de viagens do arquivo Excel."""
    cols_padrao = [
        "ID", "NUMERO_VIAGEM", "COLABORADOR", "LOJA", "ORIGEM", "DESTINO",
        "MOTIVO", "DATA_SAIDA", "DATA_RETORNO", "VALOR_LIBERADO", "TOTAL_GASTO",
        "RESTANTE", "STATUS", "OBSERVACOES", "DATA_CADASTRO"
    ]
    if not os.path.exists(ARQUIVO_VIAGENS):
        return pd.DataFrame(columns=cols_padrao)
    try:
        df = pd.read_excel(ARQUIVO_VIAGENS, dtype=str, keep_default_na=False)
    except Exception:
        return pd.DataFrame(columns=cols_padrao)
    for col in cols_padrao:
        if col not in df.columns:
            df[col] = ""
    for col in ["VALOR_LIBERADO", "TOTAL_GASTO", "RESTANTE"]:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace("", "0.00")
    
    # PROTEÇÃO CONTRA DADOS VAZIOS: tentar restaurar do último backup se a planilha foi corrompida
    if _viagens_estao_vazias(df):
        df_restaurado = _tentar_restaurar_dados("viagens", cols_padrao)
        if df_restaurado is not None:
            st.toast("🛡️ Viagens recuperadas automaticamente do backup!")
            return df_restaurado
    
    return df[cols_padrao]


def salvar_viagens(df_viagens):
    """Salva o registro de viagens no arquivo Excel."""
    # BACKUP AUTOMÁTICO antes de salvar
    try:
        pasta_bkp, arqs = criar_backup_automatico()
    except Exception:
        pass
    
    try:
        with pd.ExcelWriter(ARQUIVO_VIAGENS, engine="openpyxl", mode="w") as f:
            df_viagens.to_excel(f, sheet_name="Viagens", index=False)
    except Exception:
        pass
    st.cache_data.clear()

# ====================== BACKUP / RESTORE ======================
import zipfile
import io

def criar_backup_zip():
    """Cria um arquivo ZIP em memória com todos os dados e anexos."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # Arquivos Excel principais (nome relativo no ZIP)
        for arq in [ARQUIVO, ARQUIVO_DIARIAS, ARQUIVO_VIAGENS, ARQUIVO_COMPRAS]:
            if os.path.exists(arq):
                zf.write(arq, os.path.basename(arq))
        # Pastas de documentos, fotos e comprovantes (caminho relativo ao BASE_DIR)
        for pasta in [PASTA_DOCS, PASTA_DOCS_FUNC, PASTA_FOTOS, PASTA_COMPROVANTES]:
            if os.path.exists(pasta):
                for root, dirs, files in os.walk(pasta):
                    for file in files:
                        caminho_completo = os.path.join(root, file)
                        caminho_zip = os.path.relpath(caminho_completo, start=BASE_DIR)
                        zf.write(caminho_completo, caminho_zip)
    zip_buffer.seek(0)
    return zip_buffer

def restaurar_backup_zip(zip_file):
    """Restaura todos os dados e anexos a partir de um arquivo ZIP."""
    arquivos_extraidos = []
    with zipfile.ZipFile(zip_file, "r") as zf:
        for item in zf.namelist():
            # Ignora arquivos de sistema do Mac/Windows
            if item.startswith("__MACOSX") or item.startswith("."):
                continue
            # Garante que a pasta de destino exista dentro do BASE_DIR
            destino = os.path.join(BASE_DIR, item)
            os.makedirs(os.path.dirname(destino), exist_ok=True)
            with open(destino, "wb") as f_out:
                f_out.write(zf.read(item))
            arquivos_extraidos.append(item)
    return arquivos_extraidos

def lista_lojas():
    d = carregar_dados()
    todas = sorted(set(
        [str(l).strip() for l in d["Base_Dados"]["Loja"] if str(l).strip() != ""] +
        [str(l).strip() for l in d["Auxiliares"]["Loja"] if str(l).strip() != ""]
    ))
    return todas if todas else ["Sem Loja"]

def lista_cargos():
    d = carregar_dados()
    todas = sorted(set(
        [str(c).strip() for c in d["Base_Dados"]["Cargo"] if str(c).strip() != ""] +
        [str(c).strip() for c in d["Auxiliares"]["Cargo"] if str(c).strip() != ""]
    ))
    return todas if todas else ["Sem Cargo"]

def lista_clientes():
    """Retorna lista dinâmica de clientes lendo da aba Auxiliares."""
    d = carregar_dados()
    if "Cliente" in d["Auxiliares"].columns:
        clientes = sorted(set(
            [str(c).strip() for c in d["Auxiliares"]["Cliente"] if str(c).strip() != ""]
        ))
        if clientes:
            return clientes
    # Fallback para dados estáticos
    return list(CLIENTES_FALLBACK)

def lista_lojas_por_cliente(cliente=None):
    """Retorna dict {cliente: [lojas]} ou lista de lojas de um cliente específico."""
    d = carregar_dados()
    if "Cliente" in d["Auxiliares"].columns and "Loja" in d["Auxiliares"].columns:
        df_cli = d["Auxiliares"][d["Auxiliares"]["Cliente"].str.strip() != ""]
        if not df_cli.empty:
            resultado = {}
            for cli in df_cli["Cliente"].unique():
                cli_str = str(cli).strip()
                lojas = sorted(set(
                    [str(l).strip() for l in df_cli[df_cli["Cliente"].str.strip() == cli_str]["Loja"] if str(l).strip() != ""]
                ))
                if lojas:
                    resultado[cli_str] = lojas
            if resultado:
                if cliente:
                    return resultado.get(cliente, [])
                return resultado
    # Fallback para dados estáticos
    if cliente:
        return LOJAS_POR_CLIENTE_FALLBACK.get(cliente, [])
    return dict(LOJAS_POR_CLIENTE_FALLBACK)

def lista_materiais_por_cliente(cliente=None):
    """Retorna dict {cliente: [materiais]} ou lista de materiais de um cliente específico."""
    d = carregar_dados()
    if "Cliente" in d["Auxiliares"].columns and "Material" in d["Auxiliares"].columns:
        df_mat = d["Auxiliares"][d["Auxiliares"]["Material"].str.strip() != ""]
        if not df_mat.empty:
            resultado = {}
            for cli in df_mat["Cliente"].unique():
                cli_str = str(cli).strip()
                mats = sorted(set(
                    [str(m).strip() for m in df_mat[df_mat["Cliente"].str.strip() == cli_str]["Material"] if str(m).strip() != ""]
                ))
                if mats:
                    resultado[cli_str] = mats
            if resultado:
                if cliente:
                    return resultado.get(cliente, [])
                return resultado
    # Fallback para dados estáticos
    if cliente:
        return MATERIAIS_POR_CLIENTE_FALLBACK.get(cliente, [])
    return dict(MATERIAIS_POR_CLIENTE_FALLBACK)

def lista_epis_por_cliente(cliente=None):
    """Retorna dict {cliente: [epis]} ou lista de epis de um cliente específico."""
    d = carregar_dados()
    if "Cliente" in d["Auxiliares"].columns and "EPI" in d["Auxiliares"].columns:
        df_epi = d["Auxiliares"][d["Auxiliares"]["EPI"].str.strip() != ""]
        if not df_epi.empty:
            resultado = {}
            for cli in df_epi["Cliente"].unique():
                cli_str = str(cli).strip()
                epis = sorted(set(
                    [str(e).strip() for e in df_epi[df_epi["Cliente"].str.strip() == cli_str]["EPI"] if str(e).strip() != ""]
                ))
                if epis:
                    resultado[cli_str] = epis
            if resultado:
                if cliente:
                    return resultado.get(cliente, [])
                return resultado
    # Fallback para dados estáticos
    if cliente:
        return EPIS_POR_CLIENTE_FALLBACK.get(cliente, [])
    return dict(EPIS_POR_CLIENTE_FALLBACK)

def calcular_e_atualizar(form):
    hoje = datetime.now().date()

    # Aviso Prévio
    if form.get("dt_aviso") and form.get("dias_aviso") and str(form["dias_aviso"]).isdigit():
        try:
            dt = datetime.strptime(form["dt_aviso"], "%d/%m/%Y")
            form["termino_aviso"] = (dt + timedelta(days=int(form["dias_aviso"]) - 1)).strftime("%d/%m/%Y")
            termino_aviso_date = datetime.strptime(form["termino_aviso"], "%d/%m/%Y").date()
            # Se o aviso prévio venceu e não tem outro evento, muda para Demitido S/JC
            if termino_aviso_date < hoje and not any([
                form.get("dt_pedido","").strip(), form.get("dt_rescisao","").strip(),
                form.get("dt_abandono","").strip(), form.get("dt_desistencia","").strip(),
                form.get("dt_termino_cont","").strip()
            ]):
                form["situacao"] = "Demitido S/JC"
            elif termino_aviso_date >= hoje and not any([
                form.get("dt_pedido","").strip(), form.get("dt_rescisao","").strip(),
                form.get("dt_abandono","").strip(), form.get("dt_desistencia","").strip(),
                form.get("dt_termino_cont","").strip()
            ]):
                form["situacao"] = "Aviso Prévio"
        except: form["termino_aviso"] = ""
    else: form["termino_aviso"] = ""

    # Licença
    if form.get("dt_lic") and form.get("dias_lic") and str(form["dias_lic"]).isdigit():
        try:
            dt = datetime.strptime(form["dt_lic"], "%d/%m/%Y")
            form["termino_lic"] = (dt + timedelta(days=int(form["dias_lic"]) - 1)).strftime("%d/%m/%Y")
            termino_lic_date = datetime.strptime(form["termino_lic"], "%d/%m/%Y").date()
            # Se a licença venceu e não tem outro evento, volta para Ativo
            if termino_lic_date < hoje and not any([
                form.get("dt_pedido","").strip(), form.get("dt_rescisao","").strip(),
                form.get("dt_abandono","").strip(), form.get("dt_desistencia","").strip(),
                form.get("dt_termino_cont","").strip(),
                form.get("dt_aviso","").strip(), form.get("dt_fer","").strip(),
                form.get("dt_af","").strip()
            ]):
                form["situacao"] = "Ativo"
            elif termino_lic_date >= hoje and not any([
                form.get("dt_pedido","").strip(), form.get("dt_rescisao","").strip(),
                form.get("dt_abandono","").strip(), form.get("dt_desistencia","").strip(),
                form.get("dt_termino_cont","").strip(),
                form.get("dt_aviso","").strip(), form.get("dt_fer","").strip(),
                form.get("dt_af","").strip()
            ]):
                form["situacao"] = "Licença"
        except: form["termino_lic"] = ""
    else: form["termino_lic"] = ""

    # Férias
    if form.get("dt_fer") and form.get("dias_fer") and str(form["dias_fer"]).isdigit():
        try:
            dt = datetime.strptime(form["dt_fer"], "%d/%m/%Y")
            form["retorno_fer"] = (dt + timedelta(days=int(form["dias_fer"]) - 1)).strftime("%d/%m/%Y")
            retorno_date = datetime.strptime(form["retorno_fer"], "%d/%m/%Y").date()
            if retorno_date < hoje and not any([
                form.get("dt_pedido","").strip(), form.get("dt_rescisao","").strip(),
                form.get("dt_abandono","").strip(), form.get("dt_desistencia","").strip(),
                form.get("dt_termino_cont","").strip(),
                form.get("dt_aviso","").strip(), form.get("dt_lic","").strip(),
                form.get("dt_af","").strip()
            ]):
                form["situacao"] = "Ativo"
            elif retorno_date >= hoje and not any([
                form.get("dt_pedido","").strip(), form.get("dt_rescisao","").strip(),
                form.get("dt_abandono","").strip(), form.get("dt_desistencia","").strip(),
                form.get("dt_termino_cont","").strip(),
                form.get("dt_aviso","").strip(), form.get("dt_lic","").strip(),
                form.get("dt_af","").strip()
            ]):
                form["situacao"] = "Férias"
        except:
            form["retorno_fer"] = ""
    else:
        form["retorno_fer"] = ""

    # Afastamento
    if form.get("dt_af") and form.get("dias_af") and str(form["dias_af"]).isdigit():
        try:
            dt = datetime.strptime(form["dt_af"], "%d/%m/%Y")
            form["retorno_af"] = (dt + timedelta(days=int(form["dias_af"]) - 1)).strftime("%d/%m/%Y")
            retorno_af_date = datetime.strptime(form["retorno_af"], "%d/%m/%Y").date()
            tipo_af = form.get("tipo_af", "Nenhum")
            if tipo_af != "Nenhum" and retorno_af_date < hoje and not any([
                form.get("dt_pedido","").strip(), form.get("dt_rescisao","").strip(),
                form.get("dt_abandono","").strip(), form.get("dt_desistencia","").strip(),
                form.get("dt_termino_cont","").strip(),
                form.get("dt_aviso","").strip(), form.get("dt_lic","").strip(),
                form.get("dt_fer","").strip()
            ]):
                form["situacao"] = "Ativo"
            elif tipo_af != "Nenhum" and retorno_af_date >= hoje and not any([
                form.get("dt_pedido","").strip(), form.get("dt_rescisao","").strip(),
                form.get("dt_abandono","").strip(), form.get("dt_desistencia","").strip(),
                form.get("dt_termino_cont","").strip(),
                form.get("dt_aviso","").strip(), form.get("dt_lic","").strip(),
                form.get("dt_fer","").strip()
            ]):
                form["situacao"] = tipo_af
        except: form["retorno_af"] = ""
    else: form["retorno_af"] = ""

    # Eventos definitivos (mantêm prioridade)
    if form.get("dt_termino_cont") and form.get("dt_termino_cont").strip():
        form["situacao"] = "Término de Contrato"
    elif form.get("dt_pedido") and form.get("dt_pedido").strip():
        form["situacao"] = "Pedido de Conta"
    elif form.get("dt_rescisao") and form.get("dt_rescisao").strip():
        form["situacao"] = "Rescisão Indireta"
    elif form.get("dt_abandono") and form.get("dt_abandono").strip():
        form["situacao"] = "Abandono"
    elif form.get("dt_desistencia") and form.get("dt_desistencia").strip():
        form["situacao"] = "Desistente"

    return form

def add_historico_auto(mat, nome, acao, dados_completos):
    dados = carregar_dados()
    registro = {"DataEvento": datetime.now().strftime("%d/%m/%Y"), "TipoEvento": acao, "Detalhes": ""}
    registro.update(dados_completos)
    idx = dados["Historico"].index[dados["Historico"]["Matricula"] == mat].tolist()
    if idx:
        idx_linha = idx[0]
        for coluna, valor in registro.items():
            dados["Historico"].at[idx_linha, coluna] = valor
    else:
        dados["Historico"] = pd.concat([dados["Historico"], pd.DataFrame([registro])], ignore_index=True)
    if not salvar_dados(dados):
        st.error("❌ Erro ao salvar histórico. Verifique se o arquivo Excel não está aberto.")

def gerar_ficha_individual(fd, fh, mr):
    """Gera arquivo Excel da ficha individual usando openpyxl diretamente para evitar colunas duplicadas."""
    nome_arq = os.path.join(BASE_DIR, f"Rel_{mr}_ficha.xlsx")
    
    # Garante que usamos apenas colunas padrão na ordem correta
    colunas_dados = [
        "Matricula","Nome","CPF","RG","PIS","Nascimento","Admissao",
        "Telefone","Endereco","Loja","Cargo","Salario","Situacao",
        "DataAvisoPrevio","DiasAvisoPrevio","DataTerminoAviso",
        "DataFeriasInicio","DiasFerias","DataRetornoFerias",
        "DataPedidoConta","DataRescisao","DataAbandono","DataDesistencia",
        "DataTerminoContrato",
        "DataLicenca","DiasLicenca","DataTerminoLicenca",
        "DataAfastamento","DiasAfastamento","DataRetornoAfastamento",
        "CaminhoFoto"
    ]
    colunas_historico = [
        "DataEvento","TipoEvento","Matricula","Nome","CPF","RG","PIS",
        "Nascimento","Admissao","Telefone","Endereco","Loja","Cargo",
        "Salario","Situacao","DataAvisoPrevio","DiasAvisoPrevio","DataTerminoAviso",
        "DataFeriasInicio","DiasFerias","DataRetornoFerias",
        "DataPedidoConta","DataRescisao","DataAbandono","DataDesistencia",
        "DataTerminoContrato",
        "DataLicenca","DiasLicenca","DataTerminoLicenca",
        "DataAfastamento","DiasAfastamento","DataRetornoAfastamento","Detalhes"
    ]
    
    wb = Workbook()
    
    # Aba Dados
    ws_dados = wb.active
    ws_dados.title = "Dados"
    
    # Filtra apenas colunas que existem no DataFrame
    cols_dados_existentes = [c for c in colunas_dados if c in fd.columns]
    fd_limpo = fd[cols_dados_existentes].copy()
    
    # Escreve cabeçalho
    for col_idx, col_name in enumerate(cols_dados_existentes, 1):
        ws_dados.cell(row=1, column=col_idx, value=col_name)
    
    # Escreve dados
    for row_idx, (_, row) in enumerate(fd_limpo.iterrows(), 2):
        for col_idx, col_name in enumerate(cols_dados_existentes, 1):
            ws_dados.cell(row=row_idx, column=col_idx, value=row[col_name])
    
    # Aba Histórico
    ws_hist = wb.create_sheet(title="Histórico")
    
    if not fh.empty:
        cols_hist_existentes = [c for c in colunas_historico if c in fh.columns]
        fh_limpo = fh[cols_hist_existentes].copy()
        for col_idx, col_name in enumerate(cols_hist_existentes, 1):
            ws_hist.cell(row=1, column=col_idx, value=col_name)
        for row_idx, (_, row) in enumerate(fh_limpo.iterrows(), 2):
            for col_idx, col_name in enumerate(cols_hist_existentes, 1):
                ws_hist.cell(row=row_idx, column=col_idx, value=row[col_name])
    else:
        ws_hist.cell(row=1, column=1, value="Aviso")
        ws_hist.cell(row=2, column=1, value="Sem histórico registrado")
    
    wb.save(nome_arq)
    wb.close()
    return nome_arq

def verificar_retorno_ferias_automatico():
    """Verifica funcionários em férias que já deveriam ter retornado e atualiza para Ativo."""
    try:
        dados = carregar_dados()
        hoje = datetime.now().date()
        base = dados["Base_Dados"]
        alterados = []
        for idx, row in base.iterrows():
            if str(row.get("Situacao", "")).strip() != "Férias":
                continue
            retorno_str = str(row.get("DataRetornoFerias", "")).strip()
            if not retorno_str:
                continue
            try:
                retorno_date = datetime.strptime(retorno_str, "%d/%m/%Y").date()
                if retorno_date < hoje:
                    # Já passou a data de retorno, volta para Ativo
                    base.at[idx, "Situacao"] = "Ativo"
                    alterados.append({
                        "Matricula": row.get("Matricula", ""),
                        "Nome": row.get("Nome", ""),
                        "DataRetorno": retorno_str
                    })
            except Exception:
                continue
        if alterados:
            if not salvar_dados(dados):
                st.error("❌ Erro ao salvar retorno automático de férias. Verifique se o arquivo Excel não está aberto.")
            # Adiciona ao histórico
            for alt in alterados:
                registro = {
                    "DataEvento": datetime.now().strftime("%d/%m/%Y"),
                    "TipoEvento": "Retorno Automático de Férias",
                    "Detalhes": f"Funcionário retornou de férias em {alt['DataRetorno']}. Situação alterada para Ativo automaticamente."
                }
                # Busca dados completos do funcionário
                rf = base[base["Matricula"] == alt["Matricula"]]
                if not rf.empty:
                    registro.update(rf.iloc[0].to_dict())
                ih = dados["Historico"].index[dados["Historico"]["Matricula"] == alt["Matricula"]].tolist()
                if ih:
                    idx_linha = ih[0]
                    for coluna, valor in registro.items():
                        dados["Historico"].at[idx_linha, coluna] = valor
                else:
                    dados["Historico"] = pd.concat([dados["Historico"], pd.DataFrame([registro])], ignore_index=True)
            if not salvar_dados(dados):
                st.error("❌ Erro ao salvar histórico de férias. Verifique se o arquivo Excel não está aberto.")
    except Exception:
        pass

def verificar_retorno_afastamentos_automatico():
    """Verifica funcionários em afastamento (Doença, Acidente, Maternidade) que já deveriam ter retornado e atualiza para Ativo."""
    try:
        dados = carregar_dados()
        hoje = datetime.now().date()
        base = dados["Base_Dados"]
        situacoes_afastamento = ["Doença", "Acidente", "Maternidade"]
        alterados = []
        for idx, row in base.iterrows():
            if str(row.get("Situacao", "")).strip() not in situacoes_afastamento:
                continue
            retorno_str = str(row.get("DataRetornoAfastamento", "")).strip()
            if not retorno_str:
                continue
            try:
                retorno_date = datetime.strptime(retorno_str, "%d/%m/%Y").date()
                if retorno_date < hoje:
                    # Já passou a data de retorno, volta para Ativo
                    base.at[idx, "Situacao"] = "Ativo"
                    alterados.append({
                        "Matricula": row.get("Matricula", ""),
                        "Nome": row.get("Nome", ""),
                        "TipoAfastamento": row.get("Situacao", ""),
                        "DataRetorno": retorno_str
                    })
            except Exception:
                continue
        if alterados:
            if not salvar_dados(dados):
                st.error("❌ Erro ao salvar retorno automático de afastamento. Verifique se o arquivo Excel não está aberto.")
            # Adiciona ao histórico
            for alt in alterados:
                registro = {
                    "DataEvento": datetime.now().strftime("%d/%m/%Y"),
                    "TipoEvento": f"Retorno Automático de {alt['TipoAfastamento']}",
                    "Detalhes": f"Funcionário retornou de {alt['TipoAfastamento'].lower()} em {alt['DataRetorno']}. Situação alterada para Ativo automaticamente."
                }
                # Busca dados completos do funcionário
                rf = base[base["Matricula"] == alt["Matricula"]]
                if not rf.empty:
                    registro.update(rf.iloc[0].to_dict())
                ih = dados["Historico"].index[dados["Historico"]["Matricula"] == alt["Matricula"]].tolist()
                if ih:
                    idx_linha = ih[0]
                    for coluna, valor in registro.items():
                        dados["Historico"].at[idx_linha, coluna] = valor
                else:
                    dados["Historico"] = pd.concat([dados["Historico"], pd.DataFrame([registro])], ignore_index=True)
            if not salvar_dados(dados):
                st.error("❌ Erro ao salvar histórico de afastamento. Verifique se o arquivo Excel não está aberto.")
    except Exception:
        pass

# ====================== INTERFACE PRINCIPAL ======================
st.set_page_config(page_title="SISTEMA RH COMPLETO", layout="wide", initial_sidebar_state="collapsed")
st.title("📋 SISTEMA RH COMPLETO")

# Verifica retorno automático de férias e afastamentos no início da sessão (apenas 1x)
if "ferias_verificado" not in st.session_state:
    verificar_retorno_ferias_automatico()
    st.session_state["ferias_verificado"] = True
if "afastamentos_verificado" not in st.session_state:
    verificar_retorno_afastamentos_automatico()
    st.session_state["afastamentos_verificado"] = True

# Backup automático por intervalo de tempo (verifica a cada interação se já passou o tempo)
backup_automatico_periodico()

# ⚠️ LINHA OBRIGATÓRIA: CRIA TODAS AS ABAS ANTES DE USÁ-LAS
aba1, aba2, aba3, aba4, aba5, aba6, aba7, aba8, aba9, aba10, aba11, aba12 = st.tabs([
    "Cadastro", "Painel", "Prazos e Férias", "Histórico", "Relatórios", "📎 Documentos", "⚙️ Lojas e Cargos", "💰 CONTROLE DE DIÁRIAS", "🗺️ GUIA VIAGEM", "💾 Backup", "🌐 Tradutor", "🛒 Compras"
])



# ====================== FUNÇÕES GUIA DE VIAGEM ======================
def geocodificar(endereco):
    """Converte endereço em latitude/longitude usando Nominatim."""
    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {"q": endereco, "format": "json", "limit": 1}
        headers = {"User-Agent": "RHApp/1.0"}
        resp = requests.get(url, params=params, headers=headers, timeout=15)
        dados_geo = resp.json()
        if dados_geo:
            return float(dados_geo[0]["lat"]), float(dados_geo[0]["lon"])
    except Exception:
        pass
    return None, None

def calcular_rota(lat1, lon1, lat2, lon2):
    """Calcula distância, tempo e geometria da rota via OSRM."""
    try:
        url = f"http://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}"
        params = {"overview": "full", "geometries": "geojson", "steps": "true"}
        resp = requests.get(url, params=params, timeout=20)
        dados_r = resp.json()
        if dados_r.get("routes"):
            rota = dados_r["routes"][0]
            distancia_km = rota["distance"] / 1000
            tempo_min = rota["duration"] / 60
            geometria = rota["geometry"]
            return distancia_km, tempo_min, geometria
    except Exception:
        pass
    return None, None, None



def calcular_zoom(distancia_km):
    """Calcula o nível de zoom do mapa baseado na distância."""
    if distancia_km < 5: return 13
    elif distancia_km < 20: return 11
    elif distancia_km < 50: return 10
    elif distancia_km < 100: return 9
    elif distancia_km < 300: return 8
    elif distancia_km < 600: return 7
    elif distancia_km < 1200: return 6
    elif distancia_km < 2500: return 5
    elif distancia_km < 5000: return 4
    else: return 3

def haversine(lat1, lon1, lat2, lon2):
    """Calcula distância em linha reta entre dois pontos (km)."""
    from math import radians, sin, cos, sqrt, atan2
    R = 6371.0
    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dlambda = radians(lon2 - lon1)
    a = sin(dphi/2)**2 + cos(phi1)*cos(phi2)*sin(dlambda/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    return R * c

# ================ ABA 1 - CADASTRO ================
with aba1:
    dados = carregar_dados()
    
    # ---------- BUSCA COM AUTOCOMPLETE ----------
    st.markdown("**🔍 Buscar Colaborador**")
    
    # Prepara lista para autocomplete
    base_auto = dados["Base_Dados"].copy()
    base_auto["Matricula"] = base_auto["Matricula"].fillna("").astype(str).str.strip()
    base_auto["Nome"] = base_auto["Nome"].fillna("").astype(str).str.strip()
    base_auto["Loja"] = base_auto["Loja"].fillna("").astype(str).str.strip()
    base_auto["Situacao"] = base_auto["Situacao"].fillna("").astype(str).str.strip()
    base_auto["Cargo"] = base_auto["Cargo"].fillna("").astype(str).str.strip()
    
    # Ordena por nome
    base_auto = base_auto.sort_values("Nome", key=lambda col: col.str.upper())
    
    # Opções do autocomplete: "MATRICULA - NOME"
    opcoes_auto = [f"{row['Matricula']} - {row['Nome']}" for _, row in base_auto.iterrows()]
    
    # Autocomplete - ao selecionar já carrega automaticamente
    sel_auto = st.selectbox(
        "Digite o nome ou matrícula e selecione:",
        options=[""] + opcoes_auto,
        index=0,
        key="autocomplete_func",
        help="Comece a digitar para filtrar automaticamente"
    )
    
    # Extrai matrícula se selecionou no autocomplete
    mat_sel = ""
    if sel_auto and " - " in sel_auto:
        mat_sel = sel_auto.split(" - ")[0].strip()
        st.success(f"✅ Colaborador selecionado: {sel_auto}")
    
    # ---------- FILTROS ----------
    st.markdown("---")
    st.markdown("**📋 Filtros da Tabela**")
    
    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        filtro_loja = st.selectbox("Filtrar por Loja", ["Todas"] + lista_lojas(), key="filtro_loja_cad")
    with col_f2:
        filtro_sit = st.selectbox("Filtrar por Situação", ["Todas"] + SITUACOES, key="filtro_sit_cad")
    with col_f3:
        filtro_cargo = st.selectbox("Filtrar por Cargo", ["Todos"] + lista_cargos(), key="filtro_cargo_cad")

    lista = base_auto.copy()
    if filtro_loja != "Todas":
        lista = lista[lista["Loja"] == filtro_loja.strip()]
    if filtro_sit != "Todas":
        lista = lista[lista["Situacao"] == filtro_sit]
    if filtro_cargo != "Todos":
        lista = lista[lista["Cargo"] == filtro_cargo.strip()]

    # Contador de resultados
    total_encontrados = len(lista)
    st.markdown(f"**📊 Total encontrado: {total_encontrados} colaborador(es)**")

    st.dataframe(
        lista[["Matricula","Nome","Loja","Situacao","Cargo"]],
        use_container_width=True, hide_index=True
    )
    
    reg = pd.DataFrame()
    if mat_sel:
        mat_busca = str(mat_sel).strip()
        reg = dados["Base_Dados"][dados["Base_Dados"]["Matricula"] == mat_busca]

    val_campo = lambda nome: reg.iloc[0][nome] if (not reg.empty and nome in reg.columns) else ""

    prazos_exp = []
    if not reg.empty and val_campo("Admissao").strip():
        try:
            dt_adm = datetime.strptime(val_campo("Admissao"), "%d/%m/%Y")
            hoje = datetime.now()
            dias_corridos = (hoje - dt_adm).days
            for prazo in [30, 45, 60, 90]:
                rest = prazo - dias_corridos
                if rest > 0:
                    status = f"Faltam {rest} dias"
                elif rest == 0:
                    status = "HOJE"
                else:
                    status = f"Vencido há {abs(rest)} dias"
                prazos_exp.append([f"{prazo} dias", (dt_adm + timedelta(days=prazo - 1)).strftime("%d/%m/%Y"), status])
        except:
            pass

    if not reg.empty:
        temp = {
            "dt_aviso": val_campo("DataAvisoPrevio"), "dias_aviso": val_campo("DiasAvisoPrevio"),
            "dt_lic": val_campo("DataLicenca"), "dias_lic": val_campo("DiasLicenca"),
            "dt_fer": val_campo("DataFeriasInicio"), "dias_fer": val_campo("DiasFerias"),
            "dt_af": val_campo("DataAfastamento"), "dias_af": val_campo("DiasAfastamento"),
            "dt_pedido": val_campo("DataPedidoConta"), "dt_rescisao": val_campo("DataRescisao"),
            "dt_abandono": val_campo("DataAbandono"), "dt_termino_cont": val_campo("DataTerminoContrato"),
            "situacao": val_campo("Situacao"), "caminho_foto": val_campo("CaminhoFoto")
        }
        temp = calcular_e_atualizar(temp)
        term_aviso_val, term_lic_val, ret_fer_val, ret_af_val, situacao_val, caminho_foto_atual = temp["termino_aviso"], temp["termino_lic"], temp["retorno_fer"], temp["retorno_af"], temp["situacao"], temp["caminho_foto"]
    else:
        term_aviso_val = term_lic_val = ret_fer_val = ret_af_val = caminho_foto_atual = ""
        situacao_val = "Ativo"

    # ---------- Controle de modo (novo vs edição) ----------
    modo_edicao = bool(mat_sel.strip())
    matricula_original = mat_sel.strip() if modo_edicao else ""

    # Salva matrícula original no session_state para garantir consistência
    if modo_edicao:
        st.session_state["_matricula_original_edicao"] = matricula_original
    elif "_matricula_original_edicao" not in st.session_state:
        st.session_state["_matricula_original_edicao"] = ""

    if st.button("🗑️ LIMPAR TODOS OS CAMPOS", use_container_width=True, type="secondary"):
        if "autocomplete_func" in st.session_state:
            del st.session_state["autocomplete_func"]
        st.session_state["_matricula_original_edicao"] = ""
        st.session_state.pop("confirmar_exclusao", None)
        st.rerun()
    with st.form("form_cadastro", clear_on_submit=False):
        st.subheader("Dados Básicos")
        col_foto, col_dados = st.columns([1,3])
        
        with col_foto:
            st.markdown("**Foto do Funcionário**")
            if caminho_foto_atual and os.path.exists(caminho_foto_atual):
                st.image(caminho_foto_atual, width=180, caption="Foto atual")
            else:
                st.info("Sem foto")
            
            nova_foto = st.file_uploader("Enviar/Trocar foto", type=["jpg","jpeg","png"], key=f"foto_{mat_sel}")
            excluir_foto = st.checkbox("🗑️ Excluir foto atual", value=False)

        with col_dados:
            c1,c2,c3 = st.columns(3)
            with c1:
                if modo_edicao:
                    matricula = st.text_input("Matrícula * (igual planilha)", value=matricula_original, disabled=True)
                    st.caption("🔒 Matrícula bloqueada em modo de edição")
                else:
                    matricula = st.text_input("Matrícula * (igual planilha)", value="")
                nome = st.text_input("Nome Completo", value=val_campo("Nome"))
                cpf = st.text_input("CPF", value=val_campo("CPF"))
                rg = st.text_input("RG", value=val_campo("RG"))
                pis = st.text_input("PIS", value=val_campo("PIS"))
            with c2:
                nascimento = st.text_input("Data Nascimento (dd/mm/aaaa)", value=val_campo("Nascimento"))
                admissao = st.text_input("Data Admissão (dd/mm/aaaa)", value=val_campo("Admissao"))
                telefone = st.text_input("Telefone", value=val_campo("Telefone"))
                endereco = st.text_input("Endereço Completo", value=val_campo("Endereco"))
            with c3:
                lojas = lista_lojas()
                idx_loja = lojas.index(val_campo("Loja")) if val_campo("Loja") in lojas else 0
                loja = st.selectbox("🏬 Loja", lojas, index=idx_loja)

                cargos = lista_cargos()
                idx_cargo = cargos.index(val_campo("Cargo")) if val_campo("Cargo") in cargos else 0
                cargo = st.selectbox("💼 Cargo", cargos, index=idx_cargo)

                salario = st.text_input("Salário", value=val_campo("Salario"))

                idx_sit = SITUACOES.index(situacao_val) if situacao_val in SITUACOES else 0
                situacao = st.selectbox("📊 Situação", SITUACOES, index=idx_sit)

        if prazos_exp:
            st.markdown("---")
            st.subheader("⏳ PRAZOS DE EXPERIÊNCIA")
            st.dataframe(
                pd.DataFrame(prazos_exp, columns=["Prazo", "Data Final", "Situação"]),
                use_container_width=True, hide_index=True
            )
        elif not reg.empty:
            st.info("ℹ️ Informe a Data de Admissão para visualizar os prazos.")

        st.markdown("---")
        st.subheader("Eventos Trabalhistas")
        av1,av2,av3 = st.columns(3)
        with av1:
            st.markdown("**Aviso Prévio**")
            dt_aviso = st.text_input("Data Aviso", value=val_campo("DataAvisoPrevio"))
            dias_aviso = st.text_input("Dias Aviso", value=val_campo("DiasAvisoPrevio"))
            term_aviso = st.text_input("Término Aviso", value=term_aviso_val, disabled=True)
        with av2:
            st.markdown("**Licença**")
            dt_lic = st.text_input("Data Licença", value=val_campo("DataLicenca"))
            dias_lic = st.text_input("Dias Licença", value=val_campo("DiasLicenca"))
            term_lic = st.text_input("Término Licença", value=term_lic_val, disabled=True)
        with av3:
            st.markdown("**Férias**")
            dt_fer = st.text_input("Início Férias", value=val_campo("DataFeriasInicio"))
            dias_fer = st.text_input("Dias Férias", value=val_campo("DiasFerias"))
            ret_fer = st.text_input("Retorno Férias", value=ret_fer_val, disabled=True)

        af1,af2 = st.columns(2)
        with af1:
            st.markdown("**Afastamento**")
            dt_af = st.text_input("Data Afastamento", value=val_campo("DataAfastamento"))
            dias_af = st.text_input("Dias Afastamento", value=val_campo("DiasAfastamento"))
            ret_af = st.text_input("Retorno Afastamento", value=ret_af_val, disabled=True)
            tipo_af = st.selectbox("Tipo Afastamento", ["Nenhum", "Doença", "Acidente", "Maternidade"], index=["Nenhum", "Doença", "Acidente", "Maternidade"].index(val_campo("TipoAfastamento")) if val_campo("TipoAfastamento") in ["Nenhum", "Doença", "Acidente", "Maternidade"] else 0)
        with af2:
            st.markdown("**Desligamento**")
            dt_ped = st.text_input("Data Pedido Conta", value=val_campo("DataPedidoConta"))
            dt_res = st.text_input("Data Rescisão", value=val_campo("DataRescisao"))
            dt_aband = st.text_input("Data Abandono", value=val_campo("DataAbandono"))
            dt_desist = st.text_input("Data Desistência", value=val_campo("DataDesistencia"))
            dt_termino_cont = st.text_input("📅 Data Término de Contrato", value=val_campo("DataTerminoContrato"))

        btn_salvar = st.form_submit_button("💾 SALVAR CADASTRO", type="primary", use_container_width=True)
        if btn_salvar:
            try:
                matricula_tratada = str(matricula).strip()
                if not matricula_tratada:
                    st.error("❌ INFORME A MATRÍCULA!")
                    st.stop()
                # Usa a matrícula original do registro selecionado para garantir edição correta
                mat_para_busca = matricula_original if modo_edicao else matricula_tratada
                mat_para_busca = mat_para_busca.strip()
                caminho_final_foto = caminho_foto_atual
                if excluir_foto and caminho_final_foto and os.path.exists(caminho_final_foto):
                    os.remove(caminho_final_foto)
                    caminho_final_foto = ""
                if nova_foto:
                    if caminho_final_foto and os.path.exists(caminho_final_foto):
                        os.remove(caminho_final_foto)
                    extensao = os.path.splitext(nova_foto.name)[1].lower()
                    nome_foto = f"{matricula_tratada}_foto_{datetime.now().strftime('%Y%m%d%H%M%S')}{extensao}"
                    caminho_final_foto = os.path.join(PASTA_FOTOS, nome_foto)
                    img = Image.open(nova_foto)
                    img.save(caminho_final_foto)

                dados_form = calcular_e_atualizar({
                    "mat": matricula_tratada, "nome": nome, "cpf": cpf, "rg": rg, "pis": pis,
                    "nasc": nascimento, "adm": admissao, "tel": telefone, "end": endereco,
                    "loja": loja, "cargo": cargo, "sal": salario, "situacao": situacao,
                    "dt_aviso": dt_aviso, "dias_aviso": dias_aviso, "termino_aviso": term_aviso,
                    "dt_lic": dt_lic, "dias_lic": dias_lic, "termino_lic": term_lic,
                    "dt_fer": dt_fer, "dias_fer": dias_fer, "retorno_fer": ret_fer,
                    "dt_af": dt_af, "dias_af": dias_af, "retorno_af": ret_af, "tipo_af": tipo_af,
                    "dt_pedido": dt_ped, "dt_rescisao": dt_res, "dt_abandono": dt_aband,
                    "dt_desistencia": dt_desist, "dt_termino_cont": dt_termino_cont
                })
                registro_final = {
                    "Matricula": dados_form["mat"], "Nome": dados_form["nome"], "CPF": dados_form["cpf"],
                    "RG": dados_form["rg"], "PIS": dados_form["pis"], "Nascimento": dados_form["nasc"],
                    "Admissao": dados_form["adm"], "Telefone": dados_form["tel"], "Endereco": dados_form["end"],
                    "Loja": dados_form["loja"], "Cargo": dados_form["cargo"], "Salario": dados_form["sal"],
                    "Situacao": dados_form["situacao"], "DataAvisoPrevio": dados_form["dt_aviso"],
                    "DiasAvisoPrevio": dados_form["dias_aviso"], "DataTerminoAviso": dados_form["termino_aviso"],
                    "DataFeriasInicio": dados_form["dt_fer"], "DiasFerias": dados_form["dias_fer"],
                    "DataRetornoFerias": dados_form["retorno_fer"], "DataPedidoConta": dados_form["dt_pedido"],
                    "DataRescisao": dados_form["dt_rescisao"], "DataAbandono": dados_form["dt_abandono"],
                    "DataDesistencia": dados_form["dt_desistencia"],
                    "DataTerminoContrato": dados_form["dt_termino_cont"],
                    "DataLicenca": dados_form["dt_lic"], "DiasLicenca": dados_form["dias_lic"],
                    "DataTerminoLicenca": dados_form["termino_lic"],
                    "DataAfastamento": dados_form["dt_af"], "DiasAfastamento": dados_form["dias_af"],
                    "DataRetornoAfastamento": dados_form["retorno_af"],
                    "TipoAfastamento": dados_form.get("tipo_af", "Nenhum"),
                    "CaminhoFoto": caminho_final_foto
                }
                # Busca pela matrícula original em caso de edição, ou pela matrícula informada em novo cadastro
                indice = dados["Base_Dados"].index[dados["Base_Dados"]["Matricula"] == mat_para_busca].tolist()
                acao_hist = "Atualização Cadastral" if indice else "Novo Cadastro"
                if indice:
                    idx_linha = indice[0]
                    for coluna, valor in registro_final.items():
                        dados["Base_Dados"].at[idx_linha, coluna] = valor
                else:
                    dados["Base_Dados"] = pd.concat([dados["Base_Dados"], pd.DataFrame([registro_final])], ignore_index=True)
                if not salvar_dados(dados):
                    st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                    st.stop()
                add_historico_auto(matricula_tratada, dados_form["nome"], acao_hist, registro_final)
                # Limpa flag de edição após salvar
                if "_matricula_original_edicao" in st.session_state:
                    st.session_state["_matricula_original_edicao"] = ""
                st.success(f"✅ {'Atualização' if acao_hist == 'Atualização Cadastral' else 'Cadastro'} realizada! Matrícula: **{matricula_tratada}**")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Erro ao salvar: {e}")
                import traceback
                st.code(traceback.format_exc())

    # ---------- EXCLUSÃO DE REGISTRO ----------
    if mat_sel.strip():
        st.markdown("---")
        st.markdown("**⚠️ Zona de Exclusão**")
        
        with st.expander("🗑️ EXCLUIR REGISTRO (clique para expandir)", expanded=False):
            st.warning("⚠️ Esta ação não pode ser desfeita! Todos os dados, foto e documentos deste colaborador serão removidos permanentemente.")
            
            # Usa a matrícula original para garantir consistência
            mat_para_excluir = matricula_original if modo_edicao else mat_sel.strip()
            st.info(f"Colaborador a excluir: **{mat_para_excluir} - {val_campo('Nome')}**")
            
            confirma = st.checkbox("CONFIRMO QUE DESEJO EXCLUIR PERMANENTEMENTE", key="chk_confirma_exclusao")
            c1, c2 = st.columns(2)
            with c1:
                btn_confirmar_exclusao = st.button("✅ SIM, EXCLUIR", type="primary", use_container_width=True, key="btn_confirmar_excluir_registro", disabled=not confirma)
            with c2:
                btn_cancelar_exclusao = st.button("❌ CANCELAR", type="secondary", use_container_width=True, key="btn_cancelar_excluir_registro")
            
            if btn_cancelar_exclusao:
                st.info("Exclusão cancelada.")
            
            if btn_confirmar_exclusao and confirma:
                try:
                    indice = dados["Base_Dados"].index[dados["Base_Dados"]["Matricula"] == mat_para_excluir.strip()].tolist()
                    if indice:
                        dados_excluir = dados["Base_Dados"].loc[indice[0]].to_dict()
                        if dados_excluir.get("CaminhoFoto") and os.path.exists(str(dados_excluir["CaminhoFoto"])):
                            os.remove(str(dados_excluir["CaminhoFoto"]))
                        docs_excluir = dados["Docs_Funcionarios"][dados["Docs_Funcionarios"]["Matricula"] == mat_para_excluir.strip()]
                        for _, d in docs_excluir.iterrows():
                            if os.path.exists(d["Caminho"]): os.remove(d["Caminho"])
                        dados["Docs_Funcionarios"] = dados["Docs_Funcionarios"][dados["Docs_Funcionarios"]["Matricula"] != mat_para_excluir.strip()]
                        dados["Base_Dados"] = dados["Base_Dados"].drop(indice[0])
                        if not salvar_dados(dados):
                            st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                            st.stop()
                        add_historico_auto(mat_para_excluir.strip(), dados_excluir.get("Nome", ""), "Exclusão de Cadastro", dados_excluir)
                        # Limpa sessão
                        if "autocomplete_func" in st.session_state:
                            del st.session_state["autocomplete_func"]
                        if "_matricula_original_edicao" in st.session_state:
                            st.session_state["_matricula_original_edicao"] = ""
                        st.session_state.pop("confirmar_exclusao", None)
                        if "chk_confirma_exclusao" in st.session_state:
                            del st.session_state["chk_confirma_exclusao"]
                        st.success("✅ Registro, foto e documentos excluídos com sucesso!")
                        st.rerun()
                    else:
                        st.error("❌ Registro não encontrado! Matrícula: " + mat_para_excluir)
                except Exception as e:
                    st.error(f"❌ Erro ao excluir: {e}")
                    import traceback
                    st.code(traceback.format_exc())

    st.markdown("---")
    st.subheader("📎 DOCUMENTOS DO FUNCIONÁRIO")
    if mat_sel.strip() and not reg.empty:
        mat_atual = mat_sel.strip()
        nome_atual = val_campo("Nome")
        tipo_doc = st.selectbox("Tipo de Documento", [
            "RG", "CPF", "PIS", "Carteira de Trabalho", "Comprovante Residência",
            "Exame Admissional", "Exame Demissional", "Contrato", "Atestados",
            "Férias", "Rescisão", "Outros"
        ])
        arquivos_func = st.file_uploader("Anexar documentos", type=["pdf","doc","docx","xls","xlsx","jpg","png"], accept_multiple_files=True, key=f"up_{mat_atual}")
        if arquivos_func and st.button("SALVAR DOCUMENTOS", type="primary"):
            qtd = 0
            for arq in arquivos_func:
                nome_arq = f"{mat_atual}_{tipo_doc}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{arq.name}"
                caminho = os.path.join(PASTA_DOCS_FUNC, nome_arq)
                with open(caminho, "wb") as f: f.write(arq.read())
                dados["Docs_Funcionarios"] = pd.concat([dados["Docs_Funcionarios"], pd.DataFrame([{
                    "Matricula": mat_atual, "Nome": nome_atual, "TipoDoc": tipo_doc,
                    "NomeArquivo": arq.name, "Caminho": caminho,
                    "DataAnexado": datetime.now().strftime("%d/%m/%Y %H:%M")
                }])], ignore_index=True)
                qtd += 1
            if not salvar_dados(dados):
                st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                st.stop()
            st.success(f"✅ {qtd} documento(s) salvo(s)!")
            st.rerun()
        st.markdown("---")
        docs_func = dados["Docs_Funcionarios"][dados["Docs_Funcionarios"]["Matricula"] == mat_atual]
        if docs_func.empty: st.info("📂 Nenhum documento anexado.")
        else:
            st.markdown(f"**Total: {len(docs_func)} documento(s)**")
            for idx, doc in docs_func.iterrows():
                with st.expander(f"📄 {doc['TipoDoc']} - {doc['NomeArquivo']} | {doc['DataAnexado']}"):
                    col_v, col_b, col_e = st.columns([3,1,1])
                    with col_b:
                        if os.path.exists(doc["Caminho"]):
                            with open(doc["Caminho"], "rb") as f: st.download_button("⬇️ BAIXAR", f, file_name=doc["NomeArquivo"], key=f"dw_{idx}")
                        else:
                            st.warning("⚠️ Arquivo não encontrado no servidor.")
                    with col_e:
                        if st.button("🗑️ EXCLUIR", key=f"del_{idx}"):
                            if os.path.exists(doc["Caminho"]): os.remove(doc["Caminho"])
                            dados["Docs_Funcionarios"].drop(idx, inplace=True)
                            if not salvar_dados(dados):
                                st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                                st.stop()
                            st.rerun()
    else:
        st.info("ℹ️ Digite a Matrícula exata para ver/anexar documentos.")

# ================ ABA 2 - PAINEL ================
with aba2:
    st.subheader("📊 RESUMO GERAL")
    dados_painel = carregar_dados()
    base = dados_painel["Base_Dados"].copy()
    base["Situacao"] = base["Situacao"].fillna("").astype(str).str.strip()

    contagem = {
        "👷 Ativo": len(base[base["Situacao"] == "Ativo"]),
        "📝 Pré-cadastro": len(base[base["Situacao"] == "Pré-cadastro"]),
        "🏖️ Férias": len(base[base["Situacao"] == "Férias"]),
        "🚪 Abandono": len(base[base["Situacao"] == "Abandono"]),
        "⏹️ Término de Contrato": len(base[base["Situacao"] == "Término de Contrato"]),
        "📉 Demitido S/JC": len(base[base["Situacao"] == "Demitido S/JC"]),
        "📉 Demitido C/JC": len(base[base["Situacao"] == "Demitido C/JC"]),
        "🙋 Pedido de Conta": len(base[base["Situacao"] == "Pedido de Conta"]),
        "⚖️ Rescisão Indireta": len(base[base["Situacao"] == "Rescisão Indireta"]),
        "🏃 Desistente": len(base[base["Situacao"] == "Desistente"]),
        "🏥 Doença": len(base[base["Situacao"] == "Doença"]),
        "🚑 Acidente": len(base[base["Situacao"] == "Acidente"]),
        "🤰 Maternidade": len(base[base["Situacao"] == "Maternidade"])
    }

    cols = st.columns(3)
    for i, (rotulo, qtd) in enumerate(contagem.items()):
        cols[i % 3].metric(rotulo, qtd)

    if st.button("🔄 Atualizar Resumo"):
        st.cache_data.clear()
        st.rerun()

    st.markdown("---")
    st.subheader("📊 GRÁFICO POR SITUAÇÃO")
    if not MATPLOT:
        st.warning("⚠️ O gráfico não pode ser exibido porque a biblioteca `matplotlib` não está instalada. Adicione `matplotlib` ao `requirements.txt` e reinicie o app.")
    else:
        try:
            contagem_graf = {
                "Ativo": len(base[base["Situacao"] == "Ativo"]),
                "Pré-cadastro": len(base[base["Situacao"] == "Pré-cadastro"]),
                "Férias": len(base[base["Situacao"] == "Férias"]),
                "Abandono": len(base[base["Situacao"] == "Abandono"]),
                "Término de Contrato": len(base[base["Situacao"] == "Término de Contrato"]),
                "Demitido S/JC": len(base[base["Situacao"] == "Demitido S/JC"]),
                "Demitido C/JC": len(base[base["Situacao"] == "Demitido C/JC"]),
                "Pedido de Conta": len(base[base["Situacao"] == "Pedido de Conta"]),
                "Rescisão Indireta": len(base[base["Situacao"] == "Rescisão Indireta"]),
                "Desistente": len(base[base["Situacao"] == "Desistente"]),
                "Doença": len(base[base["Situacao"] == "Doença"]),
                "Acidente": len(base[base["Situacao"] == "Acidente"]),
                "Maternidade": len(base[base["Situacao"] == "Maternidade"])
            }
            contagem_graf = {k: v for k, v in contagem_graf.items() if v > 0}
            if contagem_graf:
                fig, ax = plt.subplots(figsize=(12, max(5, len(contagem_graf)*0.6)))
                cores = plt.cm.Set3(range(len(contagem_graf)))
                bars = ax.barh(list(contagem_graf.keys()), list(contagem_graf.values()), color=cores)
                ax.set_xlabel("Quantidade", fontsize=12)
                ax.set_title("Distribuição de Funcionários por Situação", fontsize=14, fontweight="bold", pad=15)
                for bar in bars:
                    width = bar.get_width()
                    ax.text(width + 0.3, bar.get_y() + bar.get_height()/2, str(int(width)),
                            va="center", ha="left", fontsize=10, fontweight="bold")
                ax.set_xlim(0, max(contagem_graf.values()) * 1.15)
                ax.spines["top"].set_visible(False)
                ax.spines["right"].set_visible(False)
                plt.tight_layout()
                st.pyplot(fig)
                plt.close(fig)
            else:
                st.info("ℹ️ Nenhum dado para exibir no gráfico.")
        except Exception as e:
            st.error(f"Erro ao gerar gráfico: {e}")

    st.markdown("---")
    st.subheader("🔍 Conferência Rápida - Apenas Férias")
    tab_fer = base[base["Situacao"] == "Férias"][["Matricula","Nome","Loja","DataFeriasInicio","DataRetornoFerias"]]
    if tab_fer.empty:
        st.warning("⚠️ Nenhum funcionário com situação marcada como 'Férias' no momento.")
        st.info("💡 Dica: Se a data de férias estiver preenchida mas a situação não for 'Férias', edite o cadastro e confirme se a situação está selecionada corretamente — os dados não são apagados!")
    else:
        st.dataframe(tab_fer, use_container_width=True, hide_index=True)

# ================ ABA 3 - PRAZOS E FÉRIAS ================
with aba3:
    hoje = datetime.now()
    st.subheader("⚠️ PRAZOS DE EXPERIÊNCIA PRÓXIMOS")
    tabela_exp = []
    for _, func in dados["Base_Dados"].iterrows():
        if func["Situacao"] not in ["Ativo","Pré-cadastro"]: continue
        try:
            dt_adm = datetime.strptime(str(func["Admissao"]).strip(), "%d/%m/%Y")
            dias = (hoje - dt_adm).days
            for p in [30,45,60,90]:
                if 0 <= p - dias <=10:
                    tabela_exp.append([func["Matricula"], func["Nome"], func["Loja"], f"{p} dias", f"Faltam {p-dias} dias"])
                    break
        except: pass
    st.dataframe(pd.DataFrame(tabela_exp, columns=["Matrícula","Nome","Loja","Prazo","Dias Restantes"]), use_container_width=True, hide_index=True)

    st.subheader("🗓️ FÉRIAS - POR MÊS DE ADMISSÃO")
    filtro_loja = st.selectbox("Loja", ["Todas"] + lista_lojas(), key="fl")
    filtro_mes = st.selectbox("Mês", MESES, key="fm")
    tabela_fer = []
    for _, f in dados["Base_Dados"].iterrows():
        if f["Situacao"] not in ["Ativo","Pré-cadastro","Férias"]: continue
        if filtro_loja != "Todas" and str(f["Loja"]).strip() != filtro_loja.strip(): continue
        try:
            dt = datetime.strptime(str(f["Admissao"]).strip(), "%d/%m/%Y")
            if filtro_mes != "Todos" and dt.month != [1,2,3,4,5,6,7,8,9,10,11,12][MESES.index(filtro_mes)-1]: continue
            meses = (hoje.year - dt.year)*12 + (hoje.month - dt.month) - (1 if hoje.day < dt.day else 0)
            # Mostra quem está no período 20-24 meses
            # (prestes a completar 24 meses / 2º período aquisitivo)
            if 20 <= meses <= 24:
                # Verifica status de férias
                status_fer = "🔴 Não Tirou"
                if str(f.get("Situacao","")).strip() == "Férias":
                    status_fer = "🟡 Em Férias"
                else:
                    dt_fer = str(f.get("DataFeriasInicio","")).strip()
                    ret_fer = str(f.get("DataRetornoFerias","")).strip()
                    if dt_fer and ret_fer:
                        try:
                            ret_date = datetime.strptime(ret_fer, "%d/%m/%Y").date()
                            if ret_date >= hoje.date():
                                status_fer = "🟡 Em Férias"
                            else:
                                status_fer = "🟢 Já Tirou"
                        except:
                            status_fer = "🟢 Já Tirou"
                    elif dt_fer:
                        status_fer = "🟢 Já Tirou"
                tabela_fer.append([f["Matricula"], f["Nome"], f["Loja"], f["Cargo"], f["Admissao"], f"{meses}m", status_fer])
        except: pass
    # Ordena do maior tempo para o menor (quem tem mais meses aparece primeiro — são os mais prioritários)
    tabela_fer.sort(key=lambda x: int(x[5].replace("m","")), reverse=True)
    st.dataframe(pd.DataFrame(tabela_fer, columns=["Matrícula","Nome","Loja","Cargo","Admissão","Tempo","Status Férias"]), use_container_width=True, hide_index=True)

# ================ ABA 4 - HISTÓRICO ================
with aba4:
    st.subheader("📝 HISTÓRICO GERAL")
    st.dataframe(dados["Historico"][["DataEvento","TipoEvento","Matricula","Nome","Situacao","Detalhes"]], use_container_width=True, hide_index=True)
    st.markdown("---")
    st.subheader("➕ ADICIONAR HISTÓRICO / INFORMAÇÃO INDIVIDUAL")
    # Combo com funcionários cadastrados
    funcs = dados["Base_Dados"][["Matricula","Nome"]].copy()
    funcs = funcs[funcs["Matricula"].str.strip() != ""]
    funcs = funcs.sort_values("Nome")
    func_opcoes = [f"{row['Matricula']} - {row['Nome']}" for _, row in funcs.iterrows()]
    func_sel_hist = st.selectbox("👤 Selecione o Funcionário", func_opcoes if func_opcoes else ["Nenhum funcionário cadastrado"])
    matricula_hist = func_sel_hist.split(" - ")[0] if func_sel_hist and " - " in func_sel_hist else ""
    with st.form("add_ev"):
        t,d,det = st.columns([1,1,3])
        te = t.selectbox("Tipo", ["Má conduta","Atestado","Advertência","Suspensão","Outros"])
        de = d.text_input("Data", value=datetime.now().strftime("%d/%m/%Y"))
        dee = det.text_input("Detalhes")
        if st.form_submit_button("✅ ADICIONAR") and matricula_hist.strip():
            rf = dados["Base_Dados"][dados["Base_Dados"]["Matricula"] == matricula_hist.strip()]
            if not rf.empty:
                nr = {"DataEvento":de,"TipoEvento":te,"Detalhes":dee}
                nr.update(rf.iloc[0].to_dict())
                ih = dados["Historico"].index[dados["Historico"]["Matricula"] == matricula_hist.strip()].tolist()
                if ih:
                    idx_linha = ih[0]
                    for coluna, valor in nr.items():
                        dados["Historico"].at[idx_linha, coluna] = valor
                else:
                    dados["Historico"] = pd.concat([dados["Historico"], pd.DataFrame([nr])], ignore_index=True)
                if not salvar_dados(dados):
                    st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                    st.stop()
                st.success("Adicionado!")
                st.rerun()
            else:
                st.error("Funcionário não encontrado.")

# ================ ABA 5 - RELATÓRIOS ================
with aba5:
    st.subheader("📄 RELATÓRIOS")
    rel_opcoes = [
        "Prazos Experiência","Ativos","Pré-cadastro","Férias","Afastados","Avisos",
        "Abandono","Término de Contrato","Demitido S/JC","Demitido C/JC",
        "Pedido de Conta","Rescisão Indireta","Desistente","Doença","Acidente","Maternidade",
        "Histórico","Individual"
    ]
    rel = st.selectbox("Escolha", rel_opcoes)
    if rel == "Individual":
        mr = st.text_input("Matrícula")
        if mr.strip():
            fd = dados["Base_Dados"][dados["Base_Dados"]["Matricula"] == mr.strip()]
            fh = dados["Historico"][dados["Historico"]["Matricula"] == mr.strip()]
            if fd.empty: st.error("Não encontrado")
            elif st.button("GERAR"):
                nome_arq = gerar_ficha_individual(fd, fh, mr.strip())
                with open(nome_arq, "rb") as f:
                    st.download_button("⬇️ BAIXAR", f, file_name=os.path.basename(nome_arq))
                os.remove(nome_arq)
    elif st.button("GERAR E BAIXAR"):
        if rel == "Prazos Experiência": df = pd.DataFrame(tabela_exp, columns=["Matrícula","Nome","Loja","Prazo","Dias Restantes"])
        elif rel == "Ativos": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Ativo"]
        elif rel == "Pré-cadastro": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Pré-cadastro"]
        elif rel == "Férias": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Férias"]
        elif rel == "Afastados": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"].isin(["Doença","Acidente","Maternidade"])]
        elif rel == "Avisos": df = dados["Base_Dados"][dados["Base_Dados"]["DataAvisoPrevio"].str.strip()!=""]
        elif rel == "Abandono": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Abandono"]
        elif rel == "Término de Contrato": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Término de Contrato"]
        elif rel == "Demitido S/JC": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Demitido S/JC"]
        elif rel == "Demitido C/JC": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Demitido C/JC"]
        elif rel == "Pedido de Conta": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Pedido de Conta"]
        elif rel == "Rescisão Indireta": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Rescisão Indireta"]
        elif rel == "Desistente": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Desistente"]
        elif rel == "Doença": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Doença"]
        elif rel == "Acidente": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Acidente"]
        elif rel == "Maternidade": df = dados["Base_Dados"][dados["Base_Dados"]["Situacao"] == "Maternidade"]
        else: df = dados["Historico"]
        temp_rel = os.path.join(BASE_DIR, "rel_temp.xlsx")
        with pd.ExcelWriter(temp_rel) as arq: df.to_excel(arq, index=False, sheet_name=rel)
        with open(temp_rel,"rb") as f: st.download_button("⬇️ BAIXAR", f, file_name=f"Rel_{rel.replace(' ','_')}.xlsx")
        os.remove(temp_rel)

# ================ ABA 6 - DOCUMENTOS DAS LOJAS ================
with aba6:
    st.subheader("📎 DOCUMENTOS DAS LOJAS")
    ls = lista_lojas()
    l,m,a = st.columns(3)
    sl = l.selectbox("Loja", ls)
    sm = m.selectbox("Mês", MESES)
    sa = a.selectbox("Ano", ANOS, index=ANOS.index(str(datetime.now().year)))
    st.markdown("---")
    arquivos = st.file_uploader("Anexar arquivos", type=["pdf","doc","docx","xls","xlsx","jpg","png"], accept_multiple_files=True)
    resp = st.text_input("Responsável")
    if arquivos and st.button("SALVAR TODOS", type="primary"):
        salvos = 0
        for arq in arquivos:
            nome = f"{sl}_{sm}_{sa}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{arq.name}"
            cam = os.path.join(PASTA_DOCS, nome)
            with open(cam,"wb") as f: f.write(arq.read())
            dados["Docs_Lojas"] = pd.concat([dados["Docs_Lojas"], pd.DataFrame([{
                "Loja":sl,"Mes":sm,"Ano":sa,"NomeArquivo":arq.name,"Caminho":cam,
                "DataAnexado":datetime.now().strftime("%d/%m/%Y %H:%M"),"Responsavel":resp
            }])], ignore_index=True)
            salvos += 1
        if not salvar_dados(dados):
            st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
            st.stop()
        st.success(f"✅ {salvos} arquivo(s) salvo(s)!")
        st.rerun()
    st.markdown("---")
    filt = dados["Docs_Lojas"].copy()
    if sl != "Todas": filt = filt[filt["Loja"].astype(str).str.strip()==sl]
    if sm != "Todos": filt = filt[filt["Mes"]==sm]
    filt = filt[filt["Ano"]==sa]
    if filt.empty: st.info("Nenhum documento.")
    else:
        for i,d in filt.iterrows():
            with st.expander(f"📄 {d['NomeArquivo']} | {d['Mes']}/{d['Ano']}"):
                if os.path.exists(d["Caminho"]):
                    with open(d["Caminho"],"rb") as f: st.download_button("⬇️ BAIXAR", f, file_name=d["NomeArquivo"], key=f"d{i}")
                else:
                    st.warning("⚠️ Arquivo não encontrado no servidor.")
                if st.button("🗑️ EXCLUIR", key=f"x{i}"):
                    os.remove(d["Caminho"])
                    dados["Docs_Lojas"].drop(i,inplace=True)
                    if not salvar_dados(dados):
                        st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                        st.stop()
                    st.rerun()

# ================ ABA 7 - LOJAS E CARGOS ================
with aba7:
    st.subheader("⚙️ CADASTRO DE LOJAS E CARGOS")
    dados = carregar_dados()
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**➕ Adicionar Loja**")
        nova_loja = st.text_input("Nova Loja")
        if st.button("➕ ADICIONAR LOJA", type="primary") and nova_loja.strip():
            if not dados["Auxiliares"]["Loja"].str.strip().eq(nova_loja.strip()).any():
                dados["Auxiliares"] = pd.concat([dados["Auxiliares"], pd.DataFrame([{"Loja": nova_loja.strip(), "Cargo": ""}])], ignore_index=True)
                if not salvar_dados(dados):
                    st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                    st.stop()
                st.success("✅ Loja cadastrada!")
                st.rerun()
            else: st.warning("⚠️ Já existe!")
        st.markdown("---")
        st.markdown("**🗑️ Excluir Loja**")
        lojas_existentes = sorted([str(l).strip() for l in dados["Auxiliares"]["Loja"].unique() if str(l).strip() != ""])
        loja_sel_excluir = st.selectbox("Selecione a Loja", lojas_existentes if lojas_existentes else ["Nenhuma"])
        if st.button("🗑️ EXCLUIR LOJA", type="secondary") and loja_sel_excluir != "Nenhuma":
            # Verifica se tem funcionários vinculados
            vinculados = dados["Base_Dados"][dados["Base_Dados"]["Loja"].str.strip() == loja_sel_excluir]
            if not vinculados.empty:
                st.error(f"❌ Não é possível excluir. Existem {len(vinculados)} funcionário(s) vinculado(s) a esta loja.")
            else:
                dados["Auxiliares"] = dados["Auxiliares"][dados["Auxiliares"]["Loja"].str.strip() != loja_sel_excluir]
                if not salvar_dados(dados):
                    st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                    st.stop()
                st.success(f"✅ Loja '{loja_sel_excluir}' excluída!")
                st.rerun()
    with col2:
        st.markdown("**➕ Adicionar Cargo**")
        novo_cargo = st.text_input("Novo Cargo")
        if st.button("➕ ADICIONAR CARGO", type="primary") and novo_cargo.strip():
            if not dados["Auxiliares"]["Cargo"].str.strip().eq(novo_cargo.strip()).any():
                dados["Auxiliares"] = pd.concat([dados["Auxiliares"], pd.DataFrame([{"Loja": "", "Cargo": novo_cargo.strip()}])], ignore_index=True)
                if not salvar_dados(dados):
                    st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                    st.stop()
                st.success("✅ Cargo cadastrado!")
                st.rerun()
            else: st.warning("⚠️ Já existe!")
        st.markdown("---")
        st.markdown("**🗑️ Excluir Cargo**")
        cargos_existentes = sorted([str(c).strip() for c in dados["Auxiliares"]["Cargo"].unique() if str(c).strip() != ""])
        cargo_sel_excluir = st.selectbox("Selecione o Cargo", cargos_existentes if cargos_existentes else ["Nenhum"])
        if st.button("🗑️ EXCLUIR CARGO", type="secondary") and cargo_sel_excluir != "Nenhum":
            vinculados = dados["Base_Dados"][dados["Base_Dados"]["Cargo"].str.strip() == cargo_sel_excluir]
            if not vinculados.empty:
                st.error(f"❌ Não é possível excluir. Existem {len(vinculados)} funcionário(s) com este cargo.")
            else:
                dados["Auxiliares"] = dados["Auxiliares"][dados["Auxiliares"]["Cargo"].str.strip() != cargo_sel_excluir]
                if not salvar_dados(dados):
                    st.error("❌ Não foi possível salvar os dados. Verifique se o arquivo Excel não está aberto.")
                    st.stop()
                st.success(f"✅ Cargo '{cargo_sel_excluir}' excluído!")
                st.rerun()


# ================ ABA 8 - CONTROLE DE DIÁRIAS ================
with aba8:
    st.subheader("💰 CONTROLE DE DIÁRIAS")
    st.info("ℹ️ Pagamento em até 5 dias úteis, via transferência bancária. Não permitido conta de terceiros.")

    # Upload da planilha de diárias
    st.markdown("---")
    with st.expander("📤 Como importar diárias de uma planilha externa?", expanded=False):
        st.markdown("""
        **Para que serve esta opção?**
        > Use esta opção se você já tem uma planilha Excel com diárias preenchidas e deseja importar esses dados para o sistema, sem precisar digitar tudo manualmente.
        
        **Formato esperado:**
        - A planilha pode ter uma **linha de instruções/título** na primeira linha, e o cabeçalho começando na segunda linha.
        - Ou pode ter o **cabeçalho direto na primeira linha**.
        - Colunas principais reconhecidas: LOJA, NOME COLABORADOR, CPF, DATA EXECUÇÃO, QTDE DE DIÁRIAS, VALOR UNITÁRIO, etc.
        - O sistema identifica automaticamente o formato da planilha.
        
        ⚠️ **Atenção:** ao carregar uma planilha, os dados anteriores serão substituídos pelos dados do arquivo. Faça backup se necessário.
        """)
    
    arq_diarias = st.file_uploader("Carregar planilha de Diárias (.xlsx)", type=["xlsx"], key="upload_diarias")
    if arq_diarias is not None:
        # Evita loop infinito: só processa se este arquivo ainda não foi processado nesta sessão
        chave_processado = f"_diarias_processado_{arq_diarias.name}"
        if st.session_state.get(chave_processado):
            # Já processado, apenas exibe mensagem de sucesso
            st.success(f"✅ Planilha '{arq_diarias.name}' já foi carregada. Recarregue a página (F5) se necessário.")
        else:
            # Salva temporariamente para validar
            temp_path = os.path.join(os.path.dirname(ARQUIVO_DIARIAS), "_temp_diarias.xlsx")
            with open(temp_path, "wb") as f:
                f.write(arq_diarias.read())
            
            # Valida se o arquivo tem dados legíveis
            try:
                df_test = pd.read_excel(temp_path, dtype=str, keep_default_na=False)
                if df_test.empty or df_test.shape[0] < 1:
                    st.error("❌ O arquivo parece estar vazio ou não contém dados válidos.")
                else:
                    # Move o arquivo temporário para o definitivo
                    shutil.move(temp_path, ARQUIVO_DIARIAS)
                    st.session_state[chave_processado] = True
                    st.success(f"✅ Planilha carregada com sucesso! ({df_test.shape[0]} linha(s) encontrada(s))")
                    st.info("🔄 Atualizando a página...")
                    st.rerun()
            except Exception as e:
                st.error(f"❌ Erro ao ler a planilha: {e}")
                if os.path.exists(temp_path):
                    os.remove(temp_path)

    df_diarias = carregar_diarias()
    if df_diarias.empty:
        st.warning("⚠️ Nenhuma diária cadastrada. Faça upload da planilha acima ou cadastre uma nova diária no formulário abaixo.")

    # ---------- FILTROS ----------
    col_p1, col_p2, col_p3, col_p4, col_p5 = st.columns(5)
    with col_p1: filtro_loja_d = st.selectbox("Loja", ["Todas"] + lista_lojas(), key="filtro_loja_d")
    with col_p2: filtro_mes_d = st.selectbox("Mês", MESES, key="filtro_mes_d")
    with col_p3: filtro_sem_d = st.selectbox("Semana", SEMANAS, key="filtro_sem_d")
    with col_p4: filtro_ano_d = st.selectbox("Ano", ANOS, index=ANOS.index(str(datetime.now().year)), key="filtro_ano_d")
    with col_p5: filtro_sit_d = st.selectbox("Situação", SITUACOES_DIARIA, key="filtro_sit_d")
    busca_d = st.text_input("🔍 Pesquisar por Nome ou CPF", placeholder="Digite para buscar...")

    df_filtrado = df_diarias.copy()
    if filtro_loja_d != "Todas":
        df_filtrado = df_filtrado[df_filtrado["LOJA"].astype(str).str.strip() == filtro_loja_d.strip()]
    if filtro_mes_d != "Todos":
        df_filtrado = df_filtrado[df_filtrado["MES"] == filtro_mes_d]
    if filtro_sem_d != "Todas":
        df_filtrado = df_filtrado[df_filtrado["SEMANA"] == filtro_sem_d]
    if filtro_ano_d != "Todos":
        df_filtrado = df_filtrado[df_filtrado["ANO"] == filtro_ano_d]
    if filtro_sit_d != "Todas":
        df_filtrado = df_filtrado[df_filtrado["SITUACAO"] == filtro_sit_d]
    if busca_d.strip():
        df_filtrado = df_filtrado[
            busca_palavras(df_filtrado["NOME COLABORADOR"], busca_d) |
            busca_palavras(df_filtrado["CPF"], busca_d)
        ]

    # ---------- CARDS DE RESUMO (ATUALIZADOS PELO FILTRO) ----------
    st.markdown("---")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("👥 Total de Diárias", len(df_filtrado))
    with c2:
        try:
            vfe = df_filtrado[df_filtrado["SITUACAO"] == "FALTA ENVIAR AO FINANCEIRO"]["VALOR TOTAL"].replace("", "0").astype(float).sum()
        except:
            vfe = 0
        st.metric("📤 Falta Enviar", f"R$ {vfe:,.2f}")
    with c3:
        try:
            vp = df_filtrado[df_filtrado["SITUACAO"] == "ENVIADO/PENDENTE"]["VALOR TOTAL"].replace("", "0").astype(float).sum()
        except:
            vp = 0
        st.metric("⏳ Enviado/Pendente", f"R$ {vp:,.2f}")
    with c4:
        try:
            vpg = df_filtrado[df_filtrado["SITUACAO"] == "PAGO"]["VALOR TOTAL"].replace("", "0").astype(float).sum()
        except:
            vpg = 0
        st.metric("✅ Pago", f"R$ {vpg:,.2f}")
    st.markdown("---")

    # ---------- EDITOR INLINE ----------
    if not df_filtrado.empty:
        st.markdown("**📝 Edite os dados diretamente na tabela abaixo e clique em SALVAR ALTERAÇÕES**")
        # Guarda os índices originais para salvar corretamente no DataFrame principal
        idx_original = df_filtrado.index.tolist()
        df_editable = df_filtrado.reset_index(drop=True)

        # Configurar colunas editáveis
        col_config = {
            "LOJA": st.column_config.SelectboxColumn("LOJA", options=lista_lojas(), required=True),
            "MES": st.column_config.SelectboxColumn("MÊS", options=MESES[1:], required=True),
            "SEMANA": st.column_config.SelectboxColumn("SEMANA", options=SEMANAS[1:], required=True),
            "ANO": st.column_config.SelectboxColumn("ANO", options=ANOS, required=True),
            "NOME COLABORADOR": st.column_config.TextColumn("NOME COLABORADOR", required=True),
            "CPF": st.column_config.TextColumn("CPF", required=True),
            "CARGO": st.column_config.TextColumn("CARGO"),
            "DADOS BANCÁRIOS": st.column_config.TextColumn("DADOS BANCÁRIOS"),
            "DATA EXECUCAO": st.column_config.TextColumn("DATA EXECUÇÃO"),
            "DATA PAGAMENTO": st.column_config.TextColumn("DATA PAGAMENTO"),
            "MOTIVO": st.column_config.TextColumn("MOTIVO", required=True),
            "QTDE DE DIARIAS": st.column_config.NumberColumn("QTDE", min_value=1, max_value=30, step=1, required=True),
            "VALOR UNITARIO": st.column_config.NumberColumn("VALOR UNI. (R$)", min_value=0.0, step=0.01, format="%.2f", required=True),
            "SITUACAO": st.column_config.SelectboxColumn("SITUAÇÃO", options=["FALTA ENVIAR AO FINANCEIRO", "ENVIADO/PENDENTE", "PAGO"], required=True),
            "COMPROVANTE": st.column_config.TextColumn("COMPROVANTE", disabled=True),
            "DATA CADASTRO": st.column_config.TextColumn("DATA CADASTRO", disabled=True),
            "OBSERVACAO": st.column_config.TextColumn("OBSERVAÇÃO"),
        }

        edited_df = st.data_editor(
            df_editable,
            column_config=col_config,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key="editor_diarias"
        )

        # Calcular VALOR TOTAL automaticamente
        try:
            edited_df["VALOR TOTAL"] = (edited_df["QTDE DE DIARIAS"].astype(float) * edited_df["VALOR UNITARIO"].astype(float)).apply(lambda x: f"{x:.2f}")
        except:
            pass

        col_salvar, col_excluir = st.columns([1, 1])
        with col_salvar:
            if st.button("💾 SALVAR ALTERAÇÕES", type="primary", key="salvar_diarias_editor"):
                for i, idx_orig in enumerate(idx_original):
                    if i < len(edited_df):
                        for col in df_diarias.columns:
                            if col in edited_df.columns:
                                df_diarias.at[idx_orig, col] = str(edited_df.iloc[i][col]) if col not in ["QTDE DE DIARIAS", "VALOR UNITARIO", "VALOR TOTAL"] else str(edited_df.iloc[i][col])
                # Se houver linhas novas (mais que o original)
                if len(edited_df) > len(idx_original):
                    for i in range(len(idx_original), len(edited_df)):
                        nova_linha = {col: "" for col in df_diarias.columns}
                        for col in edited_df.columns:
                            nova_linha[col] = str(edited_df.iloc[i][col])
                        nova_linha["DATA CADASTRO"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                        if not nova_linha.get("VALOR TOTAL"):
                            try:
                                q = float(edited_df.iloc[i]["QTDE DE DIARIAS"])
                                v = float(edited_df.iloc[i]["VALOR UNITARIO"])
                                nova_linha["VALOR TOTAL"] = f"{q * v:.2f}"
                            except:
                                nova_linha["VALOR TOTAL"] = ""
                        df_diarias = pd.concat([df_diarias, pd.DataFrame([nova_linha])], ignore_index=True)
                # Se houver linhas removidas
                if len(edited_df) < len(idx_original):
                    remover = idx_original[len(edited_df):]
                    for idx_rm in remover:
                        comp = str(df_diarias.at[idx_rm, "COMPROVANTE"])
                        if comp and os.path.exists(comp):
                            os.remove(comp)
                    df_diarias.drop(index=remover, inplace=True)
                    df_diarias.reset_index(drop=True, inplace=True)
                salvar_diarias(df_diarias)
                st.success("✅ Alterações salvas com sucesso!")
                st.rerun()

        with col_excluir:
            st.markdown("**🗑️ Excluir linhas selecionadas:** marque a caixa na primeira coluna da tabela acima e depois clique abaixo.")
            if st.button("🗑️ EXCLUIR SELECIONADOS", key="excluir_diarias_editor"):
                st.info("Para excluir, delete as linhas diretamente na tabela usando a tecla Delete ou botão de lixeira do editor, depois clique em SALVAR ALTERAÇÕES.")

    else:
        st.info("Nenhuma diária encontrada com os filtros aplicados.")

    # ---------- NOVA DIÁRIA (CADASTRO RÁPIDO) ----------
    st.markdown("---")
    st.subheader("➕ CADASTRAR NOVA DIÁRIA")

    # Gerenciamento de itens de diária (fora do form para permitir botões dinâmicos)
    if "itens_diaria" not in st.session_state:
        st.session_state.itens_diaria = [{"qtde": 1, "valor": 0.0}]

    st.markdown("**📋 Itens de Diária** — adicione quantos itens quiser com quantidades e valores diferentes:")
    for i, item in enumerate(st.session_state.itens_diaria):
        cols = st.columns([2, 2, 1])
        with cols[0]:
            item["qtde"] = st.number_input(
                f"Qtde Item {i+1}", min_value=1, max_value=30, value=int(item["qtde"]),
                key=f"qtde_item_d_{i}"
            )
        with cols[1]:
            item["valor"] = st.number_input(
                f"Valor Unit. Item {i+1} (R$)", min_value=0.0, format="%.2f", value=float(item["valor"]),
                key=f"valor_item_d_{i}"
            )
        with cols[2]:
            st.markdown("<br>", unsafe_allow_html=True)
            if len(st.session_state.itens_diaria) > 1:
                if st.button("🗑️ Remover", key=f"rm_item_d_{i}"):
                    st.session_state.itens_diaria.pop(i)
                    st.rerun()

    if st.button("➕ Adicionar Item de Diária", key="add_item_diaria"):
        st.session_state.itens_diaria.append({"qtde": 1, "valor": 0.0})
        st.rerun()

    # Calcular totais
    total_qtde = sum(int(item["qtde"]) for item in st.session_state.itens_diaria)
    total_valor = sum(int(item["qtde"]) * float(item["valor"]) for item in st.session_state.itens_diaria)
    detalhe_itens = "  |  ".join(
        [f"{int(item['qtde'])}x R$ {float(item['valor']):.2f}" for item in st.session_state.itens_diaria]
    )
    st.info(f"**Resumo:** {detalhe_itens}  →  **Total: {total_qtde} diárias = R$ {total_valor:,.2f}**")

    with st.form("nova_diaria", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            loja_d = st.selectbox("Loja *", lista_lojas(), key="nova_loja_d")
            mes_d = st.selectbox("Mês *", MESES[1:], key="nova_mes_d")
            semana_d = st.selectbox("Semana *", SEMANAS[1:], key="nova_sem_d")
            ano_d = st.selectbox("Ano *", ANOS, index=ANOS.index(str(datetime.now().year)), key="nova_ano_d")
        with c2:
            nome_d = st.text_input("Nome do Colaborador *", key="nova_nome_d")
            cpf_d = st.text_input("CPF *", key="nova_cpf_d")
            cargo_d = st.text_input("Cargo", key="nova_cargo_d")
            dados_bancarios_d = st.text_input("Dados Bancários (PIX / Banco / Ag / CC)", key="nova_dados_bancarios_d")
            data_exec_d = st.text_input("Data da Execução (DD/MM/AAAA)", key="nova_data_exec_d")
        with c3:
            data_pag_d = st.text_input("Data de Pagamento (DD/MM/AAAA)", key="nova_data_pag_d")
            motivo_d = st.text_input("Motivo *", key="nova_motivo_d")
            situacao_d = st.selectbox("Situação *", ["FALTA ENVIAR AO FINANCEIRO", "ENVIADO/PENDENTE", "PAGO"], key="nova_sit_d")
        observacao_d = st.text_area("Observação (erros de pagamento, conta em nome de terceiro, conta incorreta, etc.)", key="nova_obs_d")
        submitted = st.form_submit_button("💾 SALVAR DIÁRIA", type="primary")
        if submitted:
            erros = []
            if not loja_d.strip(): erros.append("Loja")
            if not mes_d.strip(): erros.append("Mês")
            if not semana_d.strip(): erros.append("Semana")
            if not ano_d.strip(): erros.append("Ano")
            if not nome_d.strip(): erros.append("Nome do Colaborador")
            if not cpf_d.strip(): erros.append("CPF")
            if not motivo_d.strip(): erros.append("Motivo")
            if total_qtde <= 0: erros.append("Qtde total deve ser > 0")
            if total_valor <= 0: erros.append("Valor total deve ser > 0")
            if erros:
                st.error("❌ Campos obrigatórios: " + ", ".join(erros))
            else:
                # Junta os valores unitários em uma string descritiva
                valores_desc = ", ".join([f"{int(item['qtde'])}x R${float(item['valor']):.2f}" for item in st.session_state.itens_diaria])
                valor_medio = total_valor / total_qtde if total_qtde > 0 else 0
                nova_linha = {
                    "LOJA": loja_d,
                    "MES": mes_d,
                    "SEMANA": semana_d,
                    "ANO": ano_d,
                    "NOME COLABORADOR": nome_d.strip().upper(),
                    "CPF": cpf_d.strip(),
                    "CARGO": cargo_d.strip().upper(),
                    "DADOS BANCÁRIOS": dados_bancarios_d.strip().upper(),
                    "DATA EXECUCAO": data_exec_d.strip(),
                    "DATA PAGAMENTO": data_pag_d.strip(),
                    "MOTIVO": motivo_d.strip().upper(),
                    "QTDE DE DIARIAS": str(total_qtde),
                    "VALOR UNITARIO": f"{valor_medio:.2f}",
                    "VALOR TOTAL": f"{total_valor:.2f}",
                    "SITUACAO": situacao_d,
                    "DATA CADASTRO": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "COMPROVANTE": "",
                    "OBSERVACAO": (observacao_d.strip().upper() + " | ITENS: " + valores_desc) if observacao_d.strip() else "ITENS: " + valores_desc
                }
                df_diarias = pd.concat([df_diarias, pd.DataFrame([nova_linha])], ignore_index=True)
                salvar_diarias(df_diarias)
                st.session_state.itens_diaria = [{"qtde": 1, "valor": 0.0}]
                st.success("✅ Diária cadastrada com sucesso!")
                st.rerun()

    # ---------- GERENCIAR COMPROVANTES ----------
    st.markdown("---")
    st.subheader("📎 GERENCIAR COMPROVANTES DE PAGAMENTO")
    if not df_diarias.empty:
        # Dropdown para selecionar diária
        opcoes_diaria = [f"[{i}] {row['NOME COLABORADOR']} | {row['LOJA']} | {row['MES']}/{row['ANO']} | R$ {row['VALOR TOTAL']}" for i, row in df_diarias.iterrows()]
        sel_diaria = st.selectbox("Selecione a diária", options=range(len(opcoes_diaria)), format_func=lambda x: opcoes_diaria[x], key="sel_comp_diaria")
        if sel_diaria is not None:
            idx_comp = df_diarias.index[sel_diaria]
            comp_atual = str(df_diarias.at[idx_comp, "COMPROVANTE"])
            if comp_atual and os.path.exists(comp_atual):
                st.success(f"✅ Comprovante anexado: {os.path.basename(comp_atual)}")
                with open(comp_atual, "rb") as fc:
                    st.download_button("⬇️ Baixar Comprovante", fc, file_name=os.path.basename(comp_atual), key=f"dl_comp_{idx_comp}")
                if st.button("🗑️ Remover Comprovante", key=f"rm_comp_{idx_comp}"):
                    os.remove(comp_atual)
                    df_diarias.at[idx_comp, "COMPROVANTE"] = ""
                    salvar_diarias(df_diarias)
                    st.success("Comprovante removido!")
                    st.rerun()
            else:
                st.info("Nenhum comprovante anexado para esta diária.")
            arq_comp = st.file_uploader("Anexar comprovante (PDF, JPG, PNG)", type=["pdf", "jpg", "png"], key=f"up_comp_{idx_comp}")
            if arq_comp and st.button("📤 ENVIAR COMPROVANTE", type="primary", key=f"btn_comp_{idx_comp}"):
                ext = os.path.splitext(arq_comp.name)[1]
                nome_comp = f"{df_diarias.at[idx_comp, 'CPF']}_{df_diarias.at[idx_comp, 'MES']}_{df_diarias.at[idx_comp, 'ANO']}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
                cam_comp = os.path.join(PASTA_COMPROVANTES, nome_comp)
                with open(cam_comp, "wb") as f: f.write(arq_comp.read())
                df_diarias.at[idx_comp, "COMPROVANTE"] = cam_comp
                salvar_diarias(df_diarias)
                st.success("✅ Comprovante anexado!")
                st.rerun()
    else:
        st.info("Nenhuma diária cadastrada.")

    # ---------- EXPORTAR ----------
    st.markdown("---")
    st.subheader("📤 EXPORTAR PARA EXCEL")
    if not df_filtrado.empty:
        nome_arq = os.path.join(BASE_DIR, f"Diarias_{filtro_loja_d}_{filtro_mes_d}_{filtro_ano_d}.xlsx".replace("/", "-").replace(" ", "_"))
        exportar_diarias_formatado(df_filtrado, nome_arq)
        with open(nome_arq, "rb") as f:
            st.download_button("⬇️ BAIXAR EXCEL", f, file_name=os.path.basename(nome_arq))
        os.remove(nome_arq)
    else:
        st.info("Filtre os dados para exportar.")




def gerar_pdf_rota(tipo, origem, destino, distancia, tempo_info, custo_ida=None, custo_volta=None, litros=None, preco_litro=None, consumo=None):
    """Gera um PDF com o resumo da rota e custos usando reportlab."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from io import BytesIO
    except ImportError as e:
        st.error(f"Erro ao importar reportlab: {e}. Verifique se 'reportlab' está no requirements.txt e reinicie o app.")
        return None

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 2 * cm
    x = margin
    y = height - margin
    line_height = 14

    def hex_color(r, g, b):
        return colors.Color(r / 255, g / 255, b / 255)

    def draw_text(text, size=11, bold=False, italic=False, color=None, x_pos=None, y_pos=None, align="left"):
        nonlocal y
        font = "Helvetica-Bold" if bold else ("Helvetica-Oblique" if italic else "Helvetica")
        c.setFont(font, size)
        c.setFillColor(color if color else colors.black)
        px = x_pos if x_pos is not None else x
        py = y_pos if y_pos is not None else y
        if align == "center":
            c.drawCentredString(width / 2, py, text)
        elif align == "right":
            c.drawRightString(px, py, text)
        else:
            c.drawString(px, py, text)
        if y_pos is None:
            y -= size + 4
        return py

    def draw_line(y_pos=None):
        nonlocal y
        py = y_pos if y_pos is not None else y
        c.setStrokeColor(colors.lightgrey)
        c.line(margin, py, width - margin, py)
        if y_pos is None:
            y -= 8
        return py

    # Header
    draw_text("RESUMO DE VIAGEM - RH COMPLETO", size=16, bold=True, color=hex_color(33, 37, 41), align="center")
    y -= 4
    draw_line()
    y -= 8

    # Tipo de transporte
    draw_text(f"Meio de Transporte: {tipo.upper()}", size=12, bold=True, color=hex_color(0, 102, 204))
    y -= 4

    # Origem e Destino
    draw_text("ORIGEM:", size=11, bold=True, color=hex_color(33, 37, 41))
    draw_text(origem, size=11)
    y -= 2
    draw_text("DESTINO:", size=11, bold=True, color=hex_color(33, 37, 41))
    draw_text(destino, size=11)
    y -= 8
    draw_line()
    y -= 8

    # Resumo da rota
    draw_text("RESUMO DA ROTA", size=12, bold=True, color=hex_color(33, 37, 41))
    y -= 2

    c.setFont("Helvetica", 11)
    c.setFillColor(hex_color(50, 50, 50))
    c.drawString(x, y, "Distancia:")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x + 140, y, f"{distancia:.1f} km")
    y -= line_height

    c.setFont("Helvetica", 11)
    c.drawString(x, y, "Tempo Estimado:")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x + 140, y, tempo_info)
    y -= 22

    # Custo de combustível (apenas carro)
    if tipo == "Carro" and custo_ida is not None:
        draw_line()
        y -= 8
        draw_text("CUSTO ESTIMADO DE COMBUSTIVEL", size=12, bold=True, color=hex_color(33, 37, 41))
        y -= 2

        c.setFont("Helvetica", 11)
        c.setFillColor(hex_color(50, 50, 50))
        c.drawString(x, y, "Preco/Litro:")
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x + 140, y, f"R$ {preco_litro:.2f}")
        y -= line_height

        c.setFont("Helvetica", 11)
        c.drawString(x, y, "Consumo do Veiculo:")
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x + 140, y, f"{consumo:.1f} km/L")
        y -= line_height

        c.setFont("Helvetica", 11)
        c.drawString(x, y, "Litros Necessarios (ida):")
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x + 140, y, f"{litros:.1f} L")
        y -= 20

        # Tabela custos
        box_h = 20
        c.setFillColor(hex_color(230, 245, 255))
        c.rect(x, y - box_h + 4, 140, box_h, fill=1, stroke=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x + 4, y - box_h + 14, "Custo Ida:")
        c.drawRightString(x + 136, y - box_h + 14, f"R$ {custo_ida:,.2f}")
        y -= box_h + 4

        c.setFillColor(hex_color(255, 235, 230))
        c.rect(x, y - box_h + 4, 140, box_h, fill=1, stroke=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x + 4, y - box_h + 14, "Custo Ida + Volta:")
        c.drawRightString(x + 136, y - box_h + 14, f"R$ {custo_volta:,.2f}")
        y -= box_h + 12

    # Observações
    draw_line()
    y -= 8
    c.setFont("Helvetica-Oblique", 9)
    c.setFillColor(hex_color(100, 100, 100))
    obs_text = (
        "Observacoes: Os valores de combustivel sao estimados e podem variar conforme o trajeto real, condicoes de transito e precos dos postos. O calculo de pedagio deve ser consultado separadamente."
        if tipo == "Carro"
        else "Observacoes: A distancia exibida e em linha reta (trajeto aereo aproximado). O tempo inclui estimativa de taxi, decolagem e pouso. Valores de passagens devem ser consultados em companhias aereas."
    )
    text_obj = c.beginText(x, y)
    text_obj.setFont("Helvetica-Oblique", 9)
    max_width = width - 2 * margin
    words = obs_text.split(" ")
    line = ""
    for word in words:
        test = line + word + " "
        if c.stringWidth(test, "Helvetica-Oblique", 9) < max_width:
            line = test
        else:
            text_obj.textLine(line.strip())
            line = word + " "
    if line:
        text_obj.textLine(line.strip())
    c.drawText(text_obj)

    # Footer
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(hex_color(128, 128, 128))
    c.drawCentredString(width / 2, 30, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage()
    c.save()
    return buffer.getvalue()


# ================ ABA 9 - GUIA VIAGEM ================
with aba9:
    st.subheader("🗺️ GUIA DE VIAGEM")
    sub_aba_rotas, sub_aba_viagens = st.tabs(["🧭 Calculadora de Rotas", "📝 Registro de Viagens"])

    with sub_aba_rotas:
        st.info("Pesquise origem e destino para ver distância, tempo estimado, custo e rota no mapa.")

        # --- Dados dos combos ---
        PAÍSES = [
            "Brasil", "Argentina", "Bolívia", "Chile", "Colômbia", "Equador", "Guiana",
            "Paraguai", "Peru", "Suriname", "Uruguai", "Venezuela", "Estados Unidos",
            "Canadá", "México", "Portugal", "Espanha", "França", "Alemanha", "Itália",
            "Reino Unido", "Japão", "China", "Austrália", "Nova Zelândia", "África do Sul",
            "Índia", "Rússia", "Ucrânia", "Turquia", "Emirados Árabes Unidos"
        ]
        ESTADOS_BR = [
            "Acre (AC)", "Alagoas (AL)", "Amapá (AP)", "Amazonas (AM)", "Bahia (BA)",
            "Ceará (CE)", "Distrito Federal (DF)", "Espírito Santo (ES)", "Goiás (GO)",
            "Maranhão (MA)", "Mato Grosso (MT)", "Mato Grosso do Sul (MS)", "Minas Gerais (MG)",
            "Pará (PA)", "Paraíba (PB)", "Paraná (PR)", "Pernambuco (PE)", "Piauí (PI)",
            "Rio de Janeiro (RJ)", "Rio Grande do Norte (RN)", "Rio Grande do Sul (RS)",
            "Rondônia (RO)", "Roraima (RR)", "Santa Catarina (SC)", "São Paulo (SP)",
            "Sergipe (SE)", "Tocantins (TO)"
        ]
        ESTADOS_US = [
            "Alabama (AL)", "Alaska (AK)", "Arizona (AZ)", "Arkansas (AR)", "Califórnia (CA)",
            "Carolina do Norte (NC)", "Carolina do Sul (SC)", "Colorado (CO)", "Connecticut (CT)",
            "Dakota do Norte (ND)", "Dakota do Sul (SD)", "Delaware (DE)", "Flórida (FL)",
            "Geórgia (GA)", "Havaí (HI)", "Idaho (ID)", "Illinois (IL)", "Indiana (IN)",
            "Iowa (IA)", "Kansas (KS)", "Kentucky (KY)", "Louisiana (LA)", "Maine (ME)",
            "Maryland (MD)", "Massachusetts (MA)", "Michigan (MI)", "Minnesota (MN)",
            "Mississippi (MS)", "Missouri (MO)", "Montana (MT)", "Nebraska (NE)", "Nevada (NV)",
            "Nova Hampshire (NH)", "Nova Jersey (NJ)", "Nova York (NY)", "Novo México (NM)",
            "Ohio (OH)", "Oklahoma (OK)", "Oregon (OR)", "Pensilvânia (PA)", "Rhode Island (RI)",
            "Tennessee (TN)", "Texas (TX)", "Utah (UT)", "Vermont (VT)", "Virgínia (VA)",
            "Virgínia Ocidental (WV)", "Washington (WA)", "Wisconsin (WI)", "Wyoming (WY)"
        ]
        ESTADOS_MX = [
            "Aguascalientes", "Baja California", "Baja California Sur", "Campeche", "Chiapas",
            "Chihuahua", "Coahuila", "Colima", "Durango", "Guanajuato", "Guerrero", "Hidalgo",
            "Jalisco", "México", "Michoacán", "Morelos", "Nayarit", "Nuevo León", "Oaxaca",
            "Puebla", "Querétaro", "Quintana Roo", "San Luis Potosí", "Sinaloa", "Sonora",
            "Tabasco", "Tamaulipas", "Tlaxcala", "Veracruz", "Yucatán", "Zacatecas",
            "Cidade do México"
        ]

        @st.cache_data(ttl=86400, show_spinner=False)
        def buscar_cidades_ibge(uf_sigla):
            """Busca lista de municípios do IBGE pela sigla da UF."""
            try:
                url = f"https://servicodados.ibge.gov.br/api/v1/localidades/estados/{uf_sigla}/municipios"
                resp = requests.get(url, timeout=15)
                if resp.status_code == 200:
                    dados = resp.json()
                    cidades = sorted([m["nome"] for m in dados])
                    return cidades
            except Exception:
                pass
            return []

        def input_endereco(label, key_prefix):
            """Monta os inputs de endereço com combos de país, estado e cidade."""
            st.markdown(f"**{label}**")
            pais = st.selectbox(f"🌍 País", PAÍSES, key=f"{key_prefix}_pais")
            if pais == "Brasil":
                estado = st.selectbox(f"🏛️ Estado", ESTADOS_BR, key=f"{key_prefix}_estado")
                estado_limp = estado.split(" (")[0] if "(" in estado else estado
                uf_sigla = estado.split("(")[1].replace(")", "").strip() if "(" in estado else ""
                cidades = buscar_cidades_ibge(uf_sigla) if uf_sigla else []
                if cidades:
                    cidade = st.selectbox(f"🏙️ Cidade", cidades, key=f"{key_prefix}_cidade")
                else:
                    cidade = st.text_input(f"🏙️ Cidade", placeholder="Ex: Belém", key=f"{key_prefix}_cidade")
            elif pais == "Estados Unidos":
                estado = st.selectbox(f"🏛️ Estado", ESTADOS_US, key=f"{key_prefix}_estado")
                estado_limp = estado.split(" (")[0] if "(" in estado else estado
                cidade = st.text_input(f"🏙️ Cidade", placeholder="Ex: Nova York", key=f"{key_prefix}_cidade")
            elif pais == "México":
                estado = st.selectbox(f"🏛️ Estado", ESTADOS_MX, key=f"{key_prefix}_estado")
                estado_limp = estado
                cidade = st.text_input(f"🏙️ Cidade", placeholder="Ex: Cidade do México", key=f"{key_prefix}_cidade")
            else:
                estado_limp = st.text_input(f"🏛️ Estado / Província", key=f"{key_prefix}_estado")
                cidade = st.text_input(f"🏙️ Cidade", placeholder="Ex: Belém", key=f"{key_prefix}_cidade")
            endereco = f"{cidade}, {estado_limp}, {pais}" if str(cidade).strip() and str(estado_limp).strip() else ""
            return endereco, pais

        col_o, col_d = st.columns(2)
        with col_o:
            origem_str, pais_o = input_endereco("📍 ORIGEM", "orig")
        with col_d:
            destino_str, pais_d = input_endereco("📍 DESTINO", "dest")

        transporte = st.selectbox("🚗 Meio de Transporte", ["Carro", "Avião"])

        if st.button("🚀 CALCULAR ROTA", type="primary"):
            if not origem_str.strip() or not destino_str.strip():
                st.warning("⚠️ Preencha cidade, estado e país tanto na origem quanto no destino.")
            else:
                with st.spinner("Consultando rota..."):
                    lat1, lon1 = geocodificar(origem_str.strip())
                    lat2, lon2 = geocodificar(destino_str.strip())
                if lat1 is None or lat2 is None:
                    st.error("❌ Não foi possível localizar um ou ambos os endereços. Tente incluir a cidade mais próxima ou verificar a grafia.")
                else:
                    st.success("✅ Pontos localizados com sucesso!")
                    st.markdown("---")
                    cidade_origem = origem_str.split(",")[0].strip()
                    cidade_destino = destino_str.split(",")[0].strip()

                    # ---- CARRO ----
                    if transporte == "Carro":
                        distancia, tempo, geometria = calcular_rota(lat1, lon1, lat2, lon2)
                        if distancia is None:
                            st.error("❌ Não foi possível calcular a rota de carro. Tente novamente mais tarde.")
                        else:
                            st.subheader("📊 RESUMO DA ROTA (CARRO)")
                            c1, c2, c3 = st.columns(3)
                            with c1:
                                st.metric("📏 Distância", f"{distancia:.1f} km")
                            with c2:
                                horas = int(tempo // 60)
                                mins = int(tempo % 60)
                                st.metric("⏱️ Tempo Estimado", f"{horas}h {mins}min")
                            with c3:
                                st.metric("💰 Pedágio", "Consultar via app")

                            # Combustível
                            st.markdown("---")
                            st.subheader("⛽ CUSTO ESTIMADO DE COMBUSTÍVEL")
                            cc1, cc2 = st.columns(2)
                            with cc1:
                                preco_litro = st.number_input("Preço/Litro (R$)", min_value=0.0, value=5.89, step=0.01, format="%.2f", key="preco_carro")
                            with cc2:
                                consumo_km_l = st.number_input("Consumo (km/L)", min_value=0.1, value=10.0, step=0.1, format="%.1f", key="consumo_carro")
                            if preco_litro > 0 and consumo_km_l > 0:
                                litros = distancia / consumo_km_l
                                custo_ida = litros * preco_litro
                                custo_ida_volta = custo_ida * 2
                                cb1, cb2 = st.columns(2)
                                with cb1:
                                    st.metric("⛽ Ida", f"R$ {custo_ida:,.2f}")
                                with cb2:
                                    st.metric("⛽ Ida + Volta", f"R$ {custo_ida_volta:,.2f}")
                                st.info(f"💡 Litros necessários (ida): **{litros:.1f} L** | Preço/L: R$ {preco_litro:.2f} | Consumo: {consumo_km_l:.1f} km/L")

                            # PDF
                            st.markdown("---")
                            pdf_bytes = gerar_pdf_rota(
                                tipo="Carro",
                                origem=origem_str,
                                destino=destino_str,
                                distancia=distancia,
                                tempo_info=f"{horas}h {mins}min",
                                custo_ida=custo_ida if (preco_litro > 0 and consumo_km_l > 0) else None,
                                custo_volta=custo_ida_volta if (preco_litro > 0 and consumo_km_l > 0) else None,
                                litros=litros if (preco_litro > 0 and consumo_km_l > 0) else None,
                                preco_litro=preco_litro if (preco_litro > 0 and consumo_km_l > 0) else None,
                                consumo=consumo_km_l if (preco_litro > 0 and consumo_km_l > 0) else None,
                            )
                            if pdf_bytes:
                                st.download_button(
                                    label="📄 BAIXAR RESUMO EM PDF",
                                    data=pdf_bytes,
                                    file_name=f"Resumo_Viagem_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                                    mime="application/pdf"
                                )
                            else:
                                st.warning("⚠️ Não foi possível gerar o PDF. Verifique se a biblioteca `reportlab` está instalada.")

                            # Mapa com linha
                            st.markdown("---")
                            st.subheader("🗺️ Visualização da Rota")
                            try:
                                import pydeck as pdk
                                coords = geometria["coordinates"]
                                path_coords = coords  # já é [lon, lat]
                                mid_lat = (lat1 + lat2) / 2
                                mid_lon = (lon1 + lon2) / 2
                                zoom_lvl = calcular_zoom(distancia)

                                path_layer = pdk.Layer(
                                    "PathLayer",
                                    data=[{"path": path_coords, "color": [255, 60, 0]}],
                                    get_path="path",
                                    get_color="color",
                                    width_scale=20,
                                    width_min_pixels=4,
                                )
                                scatter_layer = pdk.Layer(
                                    "ScatterplotLayer",
                                    data=[
                                        {"position": [lon1, lat1], "color": [0, 200, 0]},
                                        {"position": [lon2, lat2], "color": [255, 0, 0]},
                                    ],
                                    get_position="position",
                                    get_color="color",
                                    get_radius=20000,
                                    radius_min_pixels=8,
                                    radius_max_pixels=25,
                                )
                                text_layer = pdk.Layer(
                                    "TextLayer",
                                    data=[
                                        {"position": [lon1, lat1], "text": cidade_origem, "color": [0, 200, 0]},
                                        {"position": [lon2, lat2], "text": cidade_destino, "color": [255, 0, 0]},
                                    ],
                                    get_position="position",
                                    get_text="text",
                                    get_color="color",
                                    get_size=18,
                                    get_text_anchor="middle",
                                    get_alignment_baseline="bottom",
                                    size_units="pixels",
                                )
                                view_state = pdk.ViewState(
                                    latitude=mid_lat, longitude=mid_lon,
                                    zoom=zoom_lvl, pitch=0
                                )
                                st.pydeck_chart(pdk.Deck(
                                    layers=[path_layer, scatter_layer, text_layer],
                                    initial_view_state=view_state,
                                    tooltip={"text": "Rota de carro"},
                                    height=800,
                                ))
                                st.caption("🟢 Origem  |  🔴 Destino  |  🟠 Linha = rota por estrada")
                            except Exception as e:
                                st.warning(f"Não foi possível exibir o mapa: {e}")

                            # Instruções passo a passo
                            st.markdown("---")
                            st.subheader("📝 Instruções de Rota (passo a passo)")
                            try:
                                url = f"http://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}"
                                params = {"overview": "false", "steps": "true"}
                                resp = requests.get(url, params=params, timeout=20)
                                dados_inst = resp.json()
                                if dados_inst.get("routes"):
                                    legs = dados_inst["routes"][0]["legs"][0]
                                    passos = []
                                    for step in legs.get("steps", []):
                                        nome = step.get("name", "")
                                        dist = step.get("distance", 0)
                                        instr = step.get("maneuver", {}).get("type", "continue")
                                        passos.append(f"• {instr.upper()}: siga em **{nome}** por `{dist/1000:.1f} km`")
                                    if passos:
                                        for p in passos[:25]:
                                            st.markdown(p)
                                        if len(passos) > 25:
                                            st.info(f"... e mais {len(passos)-25} instruções.")
                                    else:
                                        st.info("Nenhuma instrução detalhada disponível.")
                                else:
                                    st.info("Instruções não disponíveis.")
                            except Exception as e:
                                st.info(f"Instruções não disponíveis: {e}")

                    # ---- AVIÃO ----
                    else:
                        distancia = haversine(lat1, lon1, lat2, lon2)
                        tempo_voo_min = (distancia / 850) * 60  # 850 km/h média
                        tempo_total_min = tempo_voo_min + 90   # +1h30 taxi
                        st.subheader("📊 RESUMO DA ROTA (AVIÃO)")
                        a1, a2, a3 = st.columns(3)
                        with a1:
                            st.metric("📏 Distância (linha reta)", f"{distancia:.1f} km")
                        with a2:
                            horas_v = int(tempo_total_min // 60)
                            mins_v = int(tempo_total_min % 60)
                            st.metric("⏱️ Tempo Estimado", f"{horas_v}h {mins_v}min")
                        with a3:
                            st.metric("✈️ Veloc. Média", "~850 km/h")
                        st.info("💡 O tempo inclui aproximadamente 1h30 de taxi, decolagem e pouso.")

                        # PDF
                        st.markdown("---")
                        pdf_bytes = gerar_pdf_rota(
                            tipo="Avião",
                            origem=origem_str,
                            destino=destino_str,
                            distancia=distancia,
                            tempo_info=f"{horas_v}h {mins_v}min",
                        )
                        if pdf_bytes:
                            st.download_button(
                                label="📄 BAIXAR RESUMO EM PDF",
                                data=pdf_bytes,
                                file_name=f"Resumo_Viagem_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                                mime="application/pdf"
                            )
                        else:
                            st.warning("⚠️ Não foi possível gerar o PDF. Verifique se a biblioteca `reportlab` está instalada.")

                        # Mapa com linha reta
                        st.markdown("---")
                        st.subheader("🗺️ Visualização da Rota")
                        try:
                            import pydeck as pdk
                            path_coords = [[lon1, lat1], [lon2, lat2]]
                            mid_lat = (lat1 + lat2) / 2
                            mid_lon = (lon1 + lon2) / 2
                            zoom_lvl = calcular_zoom(distancia)

                            path_layer = pdk.Layer(
                                "PathLayer",
                                data=[{"path": path_coords, "color": [0, 100, 255]}],
                                get_path="path",
                                get_color="color",
                                width_scale=20,
                                width_min_pixels=4,
                            )
                            scatter_layer = pdk.Layer(
                                "ScatterplotLayer",
                                data=[
                                    {"position": [lon1, lat1], "color": [0, 200, 0]},
                                    {"position": [lon2, lat2], "color": [255, 0, 0]},
                                ],
                                get_position="position",
                                get_color="color",
                                get_radius=20000,
                                radius_min_pixels=8,
                                radius_max_pixels=25,
                            )
                            text_layer = pdk.Layer(
                                "TextLayer",
                                data=[
                                    {"position": [lon1, lat1], "text": cidade_origem, "color": [0, 200, 0]},
                                    {"position": [lon2, lat2], "text": cidade_destino, "color": [255, 0, 0]},
                                ],
                                get_position="position",
                                get_text="text",
                                get_color="color",
                                get_size=18,
                                get_text_anchor="middle",
                                get_alignment_baseline="bottom",
                                size_units="pixels",
                            )
                            view_state = pdk.ViewState(
                                latitude=mid_lat, longitude=mid_lon,
                                zoom=zoom_lvl, pitch=0
                            )
                            st.pydeck_chart(pdk.Deck(
                                layers=[path_layer, scatter_layer, text_layer],
                                initial_view_state=view_state,
                                tooltip={"text": "Rota aérea (linha reta)"},
                                height=800,
                            ))
                            st.caption("🟢 Origem  |  🔴 Destino  |  🔵 Linha = trajeto aéreo aproximado")
                        except Exception as e:
                            st.warning(f"Não foi possível exibir o mapa: {e}")


    with sub_aba_viagens:
        st.markdown("### 📝 REGISTRO DE VIAGENS")

        df_viagens = carregar_viagens()

        # --- CADASTRO ---
        with st.expander("➕ Cadastrar Nova Viagem", expanded=False):
            with st.form("form_cadastro_viagem", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    num_viagem = st.text_input("Número da Viagem *", key="cad_num_viagem")
                    colaborador_v = st.text_input("Colaborador *", key="cad_colab_viagem")
                    loja_v = st.selectbox("Loja", lista_lojas(), key="cad_loja_viagem")
                with c2:
                    origem_v = st.text_input("Origem", key="cad_origem_viagem")
                    destino_v = st.text_input("Destino", key="cad_destino_viagem")
                    motivo_v = st.text_input("Motivo", key="cad_motivo_viagem")
                with c3:
                    data_saida_v = st.text_input("Data Saída (DD/MM/AAAA)", key="cad_dt_saida_v")
                    data_retorno_v = st.text_input("Data Retorno (DD/MM/AAAA)", key="cad_dt_retorno_v")
                    valor_liberado_v = st.number_input("Valor Liberado (R$)", min_value=0.0, step=0.01, format="%.2f", key="cad_valor_lib_v")

                observacoes_v = st.text_area("Observações / Prestação de Conta", key="cad_obs_viagem")

                submitted_v = st.form_submit_button("💾 SALVAR VIAGEM", type="primary")
                if submitted_v:
                    if not num_viagem.strip() or not colaborador_v.strip():
                        st.error("❌ Número da Viagem e Colaborador são obrigatórios!")
                    else:
                        df_v = carregar_viagens()
                        nums_existentes = df_v["NUMERO_VIAGEM"].astype(str).str.strip()
                        if num_viagem.strip() in nums_existentes.values:
                            st.error("❌ Já existe uma viagem com este número!")
                        else:
                            novo_id = "1"
                            if not df_v.empty:
                                try:
                                    ids_numericos = pd.to_numeric(df_v["ID"], errors="coerce").dropna()
                                    if not ids_numericos.empty:
                                        novo_id = str(int(ids_numericos.max()) + 1)
                                except Exception:
                                    pass
                            nova_viagem = {
                                "ID": novo_id,
                                "NUMERO_VIAGEM": num_viagem.strip(),
                                "COLABORADOR": colaborador_v.strip().upper(),
                                "LOJA": loja_v,
                                "ORIGEM": origem_v.strip().upper(),
                                "DESTINO": destino_v.strip().upper(),
                                "MOTIVO": motivo_v.strip().upper(),
                                "DATA_SAIDA": data_saida_v.strip(),
                                "DATA_RETORNO": data_retorno_v.strip(),
                                "VALOR_LIBERADO": f"{float(valor_liberado_v):.2f}",
                                "TOTAL_GASTO": "0.00",
                                "RESTANTE": f"{float(valor_liberado_v):.2f}",
                                "STATUS": "Planejada",
                                "OBSERVACOES": observacoes_v.strip().upper(),
                                "DATA_CADASTRO": datetime.now().strftime("%d/%m/%Y %H:%M")
                            }
                            df_v = pd.concat([df_v, pd.DataFrame([nova_viagem])], ignore_index=True)
                            salvar_viagens(df_v)
                            st.success("✅ Viagem cadastrada com sucesso!")
                            time.sleep(0.5)
                            st.rerun()

        # --- FILTROS ---
        st.markdown("---")
        st.markdown("### 📋 HISTÓRICO DE VIAGENS")

        fv1, fv2, fv3, fv4 = st.columns(4)
        with fv1:
            f_num_v = st.text_input("🔍 Nº Viagem", key="filtro_num_viagem")
        with fv2:
            f_colab_v = st.text_input("🔍 Colaborador", key="filtro_colab_viagem")
        with fv3:
            f_loja_v = st.selectbox("Loja", ["Todas"] + lista_lojas(), key="filtro_loja_viagem")
        with fv4:
            f_status_v = st.selectbox("Status", ["Todos", "Planejada", "Em Andamento", "Concluída", "Cancelada"], key="filtro_status_viagem")

        df_v_filt = df_viagens.copy()
        if f_num_v.strip():
            df_v_filt = df_v_filt[df_v_filt["NUMERO_VIAGEM"].astype(str).str.contains(f_num_v.strip(), case=False, na=False)]
        if f_colab_v.strip():
            df_v_filt = df_v_filt[busca_palavras(df_v_filt["COLABORADOR"], f_colab_v)]
        if f_loja_v != "Todas":
            df_v_filt = df_v_filt[df_v_filt["LOJA"] == f_loja_v]
        if f_status_v != "Todos":
            df_v_filt = df_v_filt[df_v_filt["STATUS"] == f_status_v]

        st.markdown(f"**📊 Total: {len(df_v_filt)} viagem(ns) encontrada(s)**")

        if df_v_filt.empty:
            st.info("ℹ️ Nenhuma viagem encontrada com os filtros aplicados.")
        else:
            col_config_v = {
                "ID": st.column_config.TextColumn("ID", disabled=True),
                "NUMERO_VIAGEM": st.column_config.TextColumn("Nº VIAGEM", disabled=True),
                "COLABORADOR": st.column_config.TextColumn("COLABORADOR"),
                "LOJA": st.column_config.SelectboxColumn("LOJA", options=lista_lojas()),
                "ORIGEM": st.column_config.TextColumn("ORIGEM"),
                "DESTINO": st.column_config.TextColumn("DESTINO"),
                "MOTIVO": st.column_config.TextColumn("MOTIVO"),
                "DATA_SAIDA": st.column_config.TextColumn("DATA SAÍDA"),
                "DATA_RETORNO": st.column_config.TextColumn("DATA RETORNO"),
                "VALOR_LIBERADO": st.column_config.NumberColumn("VALOR LIBERADO (R$)", min_value=0.0, format="%.2f"),
                "TOTAL_GASTO": st.column_config.NumberColumn("TOTAL GASTO (R$)", min_value=0.0, format="%.2f"),
                "RESTANTE": st.column_config.NumberColumn("RESTANTE (R$)", disabled=True, format="%.2f"),
                "STATUS": st.column_config.SelectboxColumn("STATUS", options=["Planejada", "Em Andamento", "Concluída", "Cancelada"]),
                "OBSERVACOES": st.column_config.TextColumn("OBSERVAÇÕES / PRESTAÇÃO DE CONTA"),
                "DATA_CADASTRO": st.column_config.TextColumn("DATA CADASTRO", disabled=True),
            }

            idx_original_v = df_v_filt.index.tolist()
            df_v_editable = df_v_filt.reset_index(drop=True)

            edited_v = st.data_editor(
                df_v_editable,
                column_config=col_config_v,
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                key="editor_viagens"
            )

            try:
                edited_v["RESTANTE"] = (edited_v["VALOR_LIBERADO"].astype(float) - edited_v["TOTAL_GASTO"].astype(float)).apply(lambda x: f"{x:.2f}")
            except Exception:
                pass

            try:
                total_lib = edited_v["VALOR_LIBERADO"].astype(float).sum()
                total_gasto = edited_v["TOTAL_GASTO"].astype(float).sum()
                total_rest = edited_v["RESTANTE"].astype(float).sum()
                r1, r2, r3 = st.columns(3)
                with r1:
                    st.metric("💰 Total Liberado", f"R$ {total_lib:,.2f}")
                with r2:
                    st.metric("💸 Total Gasto", f"R$ {total_gasto:,.2f}")
                with r3:
                    st.metric("📊 Total Restante", f"R$ {total_rest:,.2f}")
            except Exception:
                pass

            col_sv, col_ev = st.columns([1, 1])
            with col_sv:
                if st.button("💾 SALVAR ALTERAÇÕES", type="primary", key="salvar_viagens_btn"):
                    df_v_main = carregar_viagens()
                    for i, idx_orig in enumerate(idx_original_v):
                        if i < len(edited_v):
                            for col in df_v_main.columns:
                                if col in edited_v.columns:
                                    df_v_main.at[idx_orig, col] = str(edited_v.iloc[i][col])
                            try:
                                vl = float(str(df_v_main.at[idx_orig, "VALOR_LIBERADO"]).replace(",", "."))
                                tg = float(str(df_v_main.at[idx_orig, "TOTAL_GASTO"]).replace(",", "."))
                                df_v_main.at[idx_orig, "RESTANTE"] = f"{vl - tg:.2f}"
                            except Exception:
                                pass
                    if len(edited_v) > len(idx_original_v):
                        for i in range(len(idx_original_v), len(edited_v)):
                            nova_linha = {col: "" for col in df_v_main.columns}
                            for col in edited_v.columns:
                                nova_linha[col] = str(edited_v.iloc[i][col])
                            novo_id = "1"
                            if not df_v_main.empty:
                                try:
                                    ids_num = pd.to_numeric(df_v_main["ID"], errors="coerce").dropna()
                                    if not ids_num.empty:
                                        novo_id = str(int(ids_num.max()) + 1)
                                except Exception:
                                    pass
                            nova_linha["ID"] = novo_id
                            nova_linha["DATA_CADASTRO"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                            try:
                                vl = float(str(nova_linha.get("VALOR_LIBERADO", "0")).replace(",", "."))
                                tg = float(str(nova_linha.get("TOTAL_GASTO", "0")).replace(",", "."))
                                nova_linha["RESTANTE"] = f"{vl - tg:.2f}"
                            except Exception:
                                nova_linha["RESTANTE"] = "0.00"
                            df_v_main = pd.concat([df_v_main, pd.DataFrame([nova_linha])], ignore_index=True)
                    if len(edited_v) < len(idx_original_v):
                        remover = idx_original_v[len(edited_v):]
                        df_v_main.drop(index=remover, inplace=True)
                        df_v_main.reset_index(drop=True, inplace=True)
                    salvar_viagens(df_v_main)
                    st.success("✅ Alterações salvas com sucesso!")
                    st.rerun()

            with col_ev:
                st.markdown("**🗑️ Para excluir:** delete as linhas na tabela (tecla Delete) e clique em SALVAR ALTERAÇÕES.")


        # --- RESUMO, GRÁFICOS E EXPORTAÇÃO ---
        st.markdown("---")
        st.markdown("### 📊 RESUMO E ANÁLISE DE VIAGENS")

        if df_viagens.empty:
            st.info("ℹ️ Nenhuma viagem registrada para gerar resumos e gráficos.")
        else:
            try:
                df_v_num = df_viagens.copy()
                for col in ["VALOR_LIBERADO", "TOTAL_GASTO", "RESTANTE"]:
                    df_v_num[col] = pd.to_numeric(df_v_num[col].astype(str).str.replace(",", "."), errors="coerce").fillna(0)

                total_viagens = len(df_v_num)
                total_liberado = df_v_num["VALOR_LIBERADO"].sum()
                total_gasto = df_v_num["TOTAL_GASTO"].sum()
                total_restante = df_v_num["RESTANTE"].sum()
                media_gasto = df_v_num["TOTAL_GASTO"].mean()

                r1, r2, r3, r4, r5 = st.columns(5)
                with r1:
                    st.metric("🧳 Viagens", f"{total_viagens}")
                with r2:
                    st.metric("💰 Liberado", f"R$ {total_liberado:,.2f}")
                with r3:
                    st.metric("💸 Gasto", f"R$ {total_gasto:,.2f}")
                with r4:
                    st.metric("📊 Restante", f"R$ {total_restante:,.2f}")
                with r5:
                    st.metric("📈 Média Gasto", f"R$ {media_gasto:,.2f}")

                st.markdown("---")
                st.markdown("#### 📈 Gráficos")

                g1, g2 = st.columns(2)
                with g1:
                    status_counts = df_v_num["STATUS"].value_counts()
                    if not status_counts.empty:
                        fig1, ax1 = plt.subplots(figsize=(4.5, 3.5))
                        colors = {"Planejada": "#3498db", "Em Andamento": "#f1c40f", "Concluída": "#2ecc71", "Cancelada": "#e74c3c"}
                        pie_colors = [colors.get(s, "#95a5a6") for s in status_counts.index]
                        ax1.pie(status_counts.values, labels=status_counts.index, autopct="%1.1f%%", colors=pie_colors, startangle=90)
                        ax1.set_title("Distribuição por Status", fontsize=10, fontweight="bold")
                        plt.tight_layout()
                        st.pyplot(fig1)
                        plt.close(fig1)

                with g2:
                    top_colab = df_v_num.groupby("COLABORADOR")["TOTAL_GASTO"].sum().sort_values(ascending=True).tail(10)
                    if not top_colab.empty:
                        fig2, ax2 = plt.subplots(figsize=(4.5, 3.5))
                        top_colab.plot(kind="barh", ax=ax2, color="#2ecc71")
                        ax2.set_title("Top 10 Colaboradores - Total Gasto", fontsize=10, fontweight="bold")
                        ax2.set_xlabel("R$", fontsize=8)
                        plt.tight_layout()
                        st.pyplot(fig2)
                        plt.close(fig2)

                g3, g4 = st.columns(2)
                with g3:
                    loja_gasto = df_v_num.groupby("LOJA")["TOTAL_GASTO"].sum().sort_values(ascending=False).head(10)
                    if not loja_gasto.empty:
                        fig3, ax3 = plt.subplots(figsize=(4.5, 3.5))
                        loja_gasto.plot(kind="bar", ax=ax3, color="#3498db")
                        ax3.set_title("Top 10 Lojas - Total Gasto", fontsize=10, fontweight="bold")
                        ax3.set_ylabel("R$", fontsize=8)
                        ax3.tick_params(axis="x", rotation=45, labelsize=7)
                        plt.tight_layout()
                        st.pyplot(fig3)
                        plt.close(fig3)

                with g4:
                    try:
                        df_v_num["MES"] = pd.to_datetime(df_v_num["DATA_CADASTRO"], format="%d/%m/%Y %H:%M", errors="coerce").dt.to_period("M").astype(str)
                        mes_counts = df_v_num["MES"].value_counts().sort_index().tail(12)
                        if not mes_counts.empty:
                            fig4, ax4 = plt.subplots(figsize=(4.5, 3.5))
                            mes_counts.plot(kind="line", ax=ax4, marker="o", color="#e74c3c")
                            ax4.set_title("Viagens por Mês (últimos 12)", fontsize=10, fontweight="bold")
                            ax4.set_ylabel("Quantidade", fontsize=8)
                            ax4.tick_params(axis="x", rotation=45, labelsize=7)
                            plt.tight_layout()
                            st.pyplot(fig4)
                            plt.close(fig4)
                    except Exception:
                        pass

                st.markdown("---")
                st.markdown("#### 📥 EXPORTAR DADOS")
                e1, e2 = st.columns(2)
                with e1:
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                        df_v_num.to_excel(writer, sheet_name="Viagens", index=False)
                    st.download_button(
                        label="📊 EXPORTAR EXCEL",
                        data=excel_buffer.getvalue(),
                        file_name=f"Resumo_Viagens_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                with e2:
                    try:
                        from reportlab.lib.pagesizes import letter
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors

                        pdf_buffer = io.BytesIO()
                        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
                        elements = []
                        styles = getSampleStyleSheet()
                        elements.append(Paragraph("<b>RESUMO DE VIAGENS</b>", styles["Title"]))
                        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
                        elements.append(Spacer(1, 12))

                        # Resumo em tabela
                        resumo_data = [
                            ["Total de Viagens", str(total_viagens)],
                            ["Total Liberado", f"R$ {total_liberado:,.2f}"],
                            ["Total Gasto", f"R$ {total_gasto:,.2f}"],
                            ["Total Restante", f"R$ {total_restante:,.2f}"],
                            ["Média de Gasto", f"R$ {media_gasto:,.2f}"],
                        ]
                        t_resumo = Table(resumo_data, colWidths=[200, 200])
                        t_resumo.setStyle(TableStyle([
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ]))
                        elements.append(t_resumo)
                        elements.append(Spacer(1, 20))

                        # Tabela de viagens (top 30)
                        top_df = df_v_num.sort_values("TOTAL_GASTO", ascending=False).head(30)
                        top_df = top_df[["NUMERO_VIAGEM", "COLABORADOR", "LOJA", "DESTINO", "VALOR_LIBERADO", "TOTAL_GASTO", "RESTANTE", "STATUS"]]
                        top_df = top_df.fillna("")
                        table_data = [top_df.columns.tolist()] + top_df.values.tolist()
                        t_viagens = Table(table_data, repeatRows=1)
                        t_viagens.setStyle(TableStyle([
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 9),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ]))
                        elements.append(t_viagens)
                        doc.build(elements)
                        st.download_button(
                            label="📄 EXPORTAR PDF RESUMO",
                            data=pdf_buffer.getvalue(),
                            file_name=f"Resumo_Viagens_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                        )
                    except Exception as e_pdf:
                        st.error(f"Erro ao gerar PDF: {e_pdf}")

            except Exception as e:
                st.error(f"Erro ao gerar resumos: {e}")

# ================ ABA 10 - BACKUP / RESTAURAÇÃO ================
with aba10:
    st.subheader("💾 BACKUP E RESTAURAÇÃO")
    st.warning("⚠️ **IMPORTANTE:** No Streamlit Cloud, os dados são salvos localmente e podem ser perdidos ao atualizar o código. Use esta aba para fazer backup antes de qualquer atualização!")

    st.markdown("---")
    st.markdown("### 📥 FAZER BACKUP (Exportar tudo)")
    st.info("Clique no botão abaixo para baixar um arquivo ZIP com todos os dados: planilhas Excel, documentos das lojas, documentos dos funcionários, fotos e comprovantes de diárias.")

    if st.button("💾 GERAR BACKUP COMPLETO", type="primary"):
        with st.spinner("Compactando todos os dados..."):
            zip_buffer = criar_backup_zip()
        st.success("✅ Backup gerado com sucesso!")
        st.download_button(
            label="⬇️ BAIXAR ARQUIVO ZIP",
            data=zip_buffer,
            file_name=f"Backup_RH_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
            mime="application/zip"
        )

    st.markdown("---")
    st.markdown("### 📤 RESTAURAR BACKUP (Importar tudo)")
    st.info("Selecione o arquivo ZIP de backup para restaurar todos os dados. **ATENÇÃO:** Isso irá substituir os dados atuais.")

    arquivo_backup = st.file_uploader("Selecione o arquivo ZIP de backup", type=["zip"], key="upload_backup")

    if arquivo_backup is not None:
        st.warning("⚠️ Confirme para restaurar os dados do backup. Os dados atuais serão substituídos.")
        if st.button("🔄 RESTAURAR BACKUP", type="primary"):
            with st.spinner("Restaurando dados..."):
                try:
                    arquivos_restaurados = restaurar_backup_zip(arquivo_backup)
                    st.cache_data.clear()
                except Exception as e:
                    st.error(f"❌ Erro ao restaurar backup: {e}")
                    st.stop()
            st.success(f"✅ Backup restaurado com sucesso! {len(arquivos_restaurados)} arquivo(s) restaurado(s).")
            st.info("🔄 A página será atualizada em instantes para carregar os dados restaurados...")
            time.sleep(2)
            st.rerun()

    st.markdown("---")
    st.markdown("### 📂 Arquivos Atuais no Sistema")
    col_b1, col_b2, col_b3, col_b4 = st.columns(4)
    with col_b1:
        st.metric("📎 Docs Lojas", len(os.listdir(PASTA_DOCS)) if os.path.exists(PASTA_DOCS) else 0)
    with col_b2:
        st.metric("📄 Docs Funcionários", len(os.listdir(PASTA_DOCS_FUNC)) if os.path.exists(PASTA_DOCS_FUNC) else 0)
    with col_b3:
        st.metric("🖼️ Fotos", len(os.listdir(PASTA_FOTOS)) if os.path.exists(PASTA_FOTOS) else 0)
    with col_b4:
        st.metric("📎 Comprovantes", len(os.listdir(PASTA_COMPROVANTES)) if os.path.exists(PASTA_COMPROVANTES) else 0)

    st.markdown("---")
    st.caption("Dica: Faça backup periodicamente ou sempre antes de atualizar o código no Streamlit Cloud.")


# ================ ABA 11 - TRADUTOR ================

def traduzir_texto(texto, origem="auto", destino="pt"):
    """Traduz texto usando a API gratuita do Google Translate via requests."""
    try:
        url = "https://translate.googleapis.com/translate_a/single"
        params = {
            "client": "gtx",
            "sl": origem,
            "tl": destino,
            "dt": "t",
            "q": texto,
        }
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        dados = resp.json()
        if dados and isinstance(dados, list) and dados[0]:
            partes = [p[0] for p in dados[0] if isinstance(p, list) and len(p) > 0]
            return "".join(partes)
    except Exception:
        pass
    return None


IDIOMAS = {
    "Português": "pt",
    "Inglês": "en",
    "Espanhol": "es",
    "Francês": "fr",
    "Alemão": "de",
    "Italiano": "it",
    "Chinês (Simplificado)": "zh-CN",
    "Japonês": "ja",
    "Coreano": "ko",
    "Árabe": "ar",
    "Russo": "ru",
    "Holandês": "nl",
    "Turco": "tr",
    "Hindi": "hi",
    "Polonês": "pl",
}

with aba11:
    st.subheader("🌐 TRADUTOR MULTILÍNGUE")
    st.info("Traduza textos, gere áudio a partir de textos e transcreva arquivos de áudio para texto.")

    sub_aba_txt, sub_aba_tts, sub_aba_stt = st.tabs(["📝 Texto → Texto", "🔊 Texto → Áudio", "🎤 Áudio → Texto"])

    # ---------- SUB-ABA 1: TEXTO → TEXTO ----------
    with sub_aba_txt:
        st.markdown("### 📝 Tradução de Texto")
        c1, c2 = st.columns(2)
        with c1:
            idioma_origem = st.selectbox("Idioma de Origem", ["Auto-detectar"] + list(IDIOMAS.keys()), key="trad_origem")
        with c2:
            idioma_destino = st.selectbox("Idioma de Destino", list(IDIOMAS.keys()), index=1, key="trad_destino")

        texto_origem = st.text_area("✍️ Digite o texto para traduzir", height=150, key="trad_texto_origem")

        if st.button("🔄 TRADUZIR", type="primary", use_container_width=True):
            if not texto_origem.strip():
                st.warning("⚠️ Digite um texto para traduzir.")
            else:
                with st.spinner("Traduzindo..."):
                    cod_origem = "auto" if idioma_origem == "Auto-detectar" else IDIOMAS.get(idioma_origem, "auto")
                    cod_destino = IDIOMAS.get(idioma_destino, "pt")
                    resultado = traduzir_texto(texto_origem, origem=cod_origem, destino=cod_destino)
                if resultado:
                    st.success("✅ Tradução concluída!")
                    if "traducoes_historico" not in st.session_state:
                        st.session_state["traducoes_historico"] = []
                    st.session_state["traducoes_historico"].append({
                        "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                        "origem": idioma_origem,
                        "destino": idioma_destino,
                        "texto_original": texto_origem,
                        "texto_traduzido": resultado,
                    })
                else:
                    st.error("❌ Não foi possível traduzir. Verifique a conexão com a internet.")

        # Exibe histórico de traduções acumuladas
        if "traducoes_historico" in st.session_state and st.session_state["traducoes_historico"]:
            st.markdown("---")
            st.markdown("### 📚 Histórico de Traduções")
            for i, item in enumerate(reversed(st.session_state["traducoes_historico"])):
                idx_real = len(st.session_state["traducoes_historico"]) - 1 - i
                with st.container(border=True):
                    st.caption(f"🕐 {item['timestamp']} | {item['origem']} → {item['destino']}")
                    c_orig, c_dest = st.columns(2)
                    with c_orig:
                        st.text_area("Texto Original", value=item["texto_original"], height=100, key=f"trad_hist_orig_{idx_real}", disabled=True)
                    with c_dest:
                        st.text_area("Tradução", value=item["texto_traduzido"], height=100, key=f"trad_hist_dest_{idx_real}", disabled=True)
                    st.download_button(
                        label="📥 Baixar esta tradução (.txt)",
                        data=item["texto_traduzido"].encode("utf-8"),
                        file_name=f"traducao_{item['timestamp'].replace('/', '').replace(' ', '_').replace(':', '')}.txt",
                        mime="text/plain",
                        key=f"trad_hist_down_{idx_real}",
                    )
            if st.button("🗑️ Limpar Histórico de Traduções", key="trad_limpar_hist"):
                st.session_state["traducoes_historico"] = []
                st.rerun()

    # ---------- SUB-ABA 2: TEXTO → ÁUDIO (TTS) ----------
    with sub_aba_tts:
        st.markdown("### 🔊 Texto para Áudio (TTS)")
        st.caption("Converta texto em fala. Você pode traduzir o texto antes de gerar o áudio.")

        # Inicializa histórico de TTS
        if "tts_historico" not in st.session_state:
            st.session_state["tts_historico"] = []

        c1, c2 = st.columns(2)
        with c1:
            idioma_origem_tts = st.selectbox("Idioma do Texto (Origem)", ["Auto-detectar"] + list(IDIOMAS.keys()), key="tts_idioma_origem")
        with c2:
            idioma_destino_tts = st.selectbox("Idioma do Áudio (Destino)", list(IDIOMAS.keys()), index=0, key="tts_idioma_destino")

        traduzir_antes = st.checkbox("🔄 Traduzir o texto antes de gerar o áudio", value=True, key="tts_traduzir_antes")
        texto_tts = st.text_area("✍️ Digite o texto", height=150, key="tts_texto")

        col_gerar, col_limpar = st.columns(2)
        with col_gerar:
            gerar_audio = st.button("🔊 GERAR ÁUDIO", type="primary", use_container_width=True, key="tts_btn_gerar")
        with col_limpar:
            if st.button("🗑️ Limpar Histórico de Áudio", use_container_width=True, key="tts_btn_limpar"):
                st.session_state["tts_historico"] = []
                st.rerun()

        if gerar_audio:
            if not texto_tts.strip():
                st.warning("⚠️ Digite um texto para converter.")
            else:
                try:
                    from gtts import gTTS
                except ImportError:
                    st.error("❌ A biblioteca `gTTS` não está instalada. Execute: `pip install gtts`")
                    st.stop()

                texto_final = texto_tts
                cod_tts = IDIOMAS.get(idioma_destino_tts, "pt")

                # Traduzir se necessário
                if traduzir_antes:
                    with st.spinner("Traduzindo texto..."):
                        cod_origem = "auto" if idioma_origem_tts == "Auto-detectar" else IDIOMAS.get(idioma_origem_tts, "auto")
                        texto_traduzido = traduzir_texto(texto_tts, origem=cod_origem, destino=cod_tts)
                    if texto_traduzido:
                        texto_final = texto_traduzido
                        st.success(f"✅ Texto traduzido de {idioma_origem_tts} para {idioma_destino_tts}")
                        st.text_area("Texto que será convertido em áudio", value=texto_final, height=100, disabled=True, key="tts_texto_convertido")
                    else:
                        st.warning("⚠️ Não foi possível traduzir. Usando texto original.")
                        texto_final = texto_tts

                with st.spinner("Gerando áudio..."):
                    try:
                        tts = gTTS(text=texto_final, lang=cod_tts, slow=False)
                        mp3_buffer = io.BytesIO()
                        tts.write_to_fp(mp3_buffer)
                        mp3_buffer.seek(0)
                        st.success("✅ Áudio gerado com sucesso!")
                        st.audio(mp3_buffer, format="audio/mp3")
                        st.download_button(
                            label="📥 Baixar Áudio (.mp3)",
                            data=mp3_buffer.getvalue(),
                            file_name=f"audio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp3",
                            mime="audio/mpeg",
                            key="tts_download_audio"
                        )
                        # Adiciona ao histórico
                        st.session_state["tts_historico"].append({
                            "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            "idioma_origem": idioma_origem_tts,
                            "idioma_destino": idioma_destino_tts,
                            "texto_original": texto_tts,
                            "texto_convertido": texto_final,
                            "audio_bytes": mp3_buffer.getvalue(),
                        })
                    except Exception as e_gtts:
                        st.error(f"❌ Erro ao gerar áudio com gTTS.")
                        st.info("ℹ️ O serviço Google TTS pode estar indisponível temporariamente ou o idioma selecionado pode não ser suportado no momento. Tente outro idioma ou tente novamente mais tarde.")

        # Exibe histórico de áudio gerado
        if st.session_state["tts_historico"]:
            st.markdown("---")
            st.markdown("### 📚 Histórico de Áudios Gerados")
            for i, item in enumerate(reversed(st.session_state["tts_historico"])):
                idx_real = len(st.session_state["tts_historico"]) - 1 - i
                with st.container(border=True):
                    st.caption(f"🕐 {item['timestamp']} | {item['idioma_origem']} → {item['idioma_destino']}")
                    st.text_area("Texto convertido", value=item["texto_convertido"], height=80, disabled=True, key=f"tts_hist_texto_{idx_real}")
                    st.audio(io.BytesIO(item["audio_bytes"]), format="audio/mp3")
                    st.download_button(
                        label="📥 Baixar Áudio",
                        data=item["audio_bytes"],
                        file_name=f"audio_{item['timestamp'].replace('/', '').replace(' ', '_').replace(':', '')}.mp3",
                        mime="audio/mpeg",
                        key=f"tts_hist_down_{idx_real}",
                    )

    # ---------- SUB-ABA 3: ÁUDIO → TEXTO (STT) + CORREÇÃO + TTS ----------
    with sub_aba_stt:
        st.markdown("### 🎤 Fale, Transcreva, Traduza e Corrija")
        st.caption("Grave áudio pelo microfone ou envie um arquivo. O sistema transcreve, traduz e permite correções por escrito ou por nova fala.")

        # --- Inicializa histórico ---
        if "stt_historico" not in st.session_state:
            st.session_state["stt_historico"] = []
        if "stt_ultimo_audio_bytes" not in st.session_state:
            st.session_state["stt_ultimo_audio_bytes"] = None
        if "stt_ultimo_audio_ext" not in st.session_state:
            st.session_state["stt_ultimo_audio_ext"] = ".wav"

        # --- Entrada de áudio: Microfone + Upload ---
        col_mic, col_up = st.columns(2)
        with col_mic:
            st.markdown("**🎙️ Gravar pelo Microfone**")
            audio_gravado = None
            try:
                audio_gravado = st.audio_input("Clique para gravar sua fala", key="stt_mic")
            except Exception:
                st.info("ℹ️ Para gravação direta pelo navegador, atualize o Streamlit para versão 1.37 ou superior.")
        with col_up:
            st.markdown("**📁 Ou envie um arquivo**")
            arquivo_audio = st.file_uploader("Selecione .wav, .mp3, .ogg ou .flac", type=["wav", "mp3", "ogg", "flac"], key="stt_upload")

        audio_final = audio_gravado if audio_gravado is not None else arquivo_audio

        # Salva bytes do áudio no session_state para reutilização
        if audio_final is not None:
            if hasattr(audio_final, 'name'):
                st.session_state["stt_ultimo_audio_bytes"] = audio_final.getvalue()
                st.session_state["stt_ultimo_audio_ext"] = os.path.splitext(audio_final.name)[1].lower()
            else:
                st.session_state["stt_ultimo_audio_bytes"] = audio_final.getvalue() if hasattr(audio_final, 'getvalue') else audio_final
                st.session_state["stt_ultimo_audio_ext"] = ".wav"

        c1, c2 = st.columns(2)
        with c1:
            idioma_audio = st.selectbox("Idioma do Áudio", list(IDIOMAS.keys()), index=0, key="stt_idioma")
        with c2:
            idioma_trad = st.selectbox("Traduzir para", list(IDIOMAS.keys()), index=0, key="stt_idioma_trad")

        col_proc, col_trad = st.columns(2)
        with col_proc:
            processar_audio = st.button("🎤 TRANSCREVER E TRADUZIR", type="primary", use_container_width=True, key="stt_btn_processar")
        with col_trad:
            traduzir_novamente = st.button("🔄 TRADUZIR ÚLTIMO ÁUDIO P/ OUTRO IDIOMA", type="secondary", use_container_width=True, key="stt_btn_retraduzir")

        # --- AÇÃO: TRANSCREVER E TRADUZIR ---
        if processar_audio:
            if st.session_state["stt_ultimo_audio_bytes"] is None:
                st.warning("⚠️ Grave ou envie um áudio primeiro.")
            else:
                try:
                    import speech_recognition as sr
                except ImportError:
                    st.error("❌ A biblioteca `SpeechRecognition` não está instalada. Execute: `pip install SpeechRecognition`")
                    st.stop()

                with st.spinner("Processando áudio..."):
                    audio_bytes = st.session_state["stt_ultimo_audio_bytes"]
                    ext = st.session_state["stt_ultimo_audio_ext"]
                    tmp_path = f"/tmp/stt_audio_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
                    with open(tmp_path, "wb") as f_audio:
                        f_audio.write(audio_bytes)

                    # Converter para wav se necessário
                    wav_path = tmp_path
                    if ext != ".wav":
                        try:
                            from pydub import AudioSegment
                            wav_path = tmp_path.replace(ext, ".wav")
                            audio_seg = AudioSegment.from_file(tmp_path, format=ext.replace(".", ""))
                            audio_seg.export(wav_path, format="wav")
                        except ImportError:
                            st.error("❌ Para arquivos MP3/OGG/FLAC é necessário instalar: `pip install pydub` (e ter ffmpeg instalado no sistema).")
                            os.remove(tmp_path)
                            st.stop()
                        except Exception as e_conv:
                            st.error(f"❌ Erro ao converter áudio: {e_conv}")
                            if os.path.exists(tmp_path):
                                os.remove(tmp_path)
                            st.stop()

                    recognizer = sr.Recognizer()
                    try:
                        with sr.AudioFile(wav_path) as source:
                            audio_data = recognizer.record(source)
                        cod_stt = IDIOMAS.get(idioma_audio, "pt")
                        texto_transcrito = recognizer.recognize_google(audio_data, language=cod_stt)

                        # Traduzir automaticamente
                        cod_trad = IDIOMAS.get(idioma_trad, "pt")
                        texto_traduzido = ""
                        if texto_transcrito.strip():
                            texto_traduzido = traduzir_texto(texto_transcrito, origem=cod_stt, destino=cod_trad)
                            if texto_traduzido is None:
                                texto_traduzido = ""

                        # Guarda transcrição atual para re-traduzir depois
                        st.session_state["stt_texto_transcrito_atual"] = texto_transcrito
                        st.session_state["stt_cod_idioma_audio_atual"] = cod_stt

                        # Adiciona ao histórico
                        st.session_state["stt_historico"].append({
                            "id": len(st.session_state["stt_historico"]),
                            "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            "idioma_audio": idioma_audio,
                            "cod_audio": cod_stt,
                            "idioma_trad": idioma_trad,
                            "cod_trad": cod_trad,
                            "texto_transcrito": texto_transcrito,
                            "texto_traduzido": texto_traduzido,
                        })
                        st.success("✅ Áudio transcrito e traduzido com sucesso!")

                    except sr.UnknownValueError:
                        st.error("❌ Não foi possível entender o áudio. Verifique a qualidade do arquivo ou fale mais próximo do microfone.")
                    except sr.RequestError as e_req:
                        st.error(f"❌ Erro no serviço de reconhecimento: {e_req}")
                    except Exception as e_all:
                        st.error(f"❌ Erro ao processar áudio: {e_all}")
                    finally:
                        if os.path.exists(tmp_path):
                            os.remove(tmp_path)
                        if os.path.exists(wav_path) and wav_path != tmp_path:
                            os.remove(wav_path)

        # --- AÇÃO: TRADUZIR ÚLTIMO ÁUDIO PARA OUTRO IDIOMA ---
        if traduzir_novamente:
            if "stt_texto_transcrito_atual" not in st.session_state or not st.session_state["stt_texto_transcrito_atual"]:
                st.warning("⚠️ Transcreva um áudio primeiro.")
            else:
                cod_stt = st.session_state.get("stt_cod_idioma_audio_atual", "pt")
                cod_trad = IDIOMAS.get(idioma_trad, "pt")
                texto_transcrito = st.session_state["stt_texto_transcrito_atual"]
                with st.spinner("Traduzindo..."):
                    texto_traduzido = traduzir_texto(texto_transcrito, origem=cod_stt, destino=cod_trad)
                    if texto_traduzido is None:
                        texto_traduzido = ""
                st.session_state["stt_historico"].append({
                    "id": len(st.session_state["stt_historico"]),
                    "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    "idioma_audio": idioma_audio,
                    "cod_audio": cod_stt,
                    "idioma_trad": idioma_trad,
                    "cod_trad": cod_trad,
                    "texto_transcrito": texto_transcrito,
                    "texto_traduzido": texto_traduzido,
                })
                st.success("✅ Nova tradução adicionada ao histórico!")

        # --- SEÇÃO DE HISTÓRICO E CORREÇÃO ---
        if st.session_state["stt_historico"]:
            st.markdown("---")
            st.markdown("### 📚 Histórico de Transcrições e Traduções")

            for item in reversed(st.session_state["stt_historico"]):
                idx = item["id"]
                with st.container(border=True):
                    st.caption(f"🕐 {item['timestamp']} | Áudio: {item['idioma_audio']} → Tradução: {item['idioma_trad']}")

                    col_res1, col_res2 = st.columns(2)
                    with col_res1:
                        st.markdown("**📝 Texto Original Transcrito**")
                        st.info(item["texto_transcrito"])
                    with col_res2:
                        st.markdown("**📋 Tradução (editável)**")
                        # Usa chave única baseada no ID para evitar conflito de widgets
                        chave_edit = f"stt_edit_{idx}"
                        texto_editado = st.text_area(
                            "Edite a tradução",
                            value=item["texto_traduzido"],
                            height=100,
                            key=chave_edit,
                            label_visibility="collapsed"
                        )
                        # Atualiza o histórico se o usuário editou
                        if texto_editado != item["texto_traduzido"]:
                            item["texto_traduzido"] = texto_editado

                    col_b1, col_b2, col_b3 = st.columns(3)
                    with col_b1:
                        if st.button("🔊 Ouvir Tradução", use_container_width=True, key=f"stt_ouvir_{idx}"):
                            texto_ouvir = item["texto_traduzido"]
                            if not texto_ouvir.strip():
                                st.warning("⚠️ Nenhuma tradução para ouvir.")
                            else:
                                try:
                                    from gtts import gTTS
                                except ImportError:
                                    st.error("❌ A biblioteca `gTTS` não está instalada. Execute: `pip install gtts`")
                                    st.stop()
                                with st.spinner("Gerando áudio..."):
                                    cod_tts = item["cod_trad"]
                                    try:
                                        from gtts.lang import tts_langs
                                        suportados = tts_langs()
                                    except Exception:
                                        suportados = {}
                                    if not cod_tts or cod_tts not in suportados:
                                        cod_tts = "pt"
                                        st.info("ℹ️ Idioma não suportado para áudio. Usando Português.")
                                    try:
                                        tts = gTTS(text=texto_ouvir, lang=cod_tts, slow=False)
                                        mp3_buffer = io.BytesIO()
                                        tts.write_to_fp(mp3_buffer)
                                        mp3_buffer.seek(0)
                                        st.success("✅ Áudio gerado!")
                                        st.audio(mp3_buffer, format="audio/mp3")
                                        st.download_button(
                                            label="📥 Baixar Áudio",
                                            data=mp3_buffer.getvalue(),
                                            file_name=f"traducao_audio_{item['timestamp'].replace('/', '').replace(' ', '_').replace(':', '')}.mp3",
                                            mime="audio/mpeg",
                                            key=f"stt_down_audio_{idx}"
                                        )
                                    except Exception as e_gtts:
                                        st.error(f"❌ Erro ao gerar áudio com gTTS: {e_gtts}")
                    with col_b2:
                        if st.button("💾 Salvar Correção", use_container_width=True, key=f"stt_salvar_{idx}"):
                            item["texto_traduzido"] = texto_editado
                            st.success("✅ Correção salva!")
                    with col_b3:
                        texto_baixar = f"=== TEXTO ORIGINAL ===\n{item['texto_transcrito']}\n\n=== TRADUÇÃO ===\n{item['texto_traduzido']}"
                        st.download_button(
                            label="📥 Baixar Textos",
                            data=texto_baixar.encode("utf-8"),
                            file_name=f"traducao_{item['timestamp'].replace('/', '').replace(' ', '_').replace(':', '')}.txt",
                            mime="text/plain",
                            key=f"stt_down_txt_{idx}"
                        )

            if st.button("🗑️ Limpar Todo o Histórico STT", key="stt_limpar_hist"):
                st.session_state["stt_historico"] = []
                st.session_state["stt_texto_transcrito_atual"] = ""
                st.session_state["stt_cod_idioma_audio_atual"] = "pt"
                st.session_state["stt_ultimo_audio_bytes"] = None
                st.rerun()


# ====================== ABA 12: COMPRAS ======================
with aba12:
    render_compras()
